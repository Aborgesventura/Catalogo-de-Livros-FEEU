# -*- coding: utf-8 -*-
"""
==============================================================================
📚 CATÁLOGO DE LIVROS - FRATERNIDADE ECLÉTICA ESPIRITUALISTA UNIVERSAL
Versão: 4.14.0
- Linhas e colunas coloridas
- Tema claro/escuro
- Alteração de senha do admin
- Auto-redimensionamento
- Estante e Prateleira
- Importação CSV/Excel/Calc corrigida
==============================================================================
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog

import sqlite3
import csv
import os
import sys
import math
import calendar
import unicodedata
import hashlib
import hmac
import re
import tempfile
import subprocess
import webbrowser
import traceback
import fnmatch

from datetime import datetime

# ===================== DEPENDÊNCIAS PORTÁTEIS =====================
import os as _os_port
import sys as _sys_port

if getattr(_sys_port, 'frozen', False):
    _APP_DIR_PORT = _os_port.path.dirname(_sys_port.executable)
else:
    _APP_DIR_PORT = _os_port.path.dirname(_os_port.path.abspath(__file__))

_LIBS_DIR_PORT = _os_port.path.join(_APP_DIR_PORT, 'libs')

if _os_port.path.isdir(_LIBS_DIR_PORT) and _LIBS_DIR_PORT not in _sys_port.path:
    _sys_port.path.insert(0, _LIBS_DIR_PORT)
# ==================================================================
# ==============================================================================
# DEPENDÊNCIAS OPCIONAIS
# ==============================================================================
try:
    import openpyxl
    from openpyxl import Workbook
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

try:
    from odf.opendocument import OpenDocumentSpreadsheet, load as odf_load
    from odf.table import Table as OdfTable, TableRow as OdfTableRow, TableCell as OdfTableCell
    from odf.text import P as OdfP
    HAS_ODS = True
except ImportError:
    HAS_ODS = False

try:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table as PdfTable, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    HAS_PDF = True
except ImportError:
    HAS_PDF = False


# ==============================================================================
# TEMAS
# ==============================================================================
TEMAS = {
    'claro': {
        'bg_janela': '#f1f5f9',
        'bg_card': '#ffffff',
        'bg_rodape': '#e2e8f0',
        'bg_botao': '#e2e8f0',
        'bg_botao_hover': '#cbd5e1',
        'texto': '#0f172a',
        'texto_suave': '#64748b',
        'titulo': '#1e40af',
        'borda': '#cbd5e1',
        'borda_cell': '#e2e8f0',
        'campo_bg': '#ffffff',
        'campo_fg': '#0f172a',
        'scroll_bg': '#cbd5e1',

        'header_par': '#dbeafe',
        'header_impar': '#eff6ff',
        'header_texto': '#1e3a8a',
        'header_canvas_bg': '#ffffff',

        'celula_par': '#ffffff',
        'celula_impar': '#f1f5f9',
        'celula_par_alt': '#f8fafc',
        'celula_impar_alt': '#e9eef5',

        'selecao': '#bfdbfe',
        'selecao_texto': '#0f172a',

        'indisponivel_par': '#fee2e2',
        'indisponivel_impar': '#fecaca',
        'indisponivel_texto': '#991b1b',

        'tv_bg': '#ffffff',
        'tv_fg': '#0f172a',
        'tv_head_bg': '#2563eb',
        'tv_head_fg': '#ffffff',
    },
    'escuro': {
        'bg_janela': '#0f172a',
        'bg_card': '#1e293b',
        'bg_rodape': '#1e293b',
        'bg_botao': '#334155',
        'bg_botao_hover': '#475569',
        'texto': '#e2e8f0',
        'texto_suave': '#94a3b8',
        'titulo': '#93c5fd',
        'borda': '#475569',
        'borda_cell': '#334155',
        'campo_bg': '#1e293b',
        'campo_fg': '#e2e8f0',
        'scroll_bg': '#475569',

        'header_par': '#1e3a8a',
        'header_impar': '#172554',
        'header_texto': '#dbeafe',
        'header_canvas_bg': '#1e293b',

        'celula_par': '#1e293b',
        'celula_impar': '#24324a',
        'celula_par_alt': '#223049',
        'celula_impar_alt': '#2a3950',

        'selecao': '#2563eb',
        'selecao_texto': '#ffffff',

        'indisponivel_par': '#450a0a',
        'indisponivel_impar': '#5f1414',
        'indisponivel_texto': '#fca5a5',

        'tv_bg': '#1e293b',
        'tv_fg': '#e2e8f0',
        'tv_head_bg': '#1d4ed8',
        'tv_head_fg': '#ffffff',
    }
}

TEMA_ATUAL = TEMAS['claro']


# ==============================================================================
# TOOLTIP
# ==============================================================================
class ToolTip:
    def __init__(self, widget, text, delay_ms=800):
        self.widget = widget
        self.text = text
        self.delay_ms = delay_ms
        self.tipwindow = None
        self.after_id = None

        self.widget.bind('<Enter>', self.schedule_show, add='+')
        self.widget.bind('<Leave>', self.hide, add='+')
        self.widget.bind('<Motion>', self.reset_timer, add='+')

    def schedule_show(self, event=None):
        if self.after_id:
            self.widget.after_cancel(self.after_id)
        self.after_id = self.widget.after(self.delay_ms, self.show)

    def reset_timer(self, event=None):
        if self.after_id:
            self.widget.after_cancel(self.after_id)
        self.after_id = self.widget.after(self.delay_ms, self.show)

    def show(self, event=None):
        if self.tipwindow or not self.text:
            return

        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + self.widget.winfo_height() + 5

        self.tipwindow = tk.Toplevel(self.widget)
        self.tipwindow.wm_overrideredirect(True)
        self.tipwindow.wm_geometry(f"+{x}+{y}")
        self.tipwindow.attributes('-topmost', True)

        tk.Label(
            self.tipwindow,
            text=self.text,
            justify='left',
            background="#fef3c7",
            relief='solid',
            borderwidth=1,
            font=('DejaVu Sans', 9),
            foreground="#78350f",
            padx=8,
            pady=5
        ).pack()

        self.tipwindow.update_idletasks()

        if x + self.tipwindow.winfo_width() > self.widget.winfo_screenwidth():
            self.tipwindow.wm_geometry(
                f"+{self.widget.winfo_screenwidth() - self.tipwindow.winfo_width() - 10}+{y}"
            )

    def hide(self, event=None):
        if self.after_id:
            self.widget.after_cancel(self.after_id)
            self.after_id = None

        if self.tipwindow:
            self.tipwindow.destroy()
            self.tipwindow = None


# ==============================================================================
# DATECHOOSER
# ==============================================================================
class DateChooser(tk.Toplevel):
    def __init__(self, parent, target_entry):
        super().__init__(parent)
        self.target_entry = target_entry

        self.title("Selecionar Data")
        self.resizable(False, False)
        self.grab_set()
        self.transient(parent)

        sw, sh = parent.winfo_screenwidth(), parent.winfo_screenheight()
        self.geometry(f"280x340+{(sw - 280) // 2}+{(sh - 340) // 2}")

        self.now = datetime.now()
        self.year, self.month = self.now.year, self.now.month

        t = TEMA_ATUAL
        self.configure(bg=t['bg_janela'])
        self.bg = t['bg_card']
        self.fg_dia = t['texto']

        self._build_ui()
        self._render_calendar()

    def _build_ui(self):
        header = ttk.Frame(self)
        header.pack(fill='x', pady=(10, 5), padx=10)

        ttk.Button(header, text="◀", width=3, command=self._prev_month).pack(side='left')

        self.lbl_month = ttk.Label(header, text=" ", font=('DejaVu Sans', 10, 'bold'))
        self.lbl_month.pack(side='left', fill='x', expand=True, padx=10)

        ttk.Button(header, text="▶", width=3, command=self._next_month).pack(side='right')

        days_frame = ttk.Frame(self)
        days_frame.pack(fill='x', padx=10, pady=(5, 0))

        for dia in ['Seg', 'Ter', 'Qua', 'Qui', 'Sex', 'Sáb', 'Dom']:
            ttk.Label(days_frame, text=dia, width=4, anchor='center').grid(
                row=0,
                column=days_frame.grid_size()[0],
                padx=1
            )

        self.grid_frame = ttk.Frame(self)
        self.grid_frame.pack(fill='both', expand=True, padx=10, pady=5)

        self.btn_days = []

        for r in range(6):
            row_btns = []
            for c in range(7):
                btn = tk.Button(
                    self.grid_frame,
                    text=" ",
                    width=4,
                    relief='flat',
                    bd=1,
                    command=lambda r=r, c=c: self._pick_date(r, c),
                    font=('DejaVu Sans', 9)
                )
                btn.grid(row=r, column=c, padx=1, pady=1)
                row_btns.append(btn)
            self.btn_days.append(row_btns)

        ttk.Button(self, text="📅 Hoje", command=self._pick_today).pack(pady=10)

    def _render_calendar(self):
        meses = [
            'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
            'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
        ]

        self.lbl_month.config(text=f"{meses[self.month - 1]} {self.year}")

        for r in range(6):
            for c in range(7):
                self.btn_days[r][c].config(text=" ", state='disabled', bg=self.bg, fg=self.fg_dia)

        first_day, num_days = calendar.monthrange(self.year, self.month)

        for day in range(1, num_days + 1):
            r = (first_day + day - 1) // 7
            c = (first_day + day - 1) % 7

            if r < 6:
                btn = self.btn_days[r][c]
                btn.config(text=str(day), state='normal')

                if self.year == self.now.year and self.month == self.now.month and day == self.now.day:
                    btn.config(bg='#fef3c7', fg='#92400e', relief='raised')
                else:
                    btn.config(bg=self.bg, fg=self.fg_dia, relief='flat')

    def _prev_month(self):
        self.month -= 1
        if self.month < 1:
            self.month = 12
            self.year -= 1
        self._render_calendar()

    def _next_month(self):
        self.month += 1
        if self.month > 12:
            self.month = 1
            self.year += 1
        self._render_calendar()

    def _pick_date(self, r, c):
        day = int(self.btn_days[r][c].cget('text'))
        self.target_entry.delete(0, tk.END)
        self.target_entry.insert(0, f"{day:02d}/{self.month:02d}/{self.year}")
        self.destroy()

    def _pick_today(self):
        self.target_entry.delete(0, tk.END)
        self.target_entry.insert(0, self.now.strftime('%d/%m/%Y'))
        self.destroy()


# ==============================================================================
# TABELA CUSTOMIZADA COM CORES POR LINHA E COLUNA
# ==============================================================================
class TabelaCatalogo(ttk.Frame):
    def __init__(self, parent, columns, on_select=None, on_double=None, row_height=26):
        super().__init__(parent)

        self.columns = columns
        self.on_select = on_select
        self.on_double = on_double
        self.row_height = row_height

        self.data = []
        self.selected_id = None
        self.sort_col = None
        self.sort_reverse = False

        self.visible_cols = []
        self.col_widths = {}

        self.tema = dict(TEMA_ATUAL)

        self._after_resize = None
        self._after_redraw = None

        self.header = tk.Canvas(self, height=28, highlightthickness=0)
        self.body = tk.Canvas(self, highlightthickness=0)
        self.vbar = ttk.Scrollbar(self, orient='vertical', command=self._scroll_y)

        self.header.pack(side='top', fill='x')
        self.body.pack(side='left', fill='both', expand=True)
        self.vbar.pack(side='right', fill='y')

        self.body.configure(yscrollcommand=self._on_yscroll)

        self.header.bind('<Button-1>', self._on_header_click)
        self.body.bind('<Button-1>', self._on_click)
        self.body.bind('<Double-Button-1>', self._on_double)

        self.body.bind('<MouseWheel>', self._on_mousewheel)
        self.body.bind('<Button-4>', lambda e: self.body.yview_scroll(-1, 'units'))
        self.body.bind('<Button-5>', lambda e: self.body.yview_scroll(1, 'units'))

        self.bind('<Configure>', self._schedule_resize)

        self._update_visible_columns()
        self._draw_header()

    def set_theme(self, tema):
        self.tema.update(tema)

        try:
            self.header.configure(bg=self.tema.get('header_canvas_bg', self.tema['header_par']))
            self.body.configure(bg=self.tema.get('celula_par', 'white'))
        except Exception:
            pass

        self._draw_header()
        self._draw_body()

    def _schedule_resize(self, event=None):
        if not self.winfo_exists():
            return

        if self._after_resize:
            self.after_cancel(self._after_resize)

        self._after_resize = self.after(150, self._on_resize)

    def _on_resize(self):
        if not self.winfo_exists():
            return

        self._update_visible_columns()
        self._draw_header()
        self._draw_body()

    def _update_visible_columns(self):
        avail = self.body.winfo_width()

        if avail < 50:
            self.visible_cols = self.columns[:]
        else:
            essential = [c for c in self.columns if c.get('priority', 1) == 0]
            optional = [c for c in self.columns if c.get('priority', 1) != 0]
            optional.sort(key=lambda c: (c.get('priority', 1), self.columns.index(c)))

            chosen = essential[:]
            total = sum(c['base'] for c in chosen)

            for c in optional:
                if total + c['base'] <= avail:
                    chosen.append(c)
                    total += c['base']

            self.visible_cols = [c for c in self.columns if c in chosen]

        self.col_widths = {c['key']: c['base'] for c in self.visible_cols}

        if avail > 50:
            total = sum(self.col_widths.values())

            if avail > total:
                extra = avail - total

                stretch_cols = [c for c in self.visible_cols if c.get('stretch')]

                if not stretch_cols:
                    stretch_cols = [c for c in self.visible_cols if c['key'] in ('titulo', 'autor')]

                if not stretch_cols:
                    stretch_cols = self.visible_cols[:]

                weights = [c['base'] for c in stretch_cols]
                soma = sum(weights) or 1
                used = 0

                for i, c in enumerate(stretch_cols):
                    add = int(extra * weights[i] / soma)
                    self.col_widths[c['key']] += add
                    used += add

                resto = extra - used

                if resto > 0 and stretch_cols:
                    self.col_widths[stretch_cols[0]['key']] += resto

    def _draw_header(self):
        self.header.delete('all')

        x = 0
        t = self.tema

        for idx, c in enumerate(self.visible_cols):
            w = self.col_widths.get(c['key'], c['base'])
            bg = t['header_par'] if idx % 2 == 0 else t['header_impar']

            self.header.create_rectangle(
                x, 0, x + w, 28,
                fill=bg,
                outline=t['borda']
            )

            arrow = ''
            if self.sort_col == c['key']:
                arrow = ' ▼' if self.sort_reverse else ' ▲'

            self.header.create_text(
                x + w // 2,
                14,
                text=c['text'] + arrow,
                anchor='center',
                font=('DejaVu Sans', 9, 'bold'),
                fill=t['header_texto']
            )

            x += w

        self.header.configure(scrollregion=(0, 0, x, 28))

    def _schedule_redraw(self):
        if not self.winfo_exists():
            return

        if self._after_redraw:
            self.after_cancel(self._after_redraw)

        self._after_redraw = self.after(10, self._draw_body)

    def _draw_body(self):
        if not self.winfo_exists():
            return

        self.body.delete('all')

        if not self.visible_cols:
            return

        t = self.tema

        total_w = sum(self.col_widths.values())
        total_h = len(self.data) * self.row_height

        cw = max(self.body.winfo_width(), 1)
        ch = max(self.body.winfo_height(), 1)

        self.body.configure(scrollregion=(0, 0, max(total_w, cw), max(total_h, ch)))

        if not self.data:
            self.body.create_text(
                cw // 2,
                ch // 2,
                text="Nenhum registro.",
                fill=t['texto_suave'],
                font=('DejaVu Sans', 10)
            )
            return

        y0 = self.body.canvasy(0)
        first = max(0, int(y0 // self.row_height) - 1)
        visible_count = int(ch // self.row_height) + 3
        last = min(len(self.data), first + visible_count)

        for i in range(first, last):
            row = self.data[i]
            y = i * self.row_height

            selected = row.get('id') == self.selected_id
            indisponivel = str(row.get('disponibilidade', '')).strip() == 'Não'

            x = 0

            for idx, c in enumerate(self.visible_cols):
                w = self.col_widths.get(c['key'], c['base'])

                if selected:
                    bg = t['selecao']
                elif indisponivel:
                    bg = t['indisponivel_par'] if i % 2 == 0 else t['indisponivel_impar']
                else:
                    if i % 2 == 0:
                        bg = t['celula_par'] if idx % 2 == 0 else t['celula_impar']
                    else:
                        bg = t['celula_par_alt'] if idx % 2 == 0 else t['celula_impar_alt']

                self.body.create_rectangle(
                    x, y, x + w, y + self.row_height,
                    fill=bg,
                    outline=t['borda_cell']
                )

                val = row.get(c['key'], '')
                txt = '' if val is None else str(val)

                max_chars = max(3, w // 8)
                if len(txt) > max_chars:
                    txt = txt[:max_chars - 1] + '…'

                anchor = c.get('anchor', 'w')
                tx = x + w // 2 if anchor == 'center' else x + 4

                if selected:
                    cor = t['selecao_texto']
                elif indisponivel:
                    cor = t.get('indisponivel_texto', t['texto'])
                else:
                    cor = t['texto']

                self.body.create_text(
                    tx,
                    y + self.row_height // 2,
                    text=txt,
                    anchor=anchor,
                    fill=cor,
                    font=('DejaVu Sans', 9)
                )

                x += w

    def _on_yscroll(self, first, last):
        self.vbar.set(first, last)
        self._schedule_redraw()

    def _scroll_y(self, *args):
        self.body.yview(*args)

    def _on_mousewheel(self, event):
        self.body.yview_scroll(int(-1 * (event.delta / 120)), 'units')

    def _row_index(self, event):
        y = self.body.canvasy(event.y)
        idx = int(y // self.row_height)

        if 0 <= idx < len(self.data):
            return idx

        return None

    def _on_click(self, event):
        idx = self._row_index(event)

        if idx is None:
            return

        self.selected_id = self.data[idx].get('id')
        self._draw_body()

        if self.on_select:
            self.on_select(self.selected_id)

    def _on_double(self, event):
        idx = self._row_index(event)

        if idx is None:
            return

        self.selected_id = self.data[idx].get('id')
        self._draw_body()

        if self.on_double:
            self.on_double(self.selected_id)

    def _on_header_click(self, event):
        x = 0

        for c in self.visible_cols:
            w = self.col_widths.get(c['key'], c['base'])

            if x <= event.x <= x + w:
                key = c['key']

                if self.sort_col == key:
                    self.sort_reverse = not self.sort_reverse
                else:
                    self.sort_col = key
                    self.sort_reverse = False

                self._sort_data()
                self._draw_header()
                self._draw_body()
                return

            x += w

    def _sort_data(self):
        if not self.sort_col:
            return

        col = self.sort_col

        def chave(r):
            v = r.get(col, '')

            if col in ('id', 'quantidade'):
                try:
                    return (0, int(str(v).strip() or 0))
                except Exception:
                    return (0, 0)

            if col in ('entrada', 'saida'):
                s = str(v or '').strip()

                if not s:
                    return (1, '')

                if len(s) >= 10 and s[4] == '-':
                    return (1, s)

                p = s.split('/')
                if len(p) == 3:
                    return (1, f"{p[2]:>4}-{p[1]:>2}-{p[0]:>2}")

                return (1, s)

            t = str(v or '').lower()
            t = unicodedata.normalize('NFD', t)
            t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')

            return (2, t)

        self.data.sort(key=chave, reverse=self.sort_reverse)

    def set_data(self, data):
        self.data = list(data)
        self.selected_id = None

        self._sort_data()

        try:
            self.body.yview_moveto(0)
        except Exception:
            pass

        self._update_visible_columns()
        self._draw_header()
        self._draw_body()


# ==============================================================================
# APLICAÇÃO PRINCIPAL
# ==============================================================================

# ---- caminho gravável do app (corrige o .exe --onefile) ----
def _v12_app_dir():
    try:
        if getattr(sys, 'frozen', False):                 # rodando como .exe
            base = os.path.dirname(os.path.abspath(sys.executable))
        else:                                             # rodando como .py
            base = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        base = os.getcwd()

    def _ok(p):
        try:
            os.makedirs(p, exist_ok=True)
            t = os.path.join(p, '.teste_feeu')
            with open(t, 'w', encoding='utf-8') as f:
                f.write('1')
            os.remove(t)
            return True
        except Exception:
            return False

    escolha = base if _ok(base) else (
        (os.path.join(os.path.expanduser('~'), 'Documents', 'FEEU_Catalogo') if os.name == 'nt'
         else os.path.join(os.path.expanduser('~'), '.feeucatalogo'))
    )
    if escolha != base and not _ok(escolha):
        escolha = base

    # grava um arquivo dizendo ONDE o banco está (ajuda a achar)
    try:
        with open(os.path.join(escolha, 'caminho_banco.txt'), 'w', encoding='utf-8') as f:
            f.write(os.path.join(escolha, 'catalogo_feeu.db') + '\n')
    except Exception:
        pass
    return escolha
class LivroCatalogApp:
    COLUNAS = [
        {'key': 'id', 'text': 'ID', 'base': 45, 'anchor': 'center', 'priority': 0, 'stretch': False},
        {'key': 'titulo', 'text': 'Título', 'base': 180, 'anchor': 'w', 'priority': 1, 'stretch': True},
        {'key': 'autor', 'text': 'Autor', 'base': 140, 'anchor': 'w', 'priority': 1, 'stretch': True},
        {'key': 'estante', 'text': 'Estante', 'base': 80, 'anchor': 'center', 'priority': 2, 'stretch': False},
        {'key': 'prateleira', 'text': 'Prateleira', 'base': 90, 'anchor': 'center', 'priority': 2, 'stretch': False},
        {'key': 'editora', 'text': 'Editora', 'base': 120, 'anchor': 'w', 'priority': 3, 'stretch': True},
        {'key': 'assunto', 'text': 'Assunto', 'base': 120, 'anchor': 'w', 'priority': 3, 'stretch': True},
        {'key': 'bibliotecario', 'text': 'Bibliotecário', 'base': 120, 'anchor': 'w', 'priority': 3, 'stretch': True},
        {'key': 'quantidade', 'text': 'Qtd.', 'base': 55, 'anchor': 'center', 'priority': 0, 'stretch': False},
        {'key': 'disponibilidade', 'text': 'Disp.', 'base': 65, 'anchor': 'center', 'priority': 0, 'stretch': False},
        {'key': 'emprestado_a', 'text': 'Emprestado a', 'base': 160, 'anchor': 'w', 'priority': 2, 'stretch': True},
        {'key': 'entrada', 'text': 'Entrada', 'base': 85, 'anchor': 'center', 'priority': 2, 'stretch': False},
        {'key': 'saida', 'text': 'Saída', 'base': 85, 'anchor': 'center', 'priority': 2, 'stretch': False},
    ]

    def __init__(self, root):
        self.root = root
        self.root.title("Catálogo de Livros - Fraternidade Eclética Espiritualista Universal")

        sw, sh = root.winfo_screenwidth(), root.winfo_screenheight()
        root.geometry(f"{int(sw * 0.92)}x{int(sh * 0.88)}+{(sw - int(sw * 0.92)) // 2}+{(sh - int(sh * 0.88)) // 2}")
        root.minsize(1000, 500)
        root.resizable(True, True)
        root.protocol("WM_DELETE_WINDOW", self.confirmar_saida)

        self.diretorio_app = os.path.dirname(os.path.abspath(__file__))
        self.db_file = os.path.join(self.diretorio_app, "catalogo_feeu.db")
        self.config_file = os.path.join(self.diretorio_app, "config.json")

        self.livros = []
        self.livro_selecionado_id = None
        self.busca_ativa = False
        self.ultima_busca = []

        self.cores = {
            'verde': '#006400',
            'dourado': '#DAA520',
            'azul': '#000080',
            'vermelho': '#8B0000',
            'ponto': '#FF0000'
        }

        self.usuario_id = None
        self.usuario_nome = None
        self.permissoes = {}

        self.clock_job = None

        self.config = self._carregar_config()
        self.tema_cores = TEMAS.get(self.config.get('tema', 'claro'), TEMAS['claro'])

        self.inicializar_banco()

        self.root.withdraw()

        if not self.autenticar():
            self.root.destroy()
            return

        self.root.deiconify()

        self.setup_styles()
        self.create_widgets()
        self.aplicar_tema()
        self.aplicar_permissoes()
        self.carregar_dados()
        self.atualizar_comboboxes()
        self.atualizar_relogio()

        self.root.bind_all('<Control-a>', self._atalho_adicionar)
        self.root.bind_all('<Control-s>', lambda e: self.salvar())
        self.root.bind_all('<Control-e>', lambda e: self.excluir())
        self.root.bind_all('<Control-l>', lambda e: self.limpar_form())
        self.root.bind('<F1>', lambda e: self.mostrar_manual())

    # ==========================================================================
    # CONFIGURAÇÃO
    # ==========================================================================
    def _carregar_config(self):
        padrao = {
            'tema': 'claro'
        }

        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    cfg = json.load(f)
                padrao.update(cfg)
        except Exception:
            pass

        return padrao

    def _salvar_config(self):
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, indent=2, ensure_ascii=False)
        except Exception:
            pass

    # ==========================================================================
    # SEGURANÇA
    # ==========================================================================
    def _hash_senha(self, senha):
        salt = os.urandom(16)
        iteracoes = 260000
        dk = hashlib.pbkdf2_hmac("sha256", senha.encode("utf-8"), salt, iteracoes)
        return f"pbkdf2_sha256${iteracoes}${salt.hex()}${dk.hex()}"

    def _verificar_senha(self, senha, armazenada):
        if not armazenada:
            return False

        armazenada = str(armazenada).strip()

        if armazenada.startswith("pbkdf2_sha256$"):
            try:
                _, iteracoes, salt, hash_hex = armazenada.split("$")
                dk = hashlib.pbkdf2_hmac(
                    "sha256",
                    senha.encode("utf-8"),
                    bytes.fromhex(salt),
                    int(iteracoes)
                )
                return hmac.compare_digest(dk.hex(), hash_hex)
            except Exception:
                return False

        if len(armazenada) == 64:
            return hmac.compare_digest(
                hashlib.sha256(senha.encode("utf-8")).hexdigest(),
                armazenada
            )

        return False

    def _checar_permissao(self, chave, mensagem):
        perms = getattr(self, 'permissoes', None)

        if perms is None:
            return True

        if perms.get('admin'):
            return True

        if not perms.get(chave, False):
            messagebox.showwarning("🔒 Permissão insuficiente", mensagem)
            return False

        return True

    # ==========================================================================
    # SUPORTE
    # ==========================================================================
    def _atalho_adicionar(self, event=None):
        w = event.widget if event else None

        if w is not None:
            try:
                classe = w.winfo_class()
            except Exception:
                classe = ""

            if isinstance(w, (tk.Entry, ttk.Entry)) or classe in ("Entry", "TEntry", "Combobox", "TCombobox"):
                return

        self.adicionar_livro()
        return "break"

    def _atualizar_estado_btn_busca(self):
        try:
            if not hasattr(self, 'btn_imp_busca'):
                return

            perms = getattr(self, 'permissoes', {}) or {}
            pode = bool(perms.get('admin') or perms.get('relatorio'))
            ativo = bool(getattr(self, 'busca_ativa', False) and getattr(self, 'ultima_busca', []))

            self.btn_imp_busca.config(state='normal' if (pode and ativo) else 'disabled')
        except Exception:
            pass

    def _recarregar_manter_busca(self):
        try:
            termo = self.busca_var.get().strip() if hasattr(self, 'busca_var') else ''

            self.carregar_dados()

            if getattr(self, 'busca_ativa', False) and termo:
                self.buscar()
            else:
                self._atualizar_estado_btn_busca()
        except Exception:
            try:
                self.carregar_dados()
            except Exception:
                messagebox.showerror("❌", traceback.format_exc())

    def _validar_quantidade(self, valor):
        s = str(valor or '').strip().replace(',', '.')

        if not s:
            return 0, True

        try:
            f = float(s)

            if f < 0:
                return 0, False

            if f.is_integer():
                return int(f), True

            return 0, False
        except Exception:
            return 0, False

    def _colunas_livros_existentes(self, cursor):
        cursor.execute("PRAGMA table_info(livros)")
        return {r[1] for r in cursor.fetchall()}

    def _data_br_para_iso(self, valor, manter_original=False):
        try:
            if hasattr(valor, 'strftime'):
                return valor.strftime('%Y-%m-%d')

            s = str(valor or '').strip()

            if not s:
                return None

            if len(s) >= 10 and s[4] == '-' and s[7] == '-':
                datetime.strptime(s[:10], '%Y-%m-%d')
                return s[:10]

            for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d'):
                try:
                    return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
                except Exception:
                    continue

            return s if manter_original else None

        except Exception:
            return str(valor or '').strip() if manter_original else None

    def _data_iso_para_br(self, valor):
        try:
            s = str(valor or '').strip()

            if not s:
                return ''

            if len(s) >= 10 and s[4] == '-' and s[7] == '-':
                return datetime.strptime(s[:10], '%Y-%m-%d').strftime('%d/%m/%Y')

            return s

        except Exception:
            return str(valor or '').strip()

    def _normalizar_texto(self, texto):
        try:
            if texto is None:
                return ''

            t = str(texto).strip().lower()
            t = unicodedata.normalize('NFD', t)
            return ''.join(c for c in t if unicodedata.category(c) != 'Mn')

        except Exception:
            return str(texto or '').strip().lower()

    def _wildcard_match(self, termo, texto):
        termo_n = self._normalizar_texto(termo)
        texto_n = self._normalizar_texto(texto)

        if not termo_n:
            return True

        if '*' in termo_n or '?' in termo_n:
            try:
                return fnmatch.fnmatch(texto_n, termo_n)
            except Exception:
                return termo_n.replace('*', '') in texto_n

        return termo_n in texto_n

    def _registrar_historico_livro(self, dados, cursor=None):
        try:
            campos = [
                'autor',
                'editora',
                'assunto',
                'bibliotecario',
                'emprestado_a',
                'estante',
                'prateleira'
            ]

            def registrar(cur):
                for campo in campos:
                    valor = str(dados.get(campo, '') or '').strip()

                    if not valor:
                        continue

                    if campo == 'emprestado_a':
                        for nome in valor.split(';'):
                            nome = nome.strip()
                            if nome:
                                cur.execute(
                                    "INSERT OR IGNORE INTO historico_campos (campo, valor) VALUES (?, ?)",
                                    (campo, nome)
                                )
                    else:
                        cur.execute(
                            "INSERT OR IGNORE INTO historico_campos (campo, valor) VALUES (?, ?)",
                            (campo, valor)
                        )

            if cursor is None:
                conn = sqlite3.connect(self.db_file)

                try:
                    registrar(conn.cursor())
                    conn.commit()
                finally:
                    conn.close()
            else:
                registrar(cursor)

        except Exception:
            pass

    def _checar_duplicata(self, titulo, autor, editora='', assunto='', estante='', prateleira='', cursor=None):
        conn = None
        fechar = False

        if cursor is None:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()
            fechar = True

        try:
            cols = self._colunas_livros_existentes(cursor)

            sql = """
                SELECT id, quantidade
                FROM livros
                WHERE LOWER(titulo)=?
                  AND LOWER(autor)=?
            """

            params = [titulo.lower().strip(), autor.lower().strip()]

            if 'editora' in cols and editora.strip():
                sql += " AND IFNULL(LOWER(editora),'')=?"
                params.append(editora.lower().strip())

            if 'assunto' in cols and assunto.strip():
                sql += " AND IFNULL(LOWER(assunto),'')=?"
                params.append(assunto.lower().strip())

            if 'estante' in cols and estante.strip():
                sql += " AND IFNULL(LOWER(estante),'')=?"
                params.append(estante.lower().strip())

            if 'prateleira' in cols and prateleira.strip():
                sql += " AND IFNULL(LOWER(prateleira),'')=?"
                params.append(prateleira.lower().strip())

            cursor.execute(sql, tuple(params))
            return cursor.fetchone()

        finally:
            if fechar and conn is not None:
                conn.close()

    # ==========================================================================
    # BANCO
    # ==========================================================================
    def inicializar_banco(self):
        try:
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS usuarios (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT UNIQUE NOT NULL,
                    senha TEXT NOT NULL,
                    perm_inserir BOOLEAN DEFAULT 0,
                    perm_editar BOOLEAN DEFAULT 0,
                    perm_apagar BOOLEAN DEFAULT 0,
                    perm_exportar BOOLEAN DEFAULT 0,
                    perm_importar BOOLEAN DEFAULT 0,
                    perm_relatorio BOOLEAN DEFAULT 0,
                    perm_admin BOOLEAN DEFAULT 0,
                    perm_emprestados BOOLEAN DEFAULT 0
                )
            ''')

            try:
                cursor.execute("ALTER TABLE usuarios ADD COLUMN perm_emprestados BOOLEAN DEFAULT 0")
                conn.commit()
            except Exception:
                pass

            conn.commit()

            cursor.execute("SELECT id FROM usuarios WHERE nome='admin'")
            if not cursor.fetchone():
                cursor.execute(
                    """
                    INSERT INTO usuarios
                    (nome, senha, perm_inserir, perm_editar, perm_apagar, perm_exportar,
                     perm_importar, perm_relatorio, perm_admin, perm_emprestados)
                    VALUES (?, ?, 1, 1, 1, 1, 1, 1, 1, 1)
                    """,
                    ('admin', self._hash_senha('admin'))
                )
                conn.commit()

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS livros (
                    id INTEGER PRIMARY KEY,
                    titulo TEXT NOT NULL,
                    autor TEXT NOT NULL,
                    estante TEXT DEFAULT '',
                    prateleira TEXT DEFAULT '',
                    editora TEXT,
                    assunto TEXT,
                    bibliotecario TEXT,
                    quantidade INTEGER,
                    disponibilidade TEXT DEFAULT 'Sim',
                    emprestado_a TEXT DEFAULT '',
                    entrada TEXT,
                    saida TEXT
                )
            ''')

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS historico_campos (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    campo TEXT NOT NULL,
                    valor TEXT NOT NULL,
                    UNIQUE(campo, valor)
                )
            ''')

            for col, tipo in [
                ('estante', "TEXT DEFAULT ''"),
                ('prateleira', "TEXT DEFAULT ''"),
                ('disponibilidade', "TEXT DEFAULT 'Sim'"),
                ('emprestado_a', "TEXT DEFAULT ''"),
                ('saida', "TEXT")
            ]:
                try:
                    cursor.execute(f"ALTER TABLE livros ADD COLUMN {col} {tipo}")
                    conn.commit()
                except Exception:
                    pass

            conn.close()

        except Exception as e:
            messagebox.showerror("Erro Banco", str(e))

    # ==========================================================================
    # LOGIN
    # ==========================================================================
    def centralizar_janela(self, win):
        try:
            win.update_idletasks()
            w, h = win.winfo_width(), win.winfo_height()
            sw = self.root.winfo_screenwidth()
            sh = self.root.winfo_screenheight()
            x = (sw - w) // 2
            y = (sh - h) // 2

            if y < 0:
                y = 0

            win.geometry(f"+{x}+{y}")
        except Exception:
            pass

    def autenticar(self):
        last_user_file = os.path.join(self.diretorio_app, "last_user.conf")
        last_user = ""

        if os.path.exists(last_user_file):
            try:
                with open(last_user_file, "r", encoding="utf-8") as f:
                    last_user = f.read().strip()
            except Exception:
                last_user = ""

        login_win = tk.Toplevel(self.root)
        login_win.title("🔐 Login do Sistema")
        login_win.geometry("520x480")
        login_win.resizable(False, False)
        login_win.grab_set()
        login_win.attributes("-topmost", True)
        self.centralizar_janela(login_win)

        bg_color = self.tema_cores['bg_janela']

        main_frame = tk.Frame(login_win, bg=bg_color)
        main_frame.pack(fill="both", expand=True)

        welcome_frame = tk.Frame(main_frame, bg=bg_color)
        welcome_frame.pack(fill="x", pady=(15, 10))

        sym_canvas = tk.Canvas(
            welcome_frame,
            width=60,
            height=60,
            bg=self.tema_cores['bg_card'],
            highlightthickness=1,
            relief="solid"
        )
        sym_canvas.pack(side="left", padx=(30, 10))

        cx, cy = 30, 30

        sym_canvas.create_oval(4, 4, 56, 56, outline=self.cores['verde'], width=3)
        sym_canvas.create_oval(12, 12, 48, 48, outline=self.cores['dourado'], width=2)

        r_tri = 18
        p1 = (cx, cy - r_tri)
        p2 = (cx - r_tri * math.cos(math.radians(30)), cy + r_tri * math.sin(math.radians(30)))
        p3 = (cx + r_tri * math.cos(math.radians(30)), cy + r_tri * math.sin(math.radians(30)))

        sym_canvas.create_polygon(p1, p2, p3, outline=self.cores['azul'], width=2, fill='')
        sym_canvas.create_oval(cx - 2, cy - 2, cx + 2, cy + 2, fill=self.cores['ponto'])

        tk.Label(
            welcome_frame,
            text="📚 Catálogo de Livros FEEU",
            font=('DejaVu Sans', 16, 'bold'),
            foreground=self.tema_cores['titulo'],
            background=bg_color,
            wraplength=480
        ).pack(side='bottom', pady=(5, 10))

        tk.Label(
            welcome_frame,
            text="Bem-vindo ao Sistema",
            font=('DejaVu Sans', 11),
            foreground=self.tema_cores['texto_suave'],
            background=bg_color
        ).pack(side='bottom')

        form_frame = tk.Frame(main_frame, bg=self.tema_cores['bg_card'], relief="solid", bd=1)
        form_frame.pack(fill="x", padx=20, pady=15)

        inp_frame = ttk.Frame(form_frame, padding=15)
        inp_frame.pack(fill='x')

        ttk.Label(inp_frame, text="Usuário: ").grid(row=0, column=0, sticky='w', pady=5)

        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()
        cursor.execute("SELECT nome FROM usuarios ORDER BY nome")
        users = [r[0] for r in cursor.fetchall()]
        conn.close()

        user_var = tk.StringVar()
        user_combo = ttk.Combobox(inp_frame, textvariable=user_var, values=users, state='readonly')
        user_combo.grid(row=0, column=1, padx=5, pady=5, sticky='we')
        user_combo.set(last_user if last_user in users else (users[0] if users else ""))

        ttk.Label(inp_frame, text="Senha: ").grid(row=1, column=0, sticky='w', pady=5)

        pass_var = tk.StringVar()
        pass_frame = ttk.Frame(inp_frame)
        pass_frame.grid(row=1, column=1, padx=5, pady=5, sticky='we')

        pass_ent = ttk.Entry(pass_frame, textvariable=pass_var, show='*')
        pass_ent.pack(side='left', fill='x', expand=True)

        senha_visivel = False

        def toggle_senha():
            nonlocal senha_visivel
            senha_visivel = not senha_visivel
            pass_ent.config(show='' if senha_visivel else '*')
            btn_olho.config(text='👁️' if senha_visivel else '🙈')

        btn_olho = ttk.Button(pass_frame, text='🙈', command=toggle_senha, width=3)
        btn_olho.pack(side='right')

        login_result = {"ok": False, "id": None, "nome": None, "perms": {}}

        def do_login():
            username = user_var.get().strip()
            password = pass_var.get().strip()

            if not username:
                messagebox.showwarning("⚠️", "Selecione um usuário.", parent=login_win)
                return

            if not password:
                messagebox.showwarning("⚠️", "Digite a senha.", parent=login_win)
                return

            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            try:
                cursor.execute("""
                    SELECT id, nome, senha, perm_inserir, perm_editar, perm_apagar, perm_exportar,
                           perm_importar, perm_relatorio, perm_admin, perm_emprestados
                    FROM usuarios
                    WHERE nome=?
                """, (username,))
            except sqlite3.OperationalError:
                cursor.execute("""
                    SELECT id, nome, senha, perm_inserir, perm_editar, perm_apagar, perm_exportar,
                           perm_importar, perm_relatorio, perm_admin, 0
                    FROM usuarios
                    WHERE nome=?
                """, (username,))

            row = cursor.fetchone()

            if row and self._verificar_senha(password, row[2]):
                if not str(row[2]).startswith('pbkdf2_sha256$'):
                    cursor.execute(
                        "UPDATE usuarios SET senha=? WHERE id=?",
                        (self._hash_senha(password), row[0])
                    )
                    conn.commit()

                try:
                    with open(last_user_file, "w", encoding="utf-8") as f:
                        f.write(username)
                except Exception:
                    pass

                perms = {
                    'inserir': bool(row[3]),
                    'editar': bool(row[4]),
                    'apagar': bool(row[5]),
                    'exportar': bool(row[6]),
                    'importar': bool(row[7]),
                    'relatorio': bool(row[8]),
                    'admin': bool(row[9]),
                    'emprestados': bool(row[10]) if len(row) > 10 else False
                }

                login_result.update({"ok": True, "id": row[0], "nome": row[1], "perms": perms})
                conn.close()
                login_win.destroy()
            else:
                conn.close()
                messagebox.showerror("❌", "Usuário ou senha incorretos.", parent=login_win)
                pass_ent.delete(0, tk.END)
                pass_ent.focus_set()

        def do_cancel():
            login_win.destroy()

        btn_frame = ttk.Frame(login_win, padding=10)
        btn_frame.pack(fill='x')

        ttk.Button(btn_frame, text="Entrar", command=do_login, style='Import.TButton').pack(side='left', expand=True, padx=5)
        ttk.Button(btn_frame, text="Cancelar", command=do_cancel).pack(side='left', expand=True, padx=5)

        login_win.bind("<Return>", lambda e: do_login())

        pass_ent.focus_set()
        login_win.after(150, lambda: pass_ent.focus_force())

        login_win.lift()
        login_win.focus_force()

        self.root.wait_window(login_win)

        if login_result["ok"]:
            self.usuario_id = login_result["id"]
            self.usuario_nome = login_result["nome"]
            self.permissoes = login_result["perms"]
            return True

        return False

    def voltar_para_login(self):
        if messagebox.askyesno("🔙 Voltar ao Login", "Deseja retornar à tela de login?\nO sistema será bloqueado até nova autenticação."):
            self.root.withdraw()

            if self.autenticar():
                self.aplicar_permissoes()
                self.carregar_dados()
                self.root.deiconify()
                self.root.lift()
            else:
                self.root.destroy()

    # ==========================================================================
    # ESTILOS / TEMA
    # ==========================================================================
    def setup_styles(self):
        style = ttk.Style()
        style.theme_use('clam')

        style.configure('Excel.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#217346', foreground='white')
        style.configure('Calc.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#007439', foreground='white')
        style.configure('PDF.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#b30b00', foreground='white')
        style.configure('Print.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#0066cc', foreground='white')
        style.configure('Import.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#059669', foreground='white')
        style.configure('Adicionar.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#16a34a', foreground='white')
        style.configure('Salvar.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#2563eb', foreground='white')
        style.configure('Excluir.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#dc2626', foreground='white')
        style.configure('Limpar.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#64748b', foreground='white')
        style.configure('Sair.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(12, 5), background='#7f1d1d', foreground='white')
        style.configure('Voltar.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(12, 5), background='#f59e0b', foreground='white')
        style.configure('Emprestar.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#ea580c', foreground='white')
        style.configure('Devolver.TButton', font=('DejaVu Sans', 9, 'bold'), padding=(10, 5), background='#059669', foreground='white')

    def aplicar_tema(self):
        global TEMA_ATUAL

        nome = self.config.get('tema', 'claro')
        t = TEMAS.get(nome, TEMAS['claro'])

        self.tema_cores = t
        TEMA_ATUAL = t

        s = ttk.Style()

        if s.theme_use() != 'clam':
            s.theme_use('clam')

        s.configure('TFrame', background=t['bg_janela'])
        s.configure('TLabel', background=t['bg_janela'], foreground=t['texto'], font=('DejaVu Sans', 10))
        s.configure('Header.TLabel', background=t['bg_janela'], foreground=t['titulo'], font=('DejaVu Sans', 12, 'bold'))
        s.configure('Status.TLabel', background=t['bg_janela'], foreground=t['texto_suave'], font=('DejaVu Sans', 9, 'italic'))
        s.configure('Footer.TLabel', background=t['bg_rodape'], foreground=t['texto_suave'], font=('DejaVu Sans', 9, 'italic'))

        s.configure('TLabelframe', background=t['bg_janela'], bordercolor=t['borda'])
        s.configure('TLabelframe.Label', background=t['bg_janela'], foreground=t['titulo'], font=('DejaVu Sans', 10, 'bold'))

        s.configure('TEntry', fieldbackground=t['campo_bg'], foreground=t['campo_fg'], insertcolor=t['campo_fg'], bordercolor=t['borda'])
        s.configure('TCombobox', fieldbackground=t['campo_bg'], foreground=t['campo_fg'], arrowcolor=t['texto'], bordercolor=t['borda'])
        s.map('TCombobox', fieldbackground=[('readonly', t['campo_bg'])])

        s.configure('Treeview', background=t['tv_bg'], fieldbackground=t['tv_bg'], foreground=t['tv_fg'], rowheight=28)
        s.configure('Treeview.Heading', background=t['tv_head_bg'], foreground=t['tv_head_fg'], font=('DejaVu Sans', 10, 'bold'))
        s.map('Treeview', background=[('selected', t['selecao'])], foreground=[('selected', t['selecao_texto'])])

        s.configure('TScrollbar', background=t['scroll_bg'], troughcolor=t['bg_janela'], bordercolor=t['bg_janela'], arrowcolor=t['texto'])
        s.configure('TPanedwindow', background=t['bg_janela'])
        s.configure('TCheckbutton', background=t['bg_janela'], foreground=t['texto'])
        s.configure('TRadiobutton', background=t['bg_janela'], foreground=t['texto'])

        s.configure('TButton', background=t['bg_botao'], foreground=t['texto'], bordercolor=t['borda'])
        s.map('TButton', background=[('active', t['bg_botao_hover'])], foreground=[('active', t['texto'])])

        try:
            self.root.configure(bg=t['bg_janela'])

            if hasattr(self, 'btn_tema'):
                self.btn_tema.config(
                    text='☀️ Tema Claro' if nome == 'escuro' else '🌙 Tema Escuro',
                    bg=t['bg_botao'],
                    fg=t['texto']
                )

            if hasattr(self, 'relogio_frame'):
                for filho in self.relogio_frame.winfo_children():
                    try:
                        filho.configure(bg=t['bg_janela'])
                    except Exception:
                        pass

            if hasattr(self, 'tree'):
                self.tree.set_theme(t)

            if hasattr(self, 'resultados_tree'):
                self.resultados_tree.set_theme(t)

        except Exception:
            pass

    def alternar_tema(self):
        atual = self.config.get('tema', 'claro')
        self.config['tema'] = 'escuro' if atual == 'claro' else 'claro'
        self._salvar_config()
        self.aplicar_tema()

    # ==========================================================================
    # INTERFACE
    # ==========================================================================
    def _draw_symbol(self):
        canvas = self.symbol_canvas
        w, h = 70, 70
        cx, cy = w // 2, h // 2

        canvas.create_oval(4, 4, w - 4, h - 4, outline=self.cores['verde'], width=4, tag='symbol')
        canvas.create_oval(12, 12, w - 12, h - 12, outline=self.cores['dourado'], width=3, tag='symbol')

        r_tri = 22
        p1 = (cx, cy - r_tri)
        p2 = (cx - r_tri * math.cos(math.radians(30)), cy + r_tri * math.sin(math.radians(30)))
        p3 = (cx + r_tri * math.cos(math.radians(30)), cy + r_tri * math.sin(math.radians(30)))

        canvas.create_polygon(p1, p2, p3, outline=self.cores['azul'], width=2.5, fill='', tag='symbol')
        canvas.create_oval(cx - 3.5, cy - 3.5, cx + 3.5, cy + 3.5, fill=self.cores['ponto'], outline='', tag='symbol')

    def create_widgets(self):
        footer = ttk.Frame(self.root)
        footer.pack(side='bottom', fill='x')

        ttk.Separator(footer, orient='horizontal').pack(fill='x')
        ttk.Label(
            footer,
            text="Desenvolvido pelos Irmãos Rodolpho e Alexandre",
            style='Footer.TLabel',
            anchor='center'
        ).pack(fill='x', pady=8)

        main = ttk.Frame(self.root, padding=10)
        main.pack(fill='both', expand=True)

        header = ttk.Frame(main)
        header.pack(fill='x', pady=(0, 10))

        self.symbol_canvas = tk.Canvas(header, width=70, height=70, bg='#fff', highlightthickness=1, relief='solid')
        self.symbol_canvas.pack(side='left', padx=(0, 12))
        self._draw_symbol()

        tf = ttk.Frame(header)
        tf.pack(side='left', fill='x', expand=True, padx=5)

        ttk.Label(
            tf,
            text="Catálogo de Livros da Fraternidade Eclética Espiritualista Universal",
            style='Header.TLabel'
        ).pack(pady=(5, 0))

        self.relogio_frame = ttk.Frame(tf)
        self.relogio_frame.pack(pady=(2, 5))

        self.lbl_hora = tk.Label(self.relogio_frame, text=" ", font=('DejaVu Sans', 11, 'bold'), fg='#2563eb')
        self.lbl_hora.pack(side='left', padx=(0, 5))

        tk.Label(self.relogio_frame, text="|", font=('DejaVu Sans', 10, 'bold'), fg='#94a3b8').pack(side='left')

        self.lbl_data = tk.Label(self.relogio_frame, text=" ", font=('DejaVu Sans', 11, 'bold'), fg='#16a34a')
        self.lbl_data.pack(side='left', padx=5)

        tk.Label(self.relogio_frame, text="|", font=('DejaVu Sans', 10, 'bold'), fg='#94a3b8').pack(side='left')

        self.lbl_dia = tk.Label(self.relogio_frame, text=" ", font=('DejaVu Sans', 10, 'bold'), fg='#d97706')
        self.lbl_dia.pack(side='left', padx=5)

        ToolTip(self.relogio_frame, "Horário atual do sistema")

        hbtns = ttk.Frame(header)
        hbtns.pack(side='right', padx=10)

        self.btn_tema = tk.Button(
            hbtns,
            text="🌙 Tema Escuro",
            command=self.alternar_tema,
            bg='#334155',
            fg='white',
            font=('DejaVu Sans', 9, 'bold'),
            relief='raised',
            padx=8,
            pady=4,
            cursor='hand2'
        )
        self.btn_tema.pack(fill='x', pady=(0, 5))

        tk.Button(
            hbtns,
            text="❓ Ajuda",
            command=self.mostrar_manual,
            bg='#d97706',
            fg='white',
            font=('DejaVu Sans', 9, 'bold'),
            relief='raised',
            padx=8,
            pady=4,
            cursor='hand2'
        ).pack(fill='x', pady=(0, 5))

        tk.Button(
            hbtns,
            text="ℹ️ Sobre",
            command=self.mostrar_sobre,
            bg='#1e40af',
            fg='white',
            font=('DejaVu Sans', 9, 'bold'),
            relief='raised',
            padx=8,
            pady=4,
            cursor='hand2'
        ).pack(fill='x')

        busca = ttk.LabelFrame(main, text=" 🔍 Busca e Filtros ", padding=8)
        busca.pack(fill='x', pady=5)

        self.filtro_var = tk.StringVar(value="Todos")
        self.busca_var = tk.StringVar()
        self.status_var = tk.StringVar(value="Exibindo 0 livros")

        ttk.Label(busca, text="Filtrar por: ").grid(row=0, column=0, padx=5, sticky='e')

        cmb = ttk.Combobox(
            busca,
            textvariable=self.filtro_var,
            values=['Todos', 'Título', 'Autor', 'Estante', 'Prateleira', 'Editora', 'Assunto', 'Bibliotecário', 'Emprestado a'],
            state='readonly',
            width=15
        )
        cmb.grid(row=0, column=1, padx=5)

        ttk.Label(busca, text="Termo: ").grid(row=0, column=2, padx=5, sticky='e')

        ent = ttk.Entry(busca, textvariable=self.busca_var, width=30)
        ent.grid(row=0, column=3, padx=5)
        ent.bind("<Return>", lambda e: self.buscar())

        ttk.Button(busca, text="🔍 Buscar", command=self.buscar).grid(row=0, column=4, padx=5)
        ttk.Button(busca, text="🔄 Limpar", command=self.limpar_busca).grid(row=0, column=5, padx=5)
        ttk.Label(busca, textvariable=self.status_var, style='Status.TLabel').grid(row=0, column=6, padx=10, sticky='e')

        exp = ttk.LabelFrame(main, text=" 📤 Exportar, Importar e Imprimir ", padding=8)
        exp.pack(fill='x', pady=5)

        ttk.Label(exp, text="Exportar: ").grid(row=0, column=0, padx=5, sticky='e')

        self.btn_export_excel = ttk.Button(
            exp,
            text="📊 Excel",
            command=lambda: self._gerar_arquivo_planilha('xlsx'),
            style='Excel.TButton',
            width=15
        )
        self.btn_export_excel.grid(row=0, column=1, padx=3, pady=3)

        self.btn_export_calc = ttk.Button(
            exp,
            text="📈 Calc",
            command=lambda: self._gerar_arquivo_planilha('ods'),
            style='Calc.TButton',
            width=15
        )
        self.btn_export_calc.grid(row=0, column=2, padx=3, pady=3)

        self.btn_export_pdf = ttk.Button(
            exp,
            text="📄 PDF",
            command=self.exportar_pdf,
            style='PDF.TButton',
            width=15
        )
        self.btn_export_pdf.grid(row=0, column=3, padx=3, pady=3)

        ttk.Separator(exp, orient='vertical').grid(row=0, column=4, padx=8, pady=5, sticky='ns')

        ttk.Label(exp, text="Importar: ").grid(row=0, column=5, padx=5, sticky='e')

        self.btn_import = ttk.Button(
            exp,
            text="📥 Planilha",
            command=self.importar_planilha,
            style='Import.TButton',
            width=20
        )
        self.btn_import.grid(row=0, column=6, padx=3, pady=3)

        ttk.Separator(exp, orient='vertical').grid(row=0, column=7, padx=8, pady=5, sticky='ns')

        ttk.Label(exp, text="Imprimir: ").grid(row=0, column=8, padx=5, sticky='e')

        self.btn_imp_cat = ttk.Button(
            exp,
            text="🖨️ Catálogo",
            command=lambda: self.imprimir_catalogo('completo'),
            style='Print.TButton',
            width=14
        )
        self.btn_imp_cat.grid(row=0, column=9, padx=3, pady=3)

        self.btn_imp_busca = ttk.Button(
            exp,
            text="🖨️ Busca",
            command=lambda: self.imprimir_catalogo('busca'),
            style='Print.TButton',
            width=14
        )
        self.btn_imp_busca.grid(row=0, column=10, padx=3, pady=3)
        self.btn_imp_busca.config(state='disabled')

        ttk.Label(
            exp,
            text="*PDF abre no visualizador para Ctrl+P   •   '🖨️ Busca' só fica ativo quando houver resultados de busca",
            font=('DejaVu Sans', 8),
            foreground='#64748b'
        ).grid(row=1, column=0, columnspan=11, sticky='w', padx=5, pady=(0, 5))

        self.main_paned = ttk.PanedWindow(main, orient='vertical')
        self.main_paned.pack(fill='both', expand=True, pady=5)

        t_frame = ttk.LabelFrame(self.main_paned, text=" 📖 Catálogo Completo ", padding=5)

        self.tree = TabelaCatalogo(
            t_frame,
            columns=self.COLUNAS,
            on_select=self._ao_selecionar_catalogo,
            on_double=self._ao_duplo_catalogo
        )
        self.tree.pack(fill='both', expand=True)

        self.main_paned.add(t_frame, weight=3)

        self.resultados_frame = ttk.LabelFrame(self.main_paned, text=" 🔍 Resultados Encontrados ", padding=5)
        rc = ttk.Frame(self.resultados_frame)
        rc.pack(fill='both', expand=True)

        self.resultados_tree = TabelaCatalogo(
            rc,
            columns=self.COLUNAS,
            on_select=self._ao_selecionar_resultado,
            on_double=self._ao_duplo_resultado
        )
        self.resultados_tree.pack(fill='both', expand=True)

        ttk.Button(self.resultados_frame, text="✖ Fechar", command=self.fechar_painel_resultados).pack(side='right', padx=10, pady=5)

        ff = ttk.LabelFrame(self.main_paned, text=" Cadastro / Edição / Localização / Empréstimos ", padding=10)

        self.entries = {}

        cfgs = [
            ('Título:', 'titulo', 0, 0, 20, False),
            ('Autor:', 'autor', 0, 2, 18, True),
            ('Estante:', 'estante', 0, 4, 10, True),
            ('Prateleira:', 'prateleira', 0, 6, 10, True),

            ('Editora:', 'editora', 1, 0, 18, True),
            ('Assunto:', 'assunto', 1, 2, 18, True),
            ('Bibliotecário:', 'bibliotecario', 1, 4, 18, True),
            ('Quantidade:', 'quantidade', 1, 6, 10, False),

            ('Emprestado a:', 'emprestado_a', 2, 0, 20, True),
            ('Entrada:', 'entrada', 2, 2, 12, False, True),
            ('Saída:', 'saida', 2, 4, 12, False, True),
        ]

        for config in cfgs:
            label, name, row, col, width, is_combo = config[:6]
            is_date = config[6] if len(config) > 6 else False

            lbl = ttk.Label(ff, text=label)
            lbl.grid(row=row, column=col, padx=5, pady=5, sticky='e')

            ef = ttk.Frame(ff)
            ef.grid(row=row, column=col + 1, padx=5, pady=5, sticky='we')

            entry = ttk.Combobox(ef, width=width - 4 if is_date else width, state='normal') if is_combo else ttk.Entry(ef, width=width - 4 if is_date else width)
            entry.pack(side='left', fill='x', expand=True)

            if is_date:
                cal_btn = tk.Button(
                    ef,
                    text="📅",
                    relief='raised',
                    bd=1,
                    bg='#2563eb',
                    fg='#ffffff',
                    activebackground='#1d4ed8',
                    activeforeground='#ffffff',
                    font=('DejaVu Sans', 12),
                    padx=4,
                    pady=2,
                    cursor='hand2',
                    command=lambda e=entry: DateChooser(self.root, e)
                )
                cal_btn.pack(side='left', padx=(5, 0))

            if name in ('estante', 'prateleira'):
                add_btn = tk.Button(
                    ef,
                    text="➕",
                    relief='raised',
                    bd=1,
                    bg='#059669',
                    fg='#ffffff',
                    activebackground='#047857',
                    activeforeground='#ffffff',
                    font=('DejaVu Sans', 10, 'bold'),
                    padx=6,
                    pady=1,
                    cursor='hand2',
                    command=lambda n=name: self._adicionar_localizacao(n)
                )
                add_btn.pack(side='left', padx=(5, 0))

            self.entries[name] = entry

        for c in (1, 3, 5, 7):
            ff.columnconfigure(c, weight=1)

        self.entries['entrada'].insert(0, datetime.now().strftime('%d/%m/%Y'))
        self.entries['saida'].insert(0, datetime.now().strftime('%d/%m/%Y'))
        self.entries['quantidade'].insert(0, '1')

        bf = ttk.Frame(ff)
        bf.grid(row=3, column=0, columnspan=8, pady=10, sticky='w')

        self.btn_adicionar = ttk.Button(bf, text="➕ Adicionar", command=self.adicionar_livro, style='Adicionar.TButton')
        self.btn_adicionar.pack(side='left', padx=5)

        self.btn_salvar = ttk.Button(bf, text="💾 Salvar", command=self.salvar, style='Salvar.TButton')
        self.btn_salvar.pack(side='left', padx=5)

        self.btn_excluir = ttk.Button(bf, text="🗑️ Excluir", command=self.excluir, style='Excluir.TButton')
        self.btn_excluir.pack(side='left', padx=5)

        self.btn_limpar = ttk.Button(bf, text="🔄 Limpar", command=self.limpar_form, style='Limpar.TButton')
        self.btn_limpar.pack(side='left', padx=5)

        ttk.Separator(bf, orient='vertical').pack(side='left', padx=10, fill='y')

        self.btn_emprestar = ttk.Button(bf, text="📤 Emprestar", command=self.emprestar_livro, style='Emprestar.TButton')
        self.btn_emprestar.pack(side='left', padx=5)

        self.btn_devolver = ttk.Button(bf, text="📥 Devolver", command=self.devolver_livro, style='Devolver.TButton')
        self.btn_devolver.pack(side='left', padx=5)

        ttk.Separator(bf, orient='vertical').pack(side='left', padx=15, fill='y')

        self.btn_emprestados = ttk.Button(bf, text="📚 Emprestados", command=self.mostrar_emprestados, style='Import.TButton')
        self.btn_emprestados.pack(side='left', padx=5)

        self.btn_users = ttk.Button(bf, text="🔑 Usuários", command=self.mostrar_gerenciar_usuarios, style='Import.TButton')
        self.btn_users.pack(side='left', padx=5)

        self.btn_senha_admin = ttk.Button(bf, text="🔒 Senha Admin", command=self.alterar_senha_admin, style='Limpar.TButton')
        self.btn_senha_admin.pack(side='left', padx=5)

        ttk.Separator(bf, orient='vertical').pack(side='left', padx=15, fill='y')

        self.btn_sair = ttk.Button(bf, text="🚪 Sair", command=self.confirmar_saida, style='Sair.TButton')
        self.btn_sair.pack(side='right', padx=5)

        self.btn_voltar_login = ttk.Button(bf, text="🔙 Voltar ao Login", command=self.voltar_para_login, style='Voltar.TButton')
        self.btn_voltar_login.pack(side='right', padx=5)

        self.main_paned.add(ff, weight=2)

    def aplicar_permissoes(self):
        p = getattr(self, 'permissoes', {}) or {}
        admin = bool(p.get('admin'))

        def habilitar(nome, ativo):
            if hasattr(self, nome):
                getattr(self, nome).config(state='normal' if ativo else 'disabled')

        habilitar('btn_adicionar', admin or p.get('inserir'))
        habilitar('btn_salvar', admin or p.get('inserir') or p.get('editar'))
        habilitar('btn_excluir', admin or p.get('apagar'))

        habilitar('btn_import', admin or p.get('importar'))

        for btn in ('btn_export_excel', 'btn_export_calc', 'btn_export_pdf'):
            habilitar(btn, admin or p.get('exportar'))

        habilitar('btn_imp_cat', admin or p.get('relatorio'))

        emprestimos = admin or p.get('emprestados')
        for btn in ('btn_emprestar', 'btn_devolver', 'btn_emprestados'):
            habilitar(btn, emprestimos)

        habilitar('btn_users', admin)
        habilitar('btn_senha_admin', admin)

        self._atualizar_estado_btn_busca()

    # ==========================================================================
    # LOCALIZAÇÃO
    # ==========================================================================
    def _adicionar_localizacao(self, campo):
        titulo = "Estante" if campo == "estante" else "Prateleira"

        valor = simpledialog.askstring(
            f"Adicionar {titulo}",
            f"Digite um novo valor para {titulo}:",
            parent=self.root
        )

        if valor and valor.strip():
            valor = valor.strip()
            self.registrar_historico_campo(campo, valor)
            self.atualizar_comboboxes()

            try:
                self.entries[campo].delete(0, tk.END)
                self.entries[campo].insert(0, valor)
            except Exception:
                pass

    def registrar_historico_campo(self, campo, valor):
        if not str(valor).strip():
            return

        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()

        cursor.execute(
            "INSERT OR IGNORE INTO historico_campos (campo, valor) VALUES (?, ?)",
            (campo, str(valor).strip())
        )

        conn.commit()
        conn.close()

    def atualizar_comboboxes(self):
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()

        for c in ['autor', 'editora', 'assunto', 'bibliotecario', 'emprestado_a', 'estante', 'prateleira']:
            cursor.execute("SELECT valor FROM historico_campos WHERE campo=? ORDER BY valor", (c,))
            self.entries[c]['values'] = [r[0] for r in cursor.fetchall()]

        conn.close()

    # ==========================================================================
    # CRUD
    # ==========================================================================
    def adicionar_livro(self):
        if not self._checar_permissao('inserir', 'Você não tem permissão para inserir livros.'):
            return

        try:
            t = self.entries['titulo'].get().strip()
            a = self.entries['autor'].get().strip()

            if not t or not a:
                messagebox.showwarning("⚠️", "Preencha Título e Autor!")
                return

            q, ok = self._validar_quantidade(self.entries['quantidade'].get())
            if not ok:
                messagebox.showwarning("⚠️", "Quantidade inválida. Use número inteiro maior ou igual a zero.")
                return

            est = self.entries['estante'].get().strip()
            prat = self.entries['prateleira'].get().strip()
            e = self.entries['editora'].get().strip()
            ass = self.entries['assunto'].get().strip()
            bib = self.entries['bibliotecario'].get().strip()
            emp = self.entries['emprestado_a'].get().strip()

            entrada_iso = self._data_br_para_iso(self.entries['entrada'].get(), manter_original=True)
            saida_iso = self._data_br_para_iso(self.entries['saida'].get(), manter_original=True)

            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            dup = self._checar_duplicata(t, a, e, ass, est, prat, cursor=cursor)

            if dup:
                lid, qat = dup
                nq = (qat if qat else 0) + q

                cursor.execute(
                    "UPDATE livros SET quantidade=?, disponibilidade=? WHERE id=?",
                    (nq, 'Sim' if nq > 0 else 'Não', lid)
                )

                msg = f"Livro existente!\nQtd somada: {q}\nTotal: {nq}"
            else:
                cursor.execute(
                    """
                    INSERT INTO livros
                    (titulo, autor, estante, prateleira, editora, assunto, bibliotecario, quantidade, disponibilidade, emprestado_a, entrada, saida)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (t, a, est, prat, e, ass, bib, q, 'Sim' if q > 0 else 'Não', emp, entrada_iso, saida_iso)
                )

                msg = f"Novo livro '{t}' adicionado!"

            self._registrar_historico_livro(
                {
                    'autor': a,
                    'editora': e,
                    'assunto': ass,
                    'bibliotecario': bib,
                    'emprestado_a': emp,
                    'estante': est,
                    'prateleira': prat
                },
                cursor=cursor
            )

            conn.commit()
            conn.close()

            self._recarregar_manter_busca()
            self.atualizar_comboboxes()
            self.limpar_form()

            messagebox.showinfo("✅", msg)

        except Exception:
            messagebox.showerror("❌", traceback.format_exc())

    def salvar(self):
        if self.livro_selecionado_id is None:
            if not self._checar_permissao('inserir', 'Você não tem permissão para inserir livros.'):
                return
        else:
            if not self._checar_permissao('editar', 'Você não tem permissão para editar livros.'):
                return

        try:
            t = self.entries['titulo'].get().strip()
            a = self.entries['autor'].get().strip()

            if not t or not a:
                messagebox.showwarning("⚠️", "Título e Autor obrigatórios!")
                return

            q, ok = self._validar_quantidade(self.entries['quantidade'].get())
            if not ok:
                messagebox.showwarning("⚠️", "Quantidade inválida. Use número inteiro maior ou igual a zero.")
                return

            est = self.entries['estante'].get().strip()
            prat = self.entries['prateleira'].get().strip()
            e = self.entries['editora'].get().strip()
            ass = self.entries['assunto'].get().strip()
            bib = self.entries['bibliotecario'].get().strip()
            emp = self.entries['emprestado_a'].get().strip()

            entrada_iso = self._data_br_para_iso(self.entries['entrada'].get(), manter_original=True)
            saida_iso = self._data_br_para_iso(self.entries['saida'].get(), manter_original=True)

            disp = 'Sim' if q > 0 else 'Não'

            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            dup = self._checar_duplicata(t, a, e, ass, est, prat, cursor=cursor)

            if dup and dup[0] != self.livro_selecionado_id:
                if messagebox.askyesno(
                    "⚠️ Registro duplicado",
                    "Já existe um livro com esses dados e localização.\nDeseja somar a quantidade ao registro existente?"
                ):
                    lid, qat = dup
                    nq = (qat if qat else 0) + q

                    cursor.execute(
                        "UPDATE livros SET quantidade=?, disponibilidade=? WHERE id=?",
                        (nq, 'Sim' if nq > 0 else 'Não', lid)
                    )

                    self._registrar_historico_livro(
                        {
                            'autor': a,
                            'editora': e,
                            'assunto': ass,
                            'bibliotecario': bib,
                            'emprestado_a': emp,
                            'estante': est,
                            'prateleira': prat
                        },
                        cursor=cursor
                    )

                    conn.commit()
                    conn.close()

                    self._recarregar_manter_busca()
                    self.atualizar_comboboxes()
                    self.limpar_form()

                    messagebox.showinfo("✅", f"Quantidade somada ao registro existente.\nTotal: {nq}")
                    return
                else:
                    conn.close()
                    return

            if self.livro_selecionado_id is None:
                cursor.execute(
                    """
                    INSERT INTO livros
                    (titulo, autor, estante, prateleira, editora, assunto, bibliotecario, quantidade, disponibilidade, emprestado_a, entrada, saida)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (t, a, est, prat, e, ass, bib, q, disp, emp, entrada_iso, saida_iso)
                )
                msg = f"'{t}' adicionado!"
            else:
                cursor.execute(
                    """
                    UPDATE livros
                    SET titulo=?, autor=?, estante=?, prateleira=?, editora=?, assunto=?, bibliotecario=?, quantidade=?, disponibilidade=?, emprestado_a=?, entrada=?, saida=?
                    WHERE id=?
                    """,
                    (t, a, est, prat, e, ass, bib, q, disp, emp, entrada_iso, saida_iso, self.livro_selecionado_id)
                )
                msg = f"'{t}' atualizado!"

            self._registrar_historico_livro(
                {
                    'autor': a,
                    'editora': e,
                    'assunto': ass,
                    'bibliotecario': bib,
                    'emprestado_a': emp,
                    'estante': est,
                    'prateleira': prat
                },
                cursor=cursor
            )

            conn.commit()
            conn.close()

            self._recarregar_manter_busca()
            self.atualizar_comboboxes()
            self.limpar_form()

            messagebox.showinfo("✅", msg)

        except Exception:
            messagebox.showerror("❌", traceback.format_exc())

    def excluir(self):
        if not self._checar_permissao('apagar', 'Você não tem permissão para excluir livros.'):
            return

        if not self.livro_selecionado_id:
            messagebox.showwarning("⚠️", "Selecione um livro!")
            return

        l = next((x for x in self.livros if x['id'] == self.livro_selecionado_id), None)

        if l and messagebox.askyesno("🗑️", f"Excluir '{l['titulo']}'?"):
            conn = sqlite3.connect(self.db_file)
            cursor = conn.cursor()

            cursor.execute("DELETE FROM livros WHERE id=?", (self.livro_selecionado_id,))

            conn.commit()
            conn.close()

            self._recarregar_manter_busca()
            self.limpar_form()

            messagebox.showinfo("✅", "Excluído!")

    def limpar_form(self):
        for n in self.entries:
            self.entries[n].delete(0, tk.END)

        self.entries['entrada'].insert(0, datetime.now().strftime('%d/%m/%Y'))
        self.entries['saida'].insert(0, datetime.now().strftime('%d/%m/%Y'))
        self.entries['quantidade'].insert(0, '1')

        self.livro_selecionado_id = None

    # ==========================================================================
    # SELEÇÃO / DADOS
    # ==========================================================================
    def _ao_selecionar_catalogo(self, livro_id):
        self._carregar_form_por_id(livro_id)

    def _ao_duplo_catalogo(self, livro_id):
        self._carregar_form_por_id(livro_id)

        try:
            self.entries['titulo'].focus_set()
        except Exception:
            pass

    def _ao_selecionar_resultado(self, livro_id):
        self._carregar_form_por_id(livro_id)

    def _ao_duplo_resultado(self, livro_id):
        self._carregar_form_por_id(livro_id)

        try:
            self.entries['titulo'].focus_set()
        except Exception:
            pass

    def _carregar_form_por_id(self, livro_id):
        l = next((x for x in self.livros if x['id'] == livro_id), None)

        if not l and self.ultima_busca:
            l = next((x for x in self.ultima_busca if x['id'] == livro_id), None)

        if not l:
            return

        self.livro_selecionado_id = livro_id

        def set_entry(nome, valor):
            self.entries[nome].delete(0, tk.END)
            self.entries[nome].insert(0, '' if valor is None else str(valor))

        set_entry('titulo', l['titulo'])
        set_entry('autor', l['autor'])
        set_entry('estante', l.get('estante', ''))
        set_entry('prateleira', l.get('prateleira', ''))
        set_entry('editora', l['editora'])
        set_entry('assunto', l['assunto'])
        set_entry('bibliotecario', l['bibliotecario'])
        set_entry('quantidade', l['quantidade'])
        set_entry('emprestado_a', l['emprestado_a'])
        set_entry('entrada', l['entrada'])
        set_entry('saida', l['saida'])

    def carregar_dados(self):
        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()

        cols = self._colunas_livros_existentes(cursor)

        ordem = [
            'id', 'titulo', 'autor', 'estante', 'prateleira', 'editora', 'assunto',
            'bibliotecario', 'quantidade', 'disponibilidade', 'emprestado_a', 'entrada', 'saida'
        ]

        defaults = {
            'id': 'NULL',
            'titulo': "''",
            'autor': "''",
            'estante': "''",
            'prateleira': "''",
            'editora': "''",
            'assunto': "''",
            'bibliotecario': "''",
            'quantidade': '0',
            'disponibilidade': "'Sim'",
            'emprestado_a': "''",
            'entrada': "''",
            'saida': "''"
        }

        select_cols = []

        for c in ordem:
            if c in cols:
                select_cols.append(c)
            else:
                select_cols.append(f"{defaults[c]} AS {c}")

        sql = f"SELECT {', '.join(select_cols)} FROM livros ORDER BY id"
        cursor.execute(sql)

        self.livros = []

        for r in cursor.fetchall():
            item = {ordem[i]: r[i] for i in range(len(ordem))}
            item['entrada'] = self._data_iso_para_br(item['entrada'])
            item['saida'] = self._data_iso_para_br(item['saida'])
            self.livros.append(item)

        conn.close()
        self.atualizar_tabelas()

    def atualizar_tabelas(self):
        if hasattr(self, 'tree'):
            self.tree.set_data(self.livros)

        self.status_var.set(f"Exibindo {len(self.livros)} livros")

    # ==========================================================================
    # BUSCA
    # ==========================================================================
    def buscar(self):
        termo = self.busca_var.get().strip()

        if not termo:
            self.limpar_busca()
            return

        filtro = self.filtro_var.get()

        cmap = {
            'Título': 'titulo',
            'Autor': 'autor',
            'Estante': 'estante',
            'Prateleira': 'prateleira',
            'Editora': 'editora',
            'Assunto': 'assunto',
            'Bibliotecário': 'bibliotecario',
            'Emprestado a': 'emprestado_a'
        }

        if filtro == 'Emprestado a' and termo == '*':
            filtrados = [l for l in self.livros if str(l.get('emprestado_a', '')).strip()]
        elif filtro != 'Todos':
            campo = cmap.get(filtro, 'titulo')
            filtrados = [l for l in self.livros if self._wildcard_match(termo, l.get(campo, ''))]
        else:
            filtrados = [
                l for l in self.livros
                if any(self._wildcard_match(termo, l.get(c, '')) for c in cmap.values())
            ]

        self.busca_ativa = True
        self.ultima_busca = filtrados
        self.status_var.set(f"🔍 {len(filtrados)} resultado(s)")

        if str(self.resultados_frame) not in self.main_paned.panes():
            self.main_paned.insert(1, self.resultados_frame)

            try:
                self.main_paned.sashpos(1, self.root.winfo_height() // 2)
            except Exception:
                pass

        self.resultados_tree.set_data(filtrados)
        self._atualizar_estado_btn_busca()

    def limpar_busca(self):
        self.busca_var.set('')
        self.filtro_var.set('Todos')
        self.busca_ativa = False
        self.ultima_busca = []

        self.fechar_painel_resultados()
        self.carregar_dados()
        self.limpar_form()
        self._atualizar_estado_btn_busca()

    def fechar_painel_resultados(self):
        try:
            self.main_paned.forget(self.resultados_frame)
        except Exception:
            pass

        self.busca_ativa = False
        self.ultima_busca = []
        self.status_var.set(f"Exibindo {len(self.livros)} livros")
        self._atualizar_estado_btn_busca()

    # ==========================================================================
    # EMPRÉSTIMOS
    # ==========================================================================
    def emprestar_livro(self):
        if not self._checar_permissao('emprestados', 'Você não tem permissão para controlar empréstimos.'):
            return

        if not self.livro_selecionado_id:
            messagebox.showwarning("⚠️", "Selecione um livro.")
            return

        l = next((x for x in self.livros if x['id'] == self.livro_selecionado_id), None)

        if not l or l['quantidade'] <= 0:
            messagebox.showerror("❌", "Estoque zerado.")
            return

        nm = simpledialog.askstring("📤 Emprestar", f"Nome de quem retira '{l['titulo']}':", parent=self.root)

        if not nm or not nm.strip():
            return

        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()

        nq = l['quantidade'] - 1
        la = str(l.get('emprestado_a', '') or '').strip()

        cursor.execute(
            "UPDATE livros SET quantidade=?, disponibilidade=?, emprestado_a=? WHERE id=?",
            (
                nq,
                'Não' if nq <= 0 else 'Sim',
                f"{la}; {nm.strip()}" if la else nm.strip(),
                self.livro_selecionado_id
            )
        )

        conn.commit()
        conn.close()

        self._recarregar_manter_busca()
        self.limpar_form()

        messagebox.showinfo("✅", f"Emprestado para: {nm.strip()}\nRestante: {nq}")

    def _dialogo_devolucao(self, nomes):
        d = tk.Toplevel(self.root)
        d.title("📥 Selecionar Devolução")
        d.transient(self.root)
        d.grab_set()
        d.resizable(True, True)

        res = {"sel": None}

        ttk.Label(d, text="Selecione quem devolve:").pack(pady=5)

        lf = ttk.Frame(d)
        lf.pack(fill='both', expand=True, padx=10, pady=5)

        lb = tk.Listbox(lf, selectmode=tk.SINGLE)
        lb.pack(side='left', fill='both', expand=True)

        tk.Scrollbar(lf, orient='vertical', command=lb.yview).pack(side='right', fill='y')

        for n in nomes:
            lb.insert(tk.END, n)

        if nomes:
            lb.selection_set(0)

        bf = ttk.Frame(d)
        bf.pack(fill='x', padx=10, pady=10)

        def ok():
            s = lb.curselection()
            if s:
                res["sel"] = lb.get(s[0])
            d.destroy()

        ttk.Button(bf, text="OK", command=ok).pack(side='left', expand=True, padx=5)
        ttk.Button(bf, text="Cancelar", command=d.destroy).pack(side='left', expand=True, padx=5)

        self.centralizar_janela(d)
        self.root.wait_window(d)

        return res["sel"]

    def devolver_livro(self):
        if not self._checar_permissao('emprestados', 'Você não tem permissão para controlar empréstimos.'):
            return

        if not self.livro_selecionado_id:
            messagebox.showwarning("⚠️", "Selecione um livro.")
            return

        l = next((x for x in self.livros if x['id'] == self.livro_selecionado_id), None)

        if not l or not str(l.get('emprestado_a', '')).strip():
            messagebox.showinfo("ℹ️", "Nenhum empréstimo pendente.")
            return

        nomes = [n.strip() for n in str(l['emprestado_a']).split(';') if n.strip()]
        dev = self._dialogo_devolucao(nomes)

        if not dev:
            return

        conn = sqlite3.connect(self.db_file)
        cursor = conn.cursor()

        q = l['quantidade'] + 1
        nl = '; '.join([n for n in nomes if n != dev])

        cursor.execute(
            "UPDATE livros SET quantidade=?, disponibilidade=?, emprestado_a=? WHERE id=?",
            (q, 'Sim' if q > 0 else 'Não', nl, self.livro_selecionado_id)
        )

        conn.commit()
        conn.close()

        self._recarregar_manter_busca()
        self.limpar_form()

        messagebox.showinfo("✅", f"Devolvido por {dev}!\nEstoque: {q}")

    def mostrar_emprestados(self):
        if not self._checar_permissao('emprestados', 'Você não tem permissão para controlar empréstimos.'):
            return

        win = tk.Toplevel(self.root)
        win.title("📚 Livros Emprestados")
        win.minsize(560, 360)
        win.transient(self.root)
        win.grab_set()
        win.resizable(True, True)

        ttk.Label(win, text="Empréstimos ativos:", font=('DejaVu Sans', 11, 'bold')).pack(anchor='w', pady=(10, 5))

        frame = ttk.Frame(win)
        frame.pack(fill='both', expand=True, padx=10, pady=10)

        cols = ('id', 'titulo', 'emprestado_a')
        tree = ttk.Treeview(frame, columns=cols, show='headings')

        tree.heading('id', text='ID')
        tree.column('id', width=50, anchor='center', stretch=False)

        tree.heading('titulo', text='Livro')
        tree.column('titulo', width=250, anchor='w', stretch=True)

        tree.heading('emprestado_a', text='Emprestado a')
        tree.column('emprestado_a', width=220, anchor='w', stretch=True)

        tree.pack(side='left', fill='both', expand=True)
        ttk.Scrollbar(frame, orient='vertical', command=tree.yview).pack(side='right', fill='y')

        def atualizar():
            for i in tree.get_children():
                tree.delete(i)

            conn = sqlite3.connect(self.db_file)
            cur = conn.cursor()

            cur.execute("""
                SELECT id, titulo, emprestado_a
                FROM livros
                WHERE COALESCE(emprestado_a, '') != ''
                ORDER BY titulo
            """)

            for r in cur.fetchall():
                tree.insert('', 'end', values=r)

            conn.close()

        def devolver_no_painel():
            sel = tree.selection()

            if not sel:
                messagebox.showwarning("⚠️", "Selecione um livro na lista.", parent=win)
                return

            vals = tree.item(sel[0])['values']

            livro_id = int(vals[0])
            tit = vals[1]
            emp = str(vals[2])

            nomes = [n.strip() for n in emp.split(';') if n.strip()]

            if not nomes:
                messagebox.showinfo("ℹ️", "Nenhum pendente.", parent=win)
                atualizar()
                return

            dev = self._dialogo_devolucao(nomes)

            if not dev:
                return

            if messagebox.askyesno("📥 Confirmar", f"Devolver '{tit}' por {dev}?", parent=win):
                conn = sqlite3.connect(self.db_file)
                cursor = conn.cursor()

                cursor.execute("SELECT quantidade, emprestado_a FROM livros WHERE id=?", (livro_id,))
                row = cursor.fetchone()

                if not row:
                    conn.close()
                    messagebox.showerror("❌", "Registro não encontrado.", parent=win)
                    atualizar()
                    return

                q = (row[0] or 0) + 1
                nomes_atuais = [n.strip() for n in str(row[1] or '').split(';') if n.strip()]
                nl = '; '.join([n for n in nomes_atuais if n != dev])

                cursor.execute(
                    "UPDATE livros SET quantidade=?, disponibilidade=?, emprestado_a=? WHERE id=?",
                    (q, 'Sim' if q > 0 else 'Não', nl, livro_id)
                )

                conn.commit()
                conn.close()

                messagebox.showinfo("✅", f"Devolvido por {dev}!\nEstoque: {q}", parent=win)
                atualizar()
                self._recarregar_manter_busca()

        bf = ttk.Frame(win)
        bf.pack(fill='x', padx=10, pady=(0, 10), side='bottom')

        ttk.Button(bf, text="📥 Devolver Selecionado", command=devolver_no_painel, style='Devolver.TButton').pack(side='left', padx=5)
        ttk.Button(bf, text="Fechar", command=win.destroy).pack(side='right', padx=5)

        atualizar()
        self.centralizar_janela(win)

    # ==========================================================================
    # USUÁRIOS
    # ==========================================================================
    def mostrar_gerenciar_usuarios(self):
        if not self._checar_permissao('admin', 'Apenas administradores podem gerenciar usuários.'):
            return

        d = tk.Toplevel(self.root)
        d.title("🔐 Gerenciar Usuários")
        d.transient(self.root)
        d.grab_set()
        d.minsize(820, 500)

        self.centralizar_janela(d)

        main = ttk.Frame(d)
        main.pack(fill='both', expand=True, padx=10, pady=5)

        ft = ttk.LabelFrame(main, text="Dados do Usuário", padding=10)
        ft.pack(fill='x', pady=5)

        ttk.Label(ft, text="Usuário: ").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        en_nome = ttk.Entry(ft, width=25)
        en_nome.grid(row=0, column=1, sticky='w', padx=5)

        ttk.Label(ft, text="Senha: ").grid(row=0, column=2, sticky='w', padx=5, pady=5)
        en_senha = ttk.Entry(ft, show='*', width=25)
        en_senha.grid(row=0, column=3, sticky='w', padx=5)

        ttk.Label(ft, text="Permissões: ").grid(row=1, column=0, columnspan=4, sticky='w', padx=5, pady=(10, 2))

        perms = ['inserir', 'editar', 'apagar', 'exportar', 'importar', 'relatorio', 'admin', 'emprestados']
        vp = {k: tk.IntVar(value=0) for k in perms}

        for k in ['inserir', 'editar', 'exportar', 'importar', 'relatorio', 'emprestados']:
            vp[k].set(1)

        for i, k in enumerate(perms):
            texto = 'Admin' if k == 'admin' else ('Emprestados' if k == 'emprestados' else k.capitalize())
            tk.Checkbutton(ft, text=texto, variable=vp[k]).grid(row=2, column=i, sticky='w', padx=5)

        fb = ttk.LabelFrame(main, text="Cadastrados", padding=10)
        fb.pack(fill='both', expand=True, pady=5)

        cols = ('nome', 'inserir', 'editar', 'apagar', 'exportar', 'importar', 'relatorio', 'admin', 'emprestados')
        tv = ttk.Treeview(fb, columns=cols, show='headings', height=8)

        for c in cols:
            tv.heading(c, text='Usuário' if c == 'nome' else c.capitalize())
            tv.column(c, width=180 if c == 'nome' else 85, anchor='w' if c == 'nome' else 'center')

        tv.pack(fill='both', expand=True, side='left')
        tk.Scrollbar(fb, orient='vertical', command=tv.yview).pack(side='right', fill='y')

        btn_frame = ttk.Frame(d, padding=5)
        btn_frame.pack(fill='x', padx=10, pady=5)

        def load_users():
            for i in tv.get_children():
                tv.delete(i)

            conn = sqlite3.connect(self.db_file)
            cur = conn.cursor()

            cur.execute("""
                SELECT nome, perm_inserir, perm_editar, perm_apagar, perm_exportar,
                       perm_importar, perm_relatorio, perm_admin, perm_emprestados
                FROM usuarios
                ORDER BY nome
            """)

            for r in cur.fetchall():
                vals = [r[0]] + ["Sim" if bool(x) else "Não" for x in r[1:]]
                tv.insert('', 'end', values=vals)

            conn.close()

        def on_select(event):
            sel = tv.selection()

            if not sel:
                return

            vals = tv.item(sel[0])['values']

            en_nome.delete(0, tk.END)
            en_nome.insert(0, vals[0])
            en_senha.delete(0, tk.END)

            vp['inserir'].set(1 if str(vals[1]).strip() == 'Sim' else 0)
            vp['editar'].set(1 if str(vals[2]).strip() == 'Sim' else 0)
            vp['apagar'].set(1 if str(vals[3]).strip() == 'Sim' else 0)
            vp['exportar'].set(1 if str(vals[4]).strip() == 'Sim' else 0)
            vp['importar'].set(1 if str(vals[5]).strip() == 'Sim' else 0)
            vp['relatorio'].set(1 if str(vals[6]).strip() == 'Sim' else 0)
            vp['admin'].set(1 if str(vals[7]).strip() == 'Sim' else 0)
            vp['emprestados'].set(1 if str(vals[8]).strip() == 'Sim' else 0)

        tv.bind('<<TreeviewSelect>>', on_select)

        def adicionar():
            nm = en_nome.get().strip()
            pw = en_senha.get().strip()

            if not nm or not pw:
                messagebox.showwarning("⚠️", "Preencha Usuário e Senha.", parent=d)
                return

            conn = sqlite3.connect(self.db_file)
            cur = conn.cursor()

            cur.execute("SELECT id FROM usuarios WHERE nome=?", (nm,))
            existe = cur.fetchone()

            if existe:
                cur.execute(
                    """
                    UPDATE usuarios
                    SET senha=?, perm_inserir=?, perm_editar=?, perm_apagar=?, perm_exportar=?,
                        perm_importar=?, perm_relatorio=?, perm_admin=?, perm_emprestados=?
                    WHERE nome=?
                    """,
                    (
                        self._hash_senha(pw),
                        vp['inserir'].get(),
                        vp['editar'].get(),
                        vp['apagar'].get(),
                        vp['exportar'].get(),
                        vp['importar'].get(),
                        vp['relatorio'].get(),
                        vp['admin'].get(),
                        vp['emprestados'].get(),
                        nm
                    )
                )
                msg = "Usuário atualizado!"
            else:
                cur.execute(
                    """
                    INSERT INTO usuarios
                    (nome, senha, perm_inserir, perm_editar, perm_apagar, perm_exportar,
                     perm_importar, perm_relatorio, perm_admin, perm_emprestados)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        nm,
                        self._hash_senha(pw),
                        vp['inserir'].get(),
                        vp['editar'].get(),
                        vp['apagar'].get(),
                        vp['exportar'].get(),
                        vp['importar'].get(),
                        vp['relatorio'].get(),
                        vp['admin'].get(),
                        vp['emprestados'].get()
                    )
                )
                msg = "Usuário adicionado!"

            conn.commit()
            conn.close()

            messagebox.showinfo("✅", msg, parent=d)
            load_users()

        def salvar():
            nm = en_nome.get().strip()

            if not nm:
                messagebox.showwarning("⚠️", "Selecione um usuário.", parent=d)
                return

            if nm == 'admin' and not vp['admin'].get():
                messagebox.showerror("❌", "O usuário admin não pode perder a permissão de administrador.", parent=d)
                return

            pw = en_senha.get().strip()

            conn = sqlite3.connect(self.db_file)
            cur = conn.cursor()

            if pw:
                cur.execute(
                    """
                    UPDATE usuarios
                    SET senha=?, perm_inserir=?, perm_editar=?, perm_apagar=?, perm_exportar=?,
                        perm_importar=?, perm_relatorio=?, perm_admin=?, perm_emprestados=?
                    WHERE nome=?
                    """,
                    (
                        self._hash_senha(pw),
                        vp['inserir'].get(),
                        vp['editar'].get(),
                        vp['apagar'].get(),
                        vp['exportar'].get(),
                        vp['importar'].get(),
                        vp['relatorio'].get(),
                        vp['admin'].get(),
                        vp['emprestados'].get(),
                        nm
                    )
                )
            else:
                cur.execute(
                    """
                    UPDATE usuarios
                    SET perm_inserir=?, perm_editar=?, perm_apagar=?, perm_exportar=?,
                        perm_importar=?, perm_relatorio=?, perm_admin=?, perm_emprestados=?
                    WHERE nome=?
                    """,
                    (
                        vp['inserir'].get(),
                        vp['editar'].get(),
                        vp['apagar'].get(),
                        vp['exportar'].get(),
                        vp['importar'].get(),
                        vp['relatorio'].get(),
                        vp['admin'].get(),
                        vp['emprestados'].get(),
                        nm
                    )
                )

            if cur.rowcount == 0:
                conn.close()
                messagebox.showwarning("⚠️", "Usuário não encontrado.", parent=d)
                return

            conn.commit()
            conn.close()

            messagebox.showinfo("✅", "Usuário atualizado!", parent=d)
            load_users()

        def excluir():
            nm = en_nome.get().strip()

            if not nm or nm == 'admin':
                messagebox.showerror("❌", "Selecione um usuário válido (não pode ser admin).", parent=d)
                return

            if messagebox.askyesno("🗑️", f"Excluir '{nm}'?", parent=d):
                conn = sqlite3.connect(self.db_file)
                cur = conn.cursor()
                cur.execute("DELETE FROM usuarios WHERE nome=?", (nm,))
                conn.commit()
                conn.close()

                messagebox.showinfo("✅", "Excluído!", parent=d)
                en_nome.delete(0, tk.END)
                en_senha.delete(0, tk.END)
                load_users()

        ttk.Button(btn_frame, text="➕ Adicionar", command=adicionar, style='Adicionar.TButton').pack(side='left', padx=5, fill='x', expand=True)
        ttk.Button(btn_frame, text="💾 Salvar", command=salvar, style='Salvar.TButton').pack(side='left', padx=5, fill='x', expand=True)
        ttk.Button(btn_frame, text="🗑️ Excluir", command=excluir, style='Excluir.TButton').pack(side='left', padx=5, fill='x', expand=True)

        load_users()
        d.wait_window(d)

    def alterar_senha_admin(self):
        if not self._checar_permissao('admin', 'Apenas administradores podem alterar a senha do admin.'):
            return

        d = tk.Toplevel(self.root)
        d.title("🔒 Alterar Senha do Admin")
        d.transient(self.root)
        d.grab_set()
        d.resizable(False, False)

        self.centralizar_janela(d)

        f = ttk.Frame(d, padding=15)
        f.pack(fill='both', expand=True)

        ttk.Label(f, text="Senha atual do admin:").grid(row=0, column=0, sticky='w', pady=5)
        ent_atual = ttk.Entry(f, show='*', width=28)
        ent_atual.grid(row=0, column=1, sticky='w', padx=5, pady=5)

        ttk.Label(f, text="Nova senha:").grid(row=1, column=0, sticky='w', pady=5)
        ent_nova = ttk.Entry(f, show='*', width=28)
        ent_nova.grid(row=1, column=1, sticky='w', padx=5, pady=5)

        ttk.Label(f, text="Confirmar nova senha:").grid(row=2, column=0, sticky='w', pady=5)
        ent_conf = ttk.Entry(f, show='*', width=28)
        ent_conf.grid(row=2, column=1, sticky='w', padx=5, pady=5)

        def salvar():
            atual = ent_atual.get()
            nova = ent_nova.get()
            conf = ent_conf.get()

            if not atual or not nova or not conf:
                messagebox.showwarning("⚠️", "Preencha todos os campos.", parent=d)
                return

            if nova != conf:
                messagebox.showerror("❌", "A nova senha e a confirmação não conferem.", parent=d)
                return

            if len(nova) < 4:
                messagebox.showwarning("⚠️", "A nova senha deve ter pelo menos 4 caracteres.", parent=d)
                return

            conn = sqlite3.connect(self.db_file)
            cur = conn.cursor()

            cur.execute("SELECT senha FROM usuarios WHERE nome='admin'")
            row = cur.fetchone()

            if not row:
                cur.execute(
                    """
                    INSERT INTO usuarios
                    (nome, senha, perm_inserir, perm_editar, perm_apagar, perm_exportar,
                     perm_importar, perm_relatorio, perm_admin, perm_emprestados)
                    VALUES (?, ?, 1, 1, 1, 1, 1, 1, 1, 1)
                    """,
                    ('admin', self._hash_senha(nova))
                )
                conn.commit()
                conn.close()

                messagebox.showinfo("✅", "Usuário admin criado com a nova senha.", parent=d)
                d.destroy()
                return

            if not self._verificar_senha(atual, row[0]):
                conn.close()
                messagebox.showerror("❌", "Senha atual do admin incorreta.", parent=d)
                return

            cur.execute(
                "UPDATE usuarios SET senha=? WHERE nome='admin'",
                (self._hash_senha(nova),)
            )

            conn.commit()
            conn.close()

            messagebox.showinfo("✅", "Senha do admin alterada com sucesso.", parent=d)
            d.destroy()

        bf = ttk.Frame(d, padding=10)
        bf.pack(fill='x')

        ttk.Button(bf, text="💾 Salvar", command=salvar, style='Salvar.TButton').pack(side='left', padx=5, expand=True)
        ttk.Button(bf, text="Cancelar", command=d.destroy).pack(side='left', padx=5, expand=True)

        ent_atual.focus_set()
        d.bind("<Return>", lambda e: salvar())

    # ==========================================================================
    # EXPORTAÇÃO
    # ==========================================================================
    def _gerar_arquivo_planilha(self, fmt):
        if not self._checar_permissao('exportar', 'Você não tem permissão para exportar dados.'):
            return

        if fmt == 'xlsx' and not HAS_XLSX:
            messagebox.showerror("📦", "Instale openpyxl")
            return

        if fmt == 'ods' and not HAS_ODS:
            messagebox.showerror("📦", "Instale odfpy")
            return

        if not self.livros:
            messagebox.showwarning("⚠️", "Catálogo vazio.")
            return

        h = [
            'ID', 'Título', 'Autor', 'Estante', 'Prateleira', 'Editora', 'Assunto',
            'Bibliotecário', 'Qtd', 'Disp', 'Emprestado', 'Entrada', 'Saída'
        ]

        if fmt == 'xlsx':
            wb = Workbook()
            ws = wb.active
            ws.title = "Catalogo"
            ws.append(h)

            for l in self.livros:
                ws.append([
                    l['id'], l['titulo'], l['autor'], l.get('estante', ''), l.get('prateleira', ''),
                    l['editora'], l['assunto'], l['bibliotecario'], l['quantidade'], l['disponibilidade'],
                    l['emprestado_a'], l['entrada'], l['saida']
                ])

            n = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])

            if n:
                wb.save(n)
                messagebox.showinfo("✅", f"Exportado em {n}")

        else:
            try:
                doc = OpenDocumentSpreadsheet()
                t = OdfTable(name="Catalogo")

                r = OdfTableRow()

                for x in h:
                    c = OdfTableCell()
                    c.addElement(OdfP(text=str(x)))
                    r.addElement(c)

                t.addElement(r)

                for l in self.livros:
                    r = OdfTableRow()

                    dados = [
                        str(l['id']), l['titulo'], l['autor'], l.get('estante', ''), l.get('prateleira', ''),
                        l.get('editora', ''), l.get('assunto', ''), l.get('bibliotecario', ''),
                        str(l['quantidade']), l['disponibilidade'], l.get('emprestado_a', ''),
                        l.get('entrada', ''), l.get('saida', '')
                    ]

                    for dado in dados:
                        c = OdfTableCell()
                        c.addElement(OdfP(text=str(dado)))
                        r.addElement(c)

                    t.addElement(r)

                doc.spreadsheet.addElement(t)

                n = filedialog.asksaveasfilename(defaultextension=".ods", filetypes=[("ODS", "*.ods")])

                if n:
                    doc.save(n)
                    messagebox.showinfo("✅", f"Exportado em {n}")

            except Exception:
                messagebox.showerror("❌", traceback.format_exc())

    def _pdf_tabela(self, lv):
        hd = [
            'ID', 'Título', 'Autor', 'Estante', 'Prateleira', 'Editora', 'Assunto',
            'Bibliotecário', 'Qtd', 'Disp', 'Emprestado', 'Entrada', 'Saída'
        ]

        dt = []

        for l in lv:
            dt.append([
                str(l['id']), l['titulo'], l['autor'], l.get('estante', ''), l.get('prateleira', ''),
                l['editora'], l['assunto'], l['bibliotecario'], str(l['quantidade']), l['disponibilidade'],
                l['emprestado_a'], l['entrada'], l['saida']
            ])

        tb = PdfTable(
            [hd] + dt,
            colWidths=[
                0.5 * cm, 2.4 * cm, 1.9 * cm, 1.1 * cm, 1.2 * cm, 1.6 * cm, 1.6 * cm,
                1.6 * cm, 0.7 * cm, 0.7 * cm, 2.2 * cm, 1.3 * cm, 1.3 * cm
            ],
            repeatRows=1
        )

        tb.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2)
        ]))

        return tb

    def exportar_pdf(self):
        if not self._checar_permissao('exportar', 'Você não tem permissão para exportar dados.'):
            return

        if not HAS_PDF:
            messagebox.showerror("📦", "Instale reportlab")
            return

        if not self.livros:
            messagebox.showwarning("⚠️", "Vazio.")
            return

        n = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])

        if not n:
            return

        doc = SimpleDocTemplate(
            n,
            pagesize=landscape(A4),
            rightMargin=1.2 * cm,
            leftMargin=1.2 * cm,
            topMargin=1.2 * cm,
            bottomMargin=1.2 * cm
        )

        el = []
        ss = getSampleStyleSheet()

        el.append(Paragraph("FEEU - Catálogo", ss['Heading1']))
        el.append(Spacer(1, 0.5 * cm))
        el.append(self._pdf_tabela(self.livros))

        doc.build(el)

        messagebox.showinfo("✅", f"PDF gerado em {n}")

    def imprimir_catalogo(self, tipo='completo'):
        if not self._checar_permissao('relatorio', 'Você não tem permissão para imprimir relatórios.'):
            return

        if not HAS_PDF:
            messagebox.showerror("📦", "Instale reportlab")
            return

        if tipo == 'busca':
            if not self.busca_ativa or not self.ultima_busca:
                messagebox.showwarning(
                    "⚠️ Nenhuma busca ativa",
                    "O botão '🖨️ Busca' imprime apenas os resultados de uma busca.\n\n"
                    "Faça uma busca primeiro e depois clique em '🖨️ Busca'.\n"
                    "Para imprimir tudo, use '🖨️ Catálogo'."
                )
                return

            lv = self.ultima_busca
        else:
            lv = self.livros

        if not lv:
            messagebox.showwarning("⚠️", "Nada para imprimir.")
            return

        tmp_handle = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf', prefix='Cat_')
        tmp = tmp_handle.name
        tmp_handle.close()

        doc = SimpleDocTemplate(
            tmp,
            pagesize=landscape(A4),
            rightMargin=1.2 * cm,
            leftMargin=1.2 * cm,
            topMargin=1.2 * cm,
            bottomMargin=1.2 * cm
        )

        el = []
        ss = getSampleStyleSheet()

        titulo_doc = "FEEU - Resultados da Busca" if tipo == 'busca' else "FEEU - Catálogo"

        el.append(Paragraph(titulo_doc, ss['Heading1']))
        el.append(Spacer(1, 0.5 * cm))
        el.append(self._pdf_tabela(lv))

        doc.build(el)

        try:
            if os.name == 'nt':
                os.startfile(tmp)
            elif sys.platform == 'darwin':
                subprocess.call(['open', tmp])
            else:
                subprocess.call(['xdg-open', tmp])
        except Exception:
            messagebox.showinfo("ℹ️", f"PDF salvo em: {tmp}")

    # ==========================================================================
    # IMPORTAÇÃO
    # ==========================================================================
    def _campos_alvo_import(self):
        return {
            'titulo': [
                'titulo', 'titulo_do_livro', 'titulo_livro', 'livro', 'obra',
                'nome', 'nome_do_livro', 'nome_livro', 'titulo_obra', 'descricao', 'descricao_livro'
            ],
            'autor': [
                'autor', 'autor_do_livro', 'autor_livro', 'author', 'escritor', 'nome_autor', 'autores'
            ],
            'estante': [
                'estante', 'estantes', 'rack', 'armario', 'modulo', 'shelf', 'local_estante', 'codigo_estante'
            ],
            'prateleira': [
                'prateleira', 'prateleiras', 'shelf_number', 'nivel', 'prateleira_numero', 'local_prateleira'
            ],
            'editora': [
                'editora', 'publisher', 'publicadora', 'casa_editorial', 'editora_livro'
            ],
            'assunto': [
                'assunto', 'categoria', 'genero', 'tema', 'classificacao', 'area', 'assunto_livro'
            ],
            'bibliotecario': [
                'bibliotecario', 'responsavel', 'admin', 'cadastro_por', 'cadastrado_por', 'usuario'
            ],
            'quantidade': [
                'quantidade', 'qtd', 'exemplares', 'qty', 'estoque', 'unidades', 'numero_exemplares', 'qtd_livros'
            ],
            'disponibilidade': [
                'disponibilidade', 'disp', 'status', 'situacao', 'disponivel', 'estado'
            ],
            'emprestado_a': [
                'emprestado_a', 'emprestado', 'para', 'leitor', 'pessoa', 'emprestado_para', 'usuario_emprestimo'
            ],
            'entrada': [
                'entrada', 'data_entrada', 'recebido_em', 'aquisicao', 'data_aquisicao', 'data_recebimento', 'data_cadastro'
            ],
            'saida': [
                'saida', 'data_saida', 'emprestado_em', 'devolucao', 'data_devolucao', 'data_emprestimo'
            ]
        }

    def _norm_header(self, t):
        t = self._normalizar_texto(t)
        t = re.sub(r"[^a-z0-9 ]+", "", t)
        t = re.sub(r"\s+", "_", t).strip("_")
        return t

    def _mapear_colunas_header(self, header):
        cm = {}

        if not header:
            return cm

        campos = self._campos_alvo_import()

        for i, h in enumerate(header):
            if h is None:
                continue

            nn = self._norm_header(str(h))

            if not nn:
                continue

            nn_tokens = set(nn.split('_'))

            best = None
            best_score = 0

            for campo, aliases in campos.items():
                for alias in aliases:
                    alias_n = self._norm_header(alias)

                    if not alias_n:
                        continue

                    alias_tokens = set(alias_n.split('_'))
                    score = 0

                    if nn == alias_n:
                        score = 100
                    elif alias_n in nn:
                        score = 75
                    elif nn in alias_n:
                        score = 65
                    else:
                        common = nn_tokens & alias_tokens
                        if common:
                            score = 35 + len(common) * 5

                    if score > best_score:
                        best_score = score
                        best = campo

            if best and best_score >= 45:
                cm[i] = best

        return cm

    def _detectar_linha_cabecalho(self, rows):
        if not rows:
            return None

        melhor_idx = None
        melhor_qtd = 0
        melhor_cm = {}

        for idx, row in enumerate(rows[:10]):
            cm = self._mapear_colunas_header(row)
            qtd = len(cm)

            if qtd > melhor_qtd:
                melhor_qtd = qtd
                melhor_idx = idx
                melhor_cm = cm

        if melhor_qtd >= 2 and any(v in ('titulo', 'autor') for v in melhor_cm.values()):
            return melhor_idx

        return None

    def _ler_csv(self, caminho):
        all_rows = []

        for enc in ('utf-8-sig', 'latin-1'):
            try:
                with open(caminho, 'r', encoding=enc, newline='') as f:
                    sample = f.read(4096)
                    f.seek(0)

                    if sample.count('\t') > max(sample.count(';'), sample.count(',')):
                        delimiter = '\t'
                    elif sample.count(';') > sample.count(','):
                        delimiter = ';'
                    else:
                        delimiter = ','

                    reader = csv.reader(f, delimiter=delimiter)
                    all_rows = [row for row in reader if any(str(c).strip() for c in row)]
                    break

            except UnicodeDecodeError:
                continue

        if not all_rows:
            return []

        header_idx = self._detectar_linha_cabecalho(all_rows[:10])

        if header_idx is None:
            return []

        cm = self._mapear_colunas_header(all_rows[header_idx])

        if len(cm) < 2:
            return []

        dados = []

        for row in all_rows[header_idx + 1:]:
            if not any(str(c).strip() for c in row):
                continue

            d = {}

            for i, v in enumerate(row):
                if i in cm:
                    d[cm[i]] = str(v).strip()

            if d:
                dados.append(d)

        return dados

    def _ler_xlsx(self, caminho):
        wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
        ws = wb.active

        rows_iter = ws.iter_rows(values_only=True)

        first_rows = []

        for _ in range(10):
            try:
                row = next(rows_iter)
            except StopIteration:
                break
            first_rows.append(row)

        if not first_rows:
            wb.close()
            return []

        header_idx = self._detectar_linha_cabecalho(first_rows)

        if header_idx is None:
            wb.close()
            return []

        cm = self._mapear_colunas_header(first_rows[header_idx])

        if len(cm) < 2:
            wb.close()
            return []

        dados = []

        def process_row(row):
            if row is None:
                return

            if not any(v not in (None, "") for v in row):
                return

            d = {}

            for i, v in enumerate(row):
                if i in cm:
                    d[cm[i]] = "" if v is None else str(v).strip()

            if d:
                dados.append(d)

        for row in first_rows[header_idx + 1:]:
            process_row(row)

        for row in rows_iter:
            process_row(row)

        wb.close()
        return dados

    def _ler_ods(self, caminho):
        doc = odf_load(caminho)
        sh = doc.spreadsheet.getElementsByType(OdfTable)[0]
        rows = sh.getElementsByType(OdfTableRow)

        def get_text(node):
            txt = []

            if hasattr(node, 'data'):
                txt.append(node.data)

            if hasattr(node, 'childNodes'):
                for child in node.childNodes:
                    txt.append(get_text(child))

            return " ".join(txt).strip()

        all_rows = []

        for row in rows:
            vals = [get_text(c) for c in row.childNodes if c.nodeType == c.ELEMENT_NODE]

            if any(str(v).strip() for v in vals):
                all_rows.append(vals)

        if not all_rows:
            return []

        header_idx = self._detectar_linha_cabecalho(all_rows[:10])

        if header_idx is None:
            return []

        cm = self._mapear_colunas_header(all_rows[header_idx])

        if len(cm) < 2:
            return []

        dados = []

        for row in all_rows[header_idx + 1:]:
            if not any(str(v).strip() for v in row):
                continue

            d = {}

            for i, v in enumerate(row):
                if i in cm:
                    d[cm[i]] = str(v).strip()

            if d:
                dados.append(d)

        return dados

    def _gravar_importacao(self, dados):
        conn = sqlite3.connect(self.db_file)
        cur = conn.cursor()

        ins = 0
        att = 0
        ign = 0

        try:
            cols = self._colunas_livros_existentes(cur)

            for l in dados:
                t = str(l.get('titulo', '') or '').strip()
                a = str(l.get('autor', '') or '').strip()

                if not t or not a:
                    ign += 1
                    continue

                e = str(l.get('editora', '') or '').strip()
                ass = str(l.get('assunto', '') or '').strip()
                bib = str(l.get('bibliotecario', '') or '').strip()
                emp = str(l.get('emprestado_a', '') or '').strip()
                est = str(l.get('estante', '') or '').strip()
                prat = str(l.get('prateleira', '') or '').strip()

                q, ok = self._validar_quantidade(l.get('quantidade', '1'))
                if not ok:
                    q = 1

                disp_raw = str(l.get('disponibilidade', '') or '').strip().lower()

                if disp_raw in ('sim', 's', 'disponivel', 'yes', 'true', '1'):
                    disp = 'Sim'
                elif disp_raw in ('nao', 'não', 'n', 'indisponivel', 'no', 'false', '0'):
                    disp = 'Não'
                else:
                    disp = 'Sim' if q > 0 else 'Não'

                entrada_iso = self._data_br_para_iso(l.get('entrada', ''), manter_original=True)
                saida_iso = self._data_br_para_iso(l.get('saida', ''), manter_original=True)

                dup = self._checar_duplicata(t, a, e, ass, est, prat, cursor=cur)

                if dup:
                    lid, qat = dup
                    nq = (qat if qat else 0) + q

                    sets = []
                    params = []

                    if 'quantidade' in cols:
                        sets.append("quantidade=?")
                        params.append(nq)

                    if 'disponibilidade' in cols:
                        sets.append("disponibilidade=?")
                        params.append('Sim' if nq > 0 else 'Não')

                    if 'estante' in cols and est:
                        sets.append("estante=CASE WHEN COALESCE(estante,'')='' THEN ? ELSE estante END")
                        params.append(est)

                    if 'prateleira' in cols and prat:
                        sets.append("prateleira=CASE WHEN COALESCE(prateleira,'')='' THEN ? ELSE prateleira END")
                        params.append(prat)

                    if sets:
                        params.append(lid)
                        cur.execute(f"UPDATE livros SET {', '.join(sets)} WHERE id=?", tuple(params))

                    att += 1

                else:
                    fields = ['titulo', 'autor']
                    values = [t, a]

                    opcionais = [
                        ('editora', e),
                        ('assunto', ass),
                        ('bibliotecario', bib),
                        ('quantidade', q),
                        ('disponibilidade', disp),
                        ('emprestado_a', emp),
                        ('entrada', entrada_iso),
                        ('saida', saida_iso),
                        ('estante', est),
                        ('prateleira', prat),
                    ]

                    for col, val in opcionais:
                        if col in cols:
                            fields.append(col)
                            values.append(val)

                    placeholders = ','.join(['?'] * len(fields))
                    cur.execute(f"INSERT INTO livros ({','.join(fields)}) VALUES ({placeholders})", tuple(values))

                    ins += 1

                self._registrar_historico_livro(
                    {
                        'autor': a,
                        'editora': e,
                        'assunto': ass,
                        'bibliotecario': bib,
                        'emprestado_a': emp,
                        'estante': est,
                        'prateleira': prat
                    },
                    cursor=cur
                )

            conn.commit()

        finally:
            conn.close()

        return ins, att, ign

    def importar_planilha(self):
        if not self._checar_permissao('importar', 'Você não tem permissão para importar dados.'):
            return

        caminho = filedialog.askopenfilename(
            title="Selecionar",
            filetypes=[
                ("Planilhas/CSV", "*.xlsx *.ods *.csv"),
                ("Excel", "*.xlsx"),
                ("Calc", "*.ods"),
                ("CSV", "*.csv")
            ]
        )

        if not caminho:
            return

        ext = os.path.splitext(caminho)[1].lower()

        if ext == '.xlsx' and not HAS_XLSX:
            messagebox.showerror("📦", "Para importar Excel (.xlsx), instale:\n\npip install openpyxl")
            return

        if ext == '.ods' and not HAS_ODS:
            messagebox.showerror("📦", "Para importar Calc (.ods), instale:\n\npip install odfpy")
            return

        try:
            with open(caminho, 'rb') as f:
                pass
        except PermissionError:
            messagebox.showerror(
                "⛔ Erro de Acesso",
                "A planilha está ABERTA em outro programa.\nFeche o Excel/LibreOffice e tente novamente."
            )
            return
        except OSError as e:
            messagebox.showerror("⛔ Erro de Acesso", f"Não foi possível abrir o arquivo:\n{e}")
            return

        self.root.config(cursor="watch")
        self.root.update_idletasks()

        try:
            if ext == '.xlsx':
                dados = self._ler_xlsx(caminho)
            elif ext == '.ods':
                dados = self._ler_ods(caminho)
            elif ext == '.csv':
                dados = self._ler_csv(caminho)
            else:
                messagebox.showwarning("⚠️", "Formato não suportado. Use .xlsx, .ods ou .csv.")
                return

            if not dados:
                messagebox.showwarning(
                    "⚠️",
                    "Nenhum dado reconhecido na planilha.\n\n"
                    "Verifique se a planilha possui cabeçalhos como:\n"
                    "Título, Autor, Estante, Prateleira, Editora, Assunto, Quantidade, Entrada, Saída."
                )
                return

            ins, att, ign = self._gravar_importacao(dados)

            self._recarregar_manter_busca()
            self.atualizar_comboboxes()

            messagebox.showinfo("✅", f"Inseridos: {ins}\nAtualizados: {att}\nIgnorados: {ign}")

        except Exception:
            messagebox.showerror("❌", traceback.format_exc())

        finally:
            self.root.config(cursor="")

    # ==========================================================================
    # AJUDA / SOBRE
    # ==========================================================================
    def mostrar_manual(self):
        manual_window = tk.Toplevel(self.root)
        manual_window.title("📖 Manual")
        manual_window.geometry("650x700")
        manual_window.resizable(False, False)
        manual_window.grab_set()

        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        manual_window.geometry(f"+{(sw - 650) // 2}+{(sh - 700) // 2}")

        container = ttk.Frame(manual_window, padding=20)
        container.pack(fill='both', expand=True)

        ttk.Label(
            container,
            text="📖 Manual do Usuário",
            font=('DejaVu Sans', 16, 'bold'),
            foreground='#1e40af'
        ).pack(pady=(0, 10))

        ttk.Separator(container, orient='horizontal').pack(fill='x', pady=5)

        canvas = tk.Canvas(container, highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        scrollable_frame = ttk.Frame(canvas)

        scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _on_linux_scroll(event):
            if event.num == 4:
                canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                canvas.yview_scroll(1, "units")

        canvas.bind("<Enter>", lambda e: canvas.focus_set())
        canvas.bind("<MouseWheel>", _on_mousewheel)
        canvas.bind("<Button-4>", _on_linux_scroll)
        canvas.bind("<Button-5>", _on_linux_scroll)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        ttk.Label(
            scrollable_frame,
            text="🌐 Contato",
            font=('DejaVu Sans', 12, 'bold'),
            foreground='#1e40af'
        ).pack(anchor='w', pady=(10, 5))

        contatos = [
            ("🏠 Site", "https://www.feeu.org/"),
            ("📘 Facebook", "https://www.facebook.com/FraternidadeEcleticaEspiritualistaUniversal/"),
            ("📷 Instagram", "https://www.instagram.com/feeu.oficial/"),
            ("📺 YouTube", "https://www.youtube.com/channel/UCERzptseTBQ6ijZq4_OxqSQ")
        ]

        for nome, url in contatos:
            f = ttk.Frame(scrollable_frame)
            f.pack(fill='x', pady=2, padx=10)

            lbl = tk.Label(
                f,
                text=f"{nome}: {url}",
                font=('DejaVu Sans', 10),
                foreground='#2563eb',
                cursor='hand2',
                underline=True
            )
            lbl.pack(anchor='w')
            lbl.bind("<Button-1>", lambda e, u=url: webbrowser.open(u))

        ttk.Separator(scrollable_frame, orient='horizontal').pack(fill='x', pady=15)

        sections = [
            ("⌨️ Atalhos", [
                ("Ctrl + A", "Adicionar"),
                ("Ctrl + S", "Salvar"),
                ("Ctrl + E", "Excluir"),
                ("Ctrl + L", "Limpar"),
                ("F1", "Manual")
            ]),
            ("🎨 Visual", [
                ("Linhas e colunas", "A grade usa cores alternadas para facilitar a leitura."),
                ("Tema", "Use o botão 🌙/☀️ para alternar entre tema claro e escuro.")
            ]),
            ("🗄️ Localização", [
                ("Estante/Prateleira", "Campos para facilitar a localização física do livro."),
                ("Botão ➕", "Adiciona rapidamente novos valores de Estante e Prateleira.")
            ]),
            ("🔐 Segurança", [
                ("Login", "Usuário padrão: admin / admin."),
                ("Senha Admin", "Use o botão 🔒 Senha Admin para alterar a senha do administrador.")
            ])
        ]

        for titulo, itens in sections:
            ttk.Label(
                scrollable_frame,
                text=titulo,
                font=('DejaVu Sans', 11, 'bold'),
                foreground='#2563eb'
            ).pack(anchor='w', pady=(15, 5))

            for item_nome, item_desc in itens:
                f_item = ttk.Frame(scrollable_frame)
                f_item.pack(fill='x', pady=2, padx=10)

                ttk.Label(
                    f_item,
                    text=item_nome,
                    width=16,
                    anchor='w',
                    font=('DejaVu Sans', 10, 'bold')
                ).pack(side='left')

                ttk.Label(
                    f_item,
                    text=item_desc,
                    wraplength=460
                ).pack(side='left', fill='x', expand=True)

        ttk.Button(container, text="Entendi, fechar", command=manual_window.destroy, style='Sobre.TButton').pack(pady=15)

    def mostrar_sobre(self):
        sobre_window = tk.Toplevel(self.root)
        sobre_window.title("Sobre")
        sobre_window.resizable(True, True)
        sobre_window.grab_set()
        sobre_window.minsize(380, 350)

        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        sobre_window.geometry(f"+{(sw - 440) // 2}+{(sh - 400) // 2}")

        container = ttk.Frame(sobre_window, padding=20)
        container.pack(fill='both', expand=True)

        sobre_canvas = tk.Canvas(container, width=60, height=60, bg='#ffffff', highlightthickness=1, relief='solid')
        sobre_canvas.pack(pady=(0, 10))

        cx, cy = 30, 30

        sobre_canvas.create_oval(5, 5, 55, 55, outline=self.cores['verde'], width=3)
        sobre_canvas.create_oval(10, 10, 50, 50, outline=self.cores['dourado'], width=2)
        sobre_canvas.create_polygon(cx, cy - 15, cx - 12, cy + 10, cx + 12, cy + 10, outline=self.cores['azul'], width=2, fill='')
        sobre_canvas.create_oval(cx - 2, cy - 2, cx + 2, cy + 2, fill=self.cores['ponto'])

        info_frame = ttk.Frame(container)
        info_frame.pack(fill='x', pady=10, expand=True)

        ttk.Label(info_frame, text="Catálogo de Livros", font=('DejaVu Sans', 14, 'bold'), foreground='#2563eb').pack()
        ttk.Label(info_frame, text="Fraternidade Eclética Espiritualista Universal", font=('DejaVu Sans', 10)).pack(pady=(2, 10))

        ttk.Separator(info_frame, orient='horizontal').pack(fill='x', pady=5)

        infos = [
            ("📅 Criação", datetime.now().strftime('%d/%m/%Y')),
            ("👥 Colaboradores", "Alexandre Borges e Ir. Rodolpho"),
            ("📍 Localidade", "Regional de Petrópolis-RJ"),
            ("💻 Versão", "4.14.0 (Final)"),
            ("🔧 Tecnologia", "Python + Tkinter + SQLite3")
        ]

        for label, value in infos:
            row = ttk.Frame(info_frame)
            row.pack(fill='x', pady=2)

            ttk.Label(row, text=f"{label}: ", font=('DejaVu Sans', 10, 'bold')).pack(side='left', padx=(0, 5))
            ttk.Label(row, text=value, font=('DejaVu Sans', 10)).pack(side='left', fill='x', expand=True)

        ttk.Separator(info_frame, orient='horizontal').pack(fill='x', pady=10)

        ttk.Label(
            info_frame,
            text="SIC TRANSIT GLORIA MUNDI",
            font=('Times New Roman', 12, 'italic'),
            foreground=self.cores['vermelho']
        ).pack(pady=5)

        ttk.Label(
            info_frame,
            text="Desenvolvido pelos Irmãos Rodolpho e Alexandre",
            font=('DejaVu Sans', 9, 'italic'),
            foreground='#64748b'
        ).pack(pady=5)

        ttk.Button(container, text="Fechar", command=sobre_window.destroy, style='Sobre.TButton').pack(pady=10)

    # ==========================================================================
    # RELÓGIO / SAÍDA
    # ==========================================================================
    def confirmar_saida(self):
        if self.clock_job:
            try:
                self.root.after_cancel(self.clock_job)
            except Exception:
                pass

        if messagebox.askyesno("🚪 Sair", "Encerrar o Catálogo?\n✅ Dados já salvos automaticamente."):
            self.root.destroy()

    def atualizar_relogio(self):
        try:
            agora = datetime.now()

            dias = {
                'Monday': 'Segunda-Feira',
                'Tuesday': 'Terça-Feira',
                'Wednesday': 'Quarta-Feira',
                'Thursday': 'Quinta-Feira',
                'Friday': 'Sexta-Feira',
                'Saturday': 'Sábado',
                'Sunday': 'Domingo'
            }

            self.lbl_hora.config(text=agora.strftime('%H:%M'))
            self.lbl_data.config(text=agora.strftime('%d/%m/%Y'))
            self.lbl_dia.config(text=dias.get(agora.strftime('%A'), ''))

        except Exception:
            pass

        finally:
            try:
                self.clock_job = self.root.after(1000, self.atualizar_relogio)
            except Exception:
                pass

# ===================== PATCH FINAL PORTABLE =====================
import csv as _csv_final
import tempfile as _tempfile_final
import webbrowser as _web_final
import traceback as _tb_final
from html import escape as _html_escape_final


def _exportar_csv_final(self, caminho=None):
    if not self.livros:
        messagebox.showwarning("⚠️", "Catálogo vazio.")
        return

    if not caminho:
        caminho = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")]
        )

    if not caminho:
        return

    try:
        with open(caminho, 'w', newline='', encoding='utf-8-sig') as f:
            w = _csv_final.writer(f, delimiter=';')

            w.writerow([
                'ID', 'Título', 'Autor', 'Editora', 'Assunto', 'Bibliotecário',
                'Qtd', 'Disp', 'Emprestado', 'Entrada', 'Saída'
            ])

            for l in self.livros:
                w.writerow([
                    l.get('id', ''),
                    l.get('titulo', ''),
                    l.get('autor', ''),
                    l.get('editora', ''),
                    l.get('assunto', ''),
                    l.get('bibliotecario', ''),
                    l.get('quantidade', ''),
                    l.get('disponibilidade', ''),
                    l.get('emprestado_a', ''),
                    l.get('entrada', ''),
                    l.get('saida', '')
                ])

        messagebox.showinfo("✅", f"CSV exportado em:\n{caminho}")

    except Exception:
        messagebox.showerror("❌", _tb_final.format_exc())


def _exportar_html_final(self, lv, titulo):
    try:
        rows_html = []

        for l in lv:
            cells = ''.join(
                f'<td>{_html_escape_final(str(l.get(k, "") or ""))}</td>'
                for k in [
                    'id', 'titulo', 'autor', 'editora', 'assunto',
                    'bibliotecario', 'quantidade', 'disponibilidade',
                    'emprestado_a', 'entrada', 'saida'
                ]
            )
            rows_html.append(f'<tr>{cells}</tr>')

        html = f'''
<html>
<head>
<meta charset="utf-8">
<title>{_html_escape_final(titulo)}</title>
<style>
body {{ font-family: Arial; font-size: 11px; }}
table {{ border-collapse: collapse; width: 100%; }}
th, td {{ border: 1px solid #cbd5e1; padding: 4px; }}
th {{ background: #2563eb; color: white; }}
tr:nth-child(even) {{ background: #f8fafc; }}
</style>
</head>
<body>
<h2>{_html_escape_final(titulo)}</h2>
<table>
<tr>
<th>ID</th><th>Título</th><th>Autor</th><th>Editora</th><th>Assunto</th>
<th>Bibliotecário</th><th>Qtd</th><th>Disp</th><th>Emprestado</th>
<th>Entrada</th><th>Saída</th>
</tr>
{''.join(rows_html)}
</table>
</body>
</html>
'''

        tmp = _tempfile_final.NamedTemporaryFile(
            delete=False,
            suffix='.html',
            prefix='Cat_',
            mode='w',
            encoding='utf-8'
        )
        tmp.write(html)
        tmp.close()

        _web_final.open('file:///' + tmp.name.replace('\\', '/'))

        messagebox.showinfo(
            "ℹ️",
            "Relatório HTML aberto.\n\nUse Ctrl+P no navegador e escolha:\n'Salvar como PDF'."
        )

    except Exception:
        messagebox.showerror("❌", _tb_final.format_exc())


def _gerar_arquivo_planilha_final(self, fmt):
    if not self.livros:
        messagebox.showwarning("⚠️", "Catálogo vazio.")
        return

    h = [
        'ID', 'Título', 'Autor', 'Editora', 'Assunto', 'Bibliotecário',
        'Qtd', 'Disp', 'Emprestado', 'Entrada', 'Saída'
    ]

    if fmt == 'xlsx':
        if HAS_XLSX:
            wb = Workbook()
            ws = wb.active
            ws.title = "Catalogo"
            ws.append(h)

            for l in self.livros:
                ws.append([
                    l.get('id', ''),
                    l.get('titulo', ''),
                    l.get('autor', ''),
                    l.get('editora', ''),
                    l.get('assunto', ''),
                    l.get('bibliotecario', ''),
                    l.get('quantidade', ''),
                    l.get('disponibilidade', ''),
                    l.get('emprestado_a', ''),
                    l.get('entrada', ''),
                    l.get('saida', '')
                ])

            n = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")]
            )

            if n:
                wb.save(n)
                messagebox.showinfo("✅", f"Exportado em {n}")

        else:
            if messagebox.askyesno(
                "📦 Biblioteca ausente",
                "openpyxl não está instalado.\n\nDeseja exportar em CSV compatível com Excel?"
            ):
                self._exportar_csv()

        return

    if fmt == 'ods':
        if HAS_ODS:
            try:
                from odf.opendocument import OpenDocumentSpreadsheet as _ODS_Doc
                from odf.table import Table as _ODS_Table
                from odf.table import TableRow as _ODS_Row
                from odf.table import TableCell as _ODS_Cell
                from odf.text import P as _ODS_P

                doc = _ODS_Doc()
                t = _ODS_Table(name="Catalogo")

                r = _ODS_Row()
                for x in h:
                    c = _ODS_Cell()
                    c.addElement(_ODS_P(text=str(x)))
                    r.addElement(c)
                t.addElement(r)

                for l in self.livros:
                    r = _ODS_Row()

                    dados = [
                        str(l.get('id', '')),
                        l.get('titulo', ''),
                        l.get('autor', ''),
                        l.get('editora', ''),
                        l.get('assunto', ''),
                        l.get('bibliotecario', ''),
                        str(l.get('quantidade', '')),
                        l.get('disponibilidade', ''),
                        l.get('emprestado_a', ''),
                        l.get('entrada', ''),
                        l.get('saida', '')
                    ]

                    for d in dados:
                        c = _ODS_Cell()
                        c.addElement(_ODS_P(text=str(d)))
                        r.addElement(c)

                    t.addElement(r)

                doc.spreadsheet.addElement(t)

                n = filedialog.asksaveasfilename(
                    defaultextension=".ods",
                    filetypes=[("ODS", "*.ods")]
                )

                if n:
                    doc.save(n)
                    messagebox.showinfo("✅", f"Exportado em {n}")

            except Exception:
                messagebox.showerror("❌", _tb_final.format_exc())

        else:
            if messagebox.askyesno(
                "📦 Biblioteca ausente",
                "odfpy não está instalado.\n\nDeseja exportar em CSV compatível com Calc?"
            ):
                self._exportar_csv()

        return


def _exportar_pdf_final(self):
    if not self.livros:
        messagebox.showwarning("⚠️", "Catálogo vazio.")
        return

    if HAS_PDF:
        n = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")]
        )

        if not n:
            return

        doc = SimpleDocTemplate(
            n,
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm
        )

        el = []
        ss = getSampleStyleSheet()

        el.append(Paragraph("FEEU - Catálogo", ss['Heading1']))
        el.append(Spacer(1, 0.5 * cm))

        hd = [
            'ID', 'Título', 'Autor', 'Editora', 'Assunto', 'Bibliotecário',
            'Qtd', 'Disp', 'Emprestado', 'Entrada', 'Saída'
        ]

        dt = [[
            str(l.get('id', '')),
            l.get('titulo', ''),
            l.get('autor', ''),
            l.get('editora', ''),
            l.get('assunto', ''),
            l.get('bibliotecario', ''),
            str(l.get('quantidade', '')),
            l.get('disponibilidade', ''),
            l.get('emprestado_a', ''),
            l.get('entrada', ''),
            l.get('saida', '')
        ] for l in self.livros]

        tb = Table(
            [hd] + dt,
            colWidths=[
                0.6 * cm, 2.5 * cm, 2 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm,
                0.8 * cm, 0.8 * cm, 2.5 * cm, 1.5 * cm, 1.5 * cm
            ],
            repeatRows=1
        )

        tb.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2)
        ]))

        el.append(tb)
        doc.build(el)

        messagebox.showinfo("✅", f"PDF gerado em {n}")

    else:
        self._exportar_html(self.livros, "FEEU - Catálogo")


def _imprimir_catalogo_final(self, tipo='completo'):
    lv = self.livros

    if tipo == 'busca':
        if not self.busca_var.get().strip():
            messagebox.showwarning(
                "⚠️ Nenhuma busca ativa",
                "Faça uma busca antes de usar '🖨️ Busca'.\n\nPara imprimir tudo, use '🖨️ Catálogo'."
            )
            return

        t = self.busca_var.get().strip().lower()
        f = self.filtro_var.get()

        m = {
            'Título': 'titulo',
            'Autor': 'autor',
            'Editora': 'editora',
            'Assunto': 'assunto',
            'Bibliotecário': 'bibliotecario',
            'Emprestado a': 'emprestado_a'
        }

        if f == 'Todos':
            lv = [l for l in self.livros if any(t in str(l.get(c, '')).lower() for c in m.values())]
        else:
            campo = m.get(f, 'titulo')
            lv = [l for l in self.livros if t in str(l.get(campo, '')).lower()]

    if not lv:
        messagebox.showwarning("⚠️", "Nada para imprimir.")
        return

    if HAS_PDF:
        tmp_handle = _tempfile_final.NamedTemporaryFile(delete=False, suffix='.pdf', prefix='Cat_')
        tmp = tmp_handle.name
        tmp_handle.close()

        doc = SimpleDocTemplate(
            tmp,
            pagesize=A4,
            rightMargin=1.5 * cm,
            leftMargin=1.5 * cm,
            topMargin=1.5 * cm,
            bottomMargin=1.5 * cm
        )

        el = []
        ss = getSampleStyleSheet()

        titulo_doc = "FEEU - Resultados da Busca" if tipo == 'busca' else "FEEU - Catálogo"

        el.append(Paragraph(titulo_doc, ss['Heading1']))
        el.append(Spacer(1, 0.5 * cm))

        hd = [
            'ID', 'Título', 'Autor', 'Editora', 'Assunto', 'Bibliotecário',
            'Qtd', 'Disp', 'Emprestado', 'Entrada', 'Saída'
        ]

        dt = [[
            str(l.get('id', '')),
            l.get('titulo', ''),
            l.get('autor', ''),
            l.get('editora', ''),
            l.get('assunto', ''),
            l.get('bibliotecario', ''),
            str(l.get('quantidade', '')),
            l.get('disponibilidade', ''),
            l.get('emprestado_a', ''),
            l.get('entrada', ''),
            l.get('saida', '')
        ] for l in lv]

        tb = Table(
            [hd] + dt,
            colWidths=[
                0.6 * cm, 2.5 * cm, 2 * cm, 1.8 * cm, 1.8 * cm, 1.8 * cm,
                0.8 * cm, 0.8 * cm, 2.5 * cm, 1.5 * cm, 1.5 * cm
            ],
            repeatRows=1
        )

        tb.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2)
        ]))

        el.append(tb)
        doc.build(el)

        try:
            if os.name == 'nt':
                os.startfile(tmp)
            elif sys.platform == 'darwin':
                subprocess.call(['open', tmp])
            else:
                subprocess.call(['xdg-open', tmp])
        except Exception:
            messagebox.showinfo("ℹ️", f"PDF salvo em: {tmp}")

    else:
        titulo_html = "FEEU - Resultados da Busca" if tipo == 'busca' else "FEEU - Catálogo"
        self._exportar_html(lv, titulo_html)


def _importar_planilha_final(self):
    caminho = filedialog.askopenfilename(
        title="Selecionar",
        filetypes=[
            ("Planilhas/CSV", "*.xlsx *.ods *.csv"),
            ("Excel", "*.xlsx"),
            ("Calc", "*.ods"),
            ("CSV", "*.csv")
        ]
    )

    if not caminho:
        return

    ext = os.path.splitext(caminho)[1].lower()

    if ext == '.xlsx' and not HAS_XLSX:
        if messagebox.askyesno(
            "📦 Biblioteca ausente",
            "openpyxl não está instalado.\n\nDeseja selecionar um arquivo CSV para importar?"
        ):
            caminho = filedialog.askopenfilename(
                title="Selecionar CSV",
                filetypes=[("CSV", "*.csv")]
            )
            if not caminho:
                return
            ext = '.csv'
        else:
            return

    if ext == '.ods' and not HAS_ODS:
        if messagebox.askyesno(
            "📦 Biblioteca ausente",
            "odfpy não está instalado.\n\nDeseja selecionar um arquivo CSV para importar?"
        ):
            caminho = filedialog.askopenfilename(
                title="Selecionar CSV",
                filetypes=[("CSV", "*.csv")]
            )
            if not caminho:
                return
            ext = '.csv'
        else:
            return

    try:
        with open(caminho, 'rb') as f:
            pass
    except PermissionError:
        messagebox.showerror(
            "⛔ Erro de Acesso",
            "A planilha está ABERTA em outro programa.\nFeche o Excel/LibreOffice e tente novamente."
        )
        return
    except OSError as e:
        messagebox.showerror("⛔ Erro de Acesso", f"Não foi possível abrir o arquivo:\n{e}")
        return

    self.root.config(cursor="watch")
    self.root.update_idletasks()

    try:
        def norm(t):
            t = unicodedata.normalize('NFD', str(t).lower())
            t = ''.join(c for c in t if unicodedata.category(c) != 'Mn')
            return t.strip().replace(' ', '_')

        campos_alvo = {
            'titulo': ['titulo', 'titulo_do_livro', 'titulo_livro', 'livro', 'obra', 'nome', 'nome_do_livro'],
            'autor': ['autor', 'autor_do_livro', 'autor_livro', 'author', 'escritor', 'nome_autor', 'autores'],
            'editora': ['editora', 'publisher', 'publicadora', 'casa_editorial'],
            'assunto': ['assunto', 'categoria', 'genero', 'tema', 'classificacao', 'area'],
            'bibliotecario': ['bibliotecario', 'responsavel', 'admin', 'cadastro_por', 'cadastrado_por'],
            'quantidade': ['quantidade', 'qtd', 'exemplares', 'qty', 'estoque', 'unidades'],
            'disponibilidade': ['disponibilidade', 'disp', 'status', 'situacao', 'disponivel', 'estado'],
            'emprestado_a': ['emprestado_a', 'emprestado', 'para', 'leitor', 'pessoa', 'emprestado_para'],
            'entrada': ['entrada', 'data_entrada', 'recebido_em', 'aquisicao', 'data_aquisicao'],
            'saida': ['saida', 'data_saida', 'emprestado_em', 'devolucao', 'data_devolucao']
        }

        def map_header(header):
            cm = {}
            if not header:
                return cm

            for i, h in enumerate(header):
                if h is None:
                    continue

                nn = norm(str(h))
                if not nn:
                    continue

                nn_tokens = set(nn.split('_'))
                best = None
                best_score = 0

                for campo, aliases in campos_alvo.items():
                    for alias in aliases:
                        an = norm(alias)
                        if not an:
                            continue

                        an_tokens = set(an.split('_'))
                        score = 0

                        if nn == an:
                            score = 100
                        elif an in nn:
                            score = 75
                        elif nn in an:
                            score = 65
                        else:
                            common = nn_tokens & an_tokens
                            if common:
                                score = 35 + len(common) * 5

                        if score > best_score:
                            best_score = score
                            best = campo

                if best and best_score >= 45:
                    cm[i] = best

            return cm

        def detect_header(rows):
            best_idx = None
            best_q = 0
            best_cm = {}

            for idx, row in enumerate(rows[:10]):
                cm = map_header(row)
                q = len(cm)

                if q > best_q:
                    best_q = q
                    best_idx = idx
                    best_cm = cm

            if best_q >= 2 and any(v in ('titulo', 'autor') for v in best_cm.values()):
                return best_idx

            return None

        dados = []

        if ext == '.csv':
            all_rows = []

            for enc in ('utf-8-sig', 'latin-1'):
                try:
                    with open(caminho, 'r', encoding=enc, newline='') as f:
                        sample = f.read(4096)
                        f.seek(0)

                        if sample.count('\t') > max(sample.count(';'), sample.count(',')):
                            delim = '\t'
                        elif sample.count(';') > sample.count(','):
                            delim = ';'
                        else:
                            delim = ','

                        reader = _csv_final.reader(f, delimiter=delim)
                        all_rows = [row for row in reader if any(str(c).strip() for c in row)]
                        break

                except UnicodeDecodeError:
                    continue

            if all_rows:
                header_idx = detect_header(all_rows[:10])

                if header_idx is None:
                    header_idx = 0

                cm = map_header(all_rows[header_idx])

                if len(cm) >= 2:
                    for row in all_rows[header_idx + 1:]:
                        if not any(str(c).strip() for c in row):
                            continue

                        d = {}

                        for i, v in enumerate(row):
                            if i in cm:
                                d[cm[i]] = str(v).strip()

                        if d:
                            dados.append(d)

        elif ext == '.xlsx':
            wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
            ws = wb.active

            rows_iter = ws.iter_rows(values_only=True)
            first_rows = []

            for _ in range(10):
                try:
                    row = next(rows_iter)
                except StopIteration:
                    break
                first_rows.append(row)

            if first_rows:
                header_idx = detect_header(first_rows)

                if header_idx is None:
                    header_idx = 0

                cm = map_header(first_rows[header_idx])

                if len(cm) >= 2:
                    def process_row(row):
                        if row is None:
                            return

                        if not any(v not in (None, '') for v in row):
                            return

                        d = {}

                        for i, v in enumerate(row):
                            if i in cm:
                                d[cm[i]] = '' if v is None else str(v).strip()

                        if d:
                            dados.append(d)

                    for row in first_rows[header_idx + 1:]:
                        process_row(row)

                    for row in rows_iter:
                        process_row(row)

            wb.close()

        elif ext == '.ods':
            from odf.opendocument import load as _odf_load_final
            from odf.table import Table as _ODS_TableFinal
            from odf.table import TableRow as _ODS_RowFinal

            def get_text(node):
                txt = []

                if hasattr(node, 'data'):
                    txt.append(node.data)

                if hasattr(node, 'childNodes'):
                    for child in node.childNodes:
                        txt.append(get_text(child))

                return ' '.join(txt).strip()

            doc = _odf_load_final(caminho)
            sh = doc.spreadsheet.getElementsByType(_ODS_TableFinal)[0]
            rows = sh.getElementsByType(_ODS_RowFinal)

            all_rows = []

            for row in rows:
                vals = [get_text(c) for c in row.childNodes if c.nodeType == c.ELEMENT_NODE]

                if any(str(v).strip() for v in vals):
                    all_rows.append(vals)

            if all_rows:
                header_idx = detect_header(all_rows[:10])

                if header_idx is None:
                    header_idx = 0

                cm = map_header(all_rows[header_idx])

                if len(cm) >= 2:
                    for row in all_rows[header_idx + 1:]:
                        if not any(str(v).strip() for v in row):
                            continue

                        d = {}

                        for i, v in enumerate(row):
                            if i in cm:
                                d[cm[i]] = str(v).strip()

                        if d:
                            dados.append(d)

        else:
            messagebox.showwarning("⚠️", "Formato não suportado. Use .xlsx, .ods ou .csv.")
            return

        if not dados:
            messagebox.showwarning(
                "⚠️",
                "Nenhum dado reconhecido na planilha.\n\n"
                "Verifique se a primeira linha contém cabeçalhos como:\n"
                "Título, Autor, Editora, Assunto, Quantidade, Entrada, Saída."
            )
            return

        conn = sqlite3.connect(self.db_file)
        cur = conn.cursor()

        cur.execute("PRAGMA table_info(livros)")
        cols = {r[1] for r in cur.fetchall()}

        ins = 0
        att = 0
        ign = 0

        for l in dados:
            t = str(l.get('titulo', '') or '').strip()
            a = str(l.get('autor', '') or '').strip()

            if not t or not a:
                ign += 1
                continue

            e = str(l.get('editora', '') or '').strip()
            ass = str(l.get('assunto', '') or '').strip()
            bib = str(l.get('bibliotecario', '') or '').strip()
            emp = str(l.get('emprestado_a', '') or '').strip()
            est = str(l.get('estante', '') or '').strip()
            prat = str(l.get('prateleira', '') or '').strip()

            try:
                q = int(float(str(l.get('quantidade', '1')).strip().replace(',', '.')))
                if q < 0:
                    q = 1
            except Exception:
                q = 1

            disp_raw = str(l.get('disponibilidade', '') or '').strip().lower()

            if disp_raw in ('sim', 's', 'disponivel', 'yes', 'true', '1'):
                disp = 'Sim'
            elif disp_raw in ('nao', 'não', 'n', 'indisponivel', 'no', 'false', '0'):
                disp = 'Não'
            else:
                disp = 'Sim' if q > 0 else 'Não'

            entrada = str(l.get('entrada', '') or '').strip()
            saida = str(l.get('saida', '') or '').strip()

            sql = "SELECT id, quantidade FROM livros WHERE LOWER(titulo)=? AND LOWER(autor)=?"
            params = [t.lower(), a.lower()]

            if 'editora' in cols and e:
                sql += " AND IFNULL(LOWER(editora),'')=?"
                params.append(e.lower())

            if 'assunto' in cols and ass:
                sql += " AND IFNULL(LOWER(assunto),'')=?"
                params.append(ass.lower())

            if 'estante' in cols and est:
                sql += " AND IFNULL(LOWER(estante),'')=?"
                params.append(est.lower())

            if 'prateleira' in cols and prat:
                sql += " AND IFNULL(LOWER(prateleira),'')=?"
                params.append(prat.lower())

            cur.execute(sql, tuple(params))
            dup = cur.fetchone()

            if dup:
                lid, qat = dup
                nq = (qat if qat else 0) + q

                sets = ["quantidade=?", "disponibilidade=?"]
                params2 = [nq, 'Sim' if nq > 0 else 'Não']

                if 'estante' in cols and est:
                    sets.append("estante=CASE WHEN COALESCE(estante,'')='' THEN ? ELSE estante END")
                    params2.append(est)

                if 'prateleira' in cols and prat:
                    sets.append("prateleira=CASE WHEN COALESCE(prateleira,'')='' THEN ? ELSE prateleira END")
                    params2.append(prat)

                params2.append(lid)

                cur.execute(f"UPDATE livros SET {', '.join(sets)} WHERE id=?", tuple(params2))

                att += 1

            else:
                fields = ['titulo', 'autor']
                values = [t, a]

                opcionais = [
                    ('editora', e),
                    ('assunto', ass),
                    ('bibliotecario', bib),
                    ('quantidade', q),
                    ('disponibilidade', disp),
                    ('emprestado_a', emp),
                    ('entrada', entrada),
                    ('saida', saida),
                    ('estante', est),
                    ('prateleira', prat),
                ]

                for col, val in opcionais:
                    if col in cols:
                        fields.append(col)
                        values.append(val)

                placeholders = ','.join(['?'] * len(fields))

                cur.execute(
                    f"INSERT INTO livros ({','.join(fields)}) VALUES ({placeholders})",
                    tuple(values)
                )

                ins += 1

            for k in ['autor', 'editora', 'assunto', 'bibliotecario', 'emprestado_a', 'estante', 'prateleira']:
                val = str(l.get(k, '') or '').strip()

                if not val:
                    continue

                if k == 'emprestado_a':
                    for nome in val.split(';'):
                        nome = nome.strip()
                        if nome:
                            cur.execute(
                                "INSERT OR IGNORE INTO historico_campos (campo, valor) VALUES (?, ?)",
                                (k, nome)
                            )
                else:
                    cur.execute(
                        "INSERT OR IGNORE INTO historico_campos (campo, valor) VALUES (?, ?)",
                        (k, val)
                    )

        conn.commit()
        conn.close()

        self.carregar_dados()
        self.atualizar_comboboxes()

        messagebox.showinfo("✅", f"Inseridos: {ins}\nAtualizados: {att}\nIgnorados: {ign}")

    except Exception:
        messagebox.showerror("❌", _tb_final.format_exc())

    finally:
        self.root.config(cursor="")


# Substitui os métodos originais pelos métodos portáveis
LivroCatalogApp._exportar_csv = _exportar_csv_final
LivroCatalogApp._exportar_html = _exportar_html_final
LivroCatalogApp._gerar_arquivo_planilha = _gerar_arquivo_planilha_final
LivroCatalogApp.exportar_pdf = _exportar_pdf_final
LivroCatalogApp.imprimir_catalogo = _imprimir_catalogo_final
LivroCatalogApp.importar_planilha = _importar_planilha_final
# ==================================================================

# =====================================================================
# PATCH DE CORREÇÃO — PDF / IMPRIMIR (v2)
# Corrige:
#   1) NameError: name 'Table' is not defined  (import local do reportlab)
#   2) "Nada para imprimir" na busca           (lê direto do painel de resultados)
# Cola este bloco ANTES de:  if __name__ == "__main__":
# =====================================================================
import tempfile as _pdf_tempfile
import subprocess as _pdf_subprocess
import traceback as _pdf_traceback


def _pdf_html_escape(s):
    return (str(s).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'))


def _pdf_html_fallback(lv, titulo):
    """Se o reportlab não estiver disponível, abre um HTML pronto para Ctrl+P -> Salvar como PDF."""
    try:
        ordem = ['id', 'titulo', 'autor', 'estante', 'prateleira', 'editora', 'assunto',
                 'bibliotecario', 'quantidade', 'disponibilidade', 'emprestado_a', 'entrada', 'saida']
        rot = {'id': 'ID', 'titulo': 'Título', 'autor': 'Autor', 'estante': 'Estante',
               'prateleira': 'Prateleira', 'editora': 'Editora', 'assunto': 'Assunto',
               'bibliotecario': 'Bibliotecário', 'quantidade': 'Qtd', 'disponibilidade': 'Disp',
               'emprestado_a': 'Emprestado', 'entrada': 'Entrada', 'saida': 'Saída'}
        cols = [c for c in ordem if any((c in l) for l in lv)] or ordem

        css = ("body{font-family:Arial,sans-serif;font-size:11px;}"
               "table{border-collapse:collapse;width:100%;}"
               "th,td{border:1px solid #cbd5e1;padding:4px;}"
               "th{background:#2563eb;color:#fff;}"
               "tr:nth-child(even){background:#f1f5f9;}")

        head = '<tr>' + ''.join('<th>' + _pdf_html_escape(rot.get(c, c)) + '</th>' for c in cols) + '</tr>'
        body = ''
        for l in lv:
            body += '<tr>' + ''.join('<td>' + _pdf_html_escape(l.get(c, '')) + '</td>' for c in cols) + '</tr>'

        html = ('<html><head><meta charset="utf-8"><title>' + _pdf_html_escape(titulo) +
                '</title><style>' + css + '</style></head><body><h2>' + _pdf_html_escape(titulo) +
                '</h2><table>' + head + body + '</table></body></html>')

        tmp = _pdf_tempfile.NamedTemporaryFile(delete=False, suffix='.html', prefix='Cat_',
                                               mode='w', encoding='utf-8')
        tmp.write(html)
        tmp.close()

        webbrowser.open('file:///' + tmp.name.replace('\\', '/'))
        messagebox.showinfo("ℹ️", "Relatório aberto no navegador.\n\nUse Ctrl+P e escolha 'Salvar como PDF'.")
    except Exception:
        messagebox.showerror("❌", _pdf_traceback.format_exc())


def _pdf_montar_tabela(lv):
    """Monta a tabela do PDF com import LOCAL do reportlab (resolve o NameError)."""
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm

    ordem = ['id', 'titulo', 'autor', 'estante', 'prateleira', 'editora', 'assunto',
             'bibliotecario', 'quantidade', 'disponibilidade', 'emprestado_a', 'entrada', 'saida']
    rot = {'id': 'ID', 'titulo': 'Título', 'autor': 'Autor', 'estante': 'Estante',
           'prateleira': 'Prateleira', 'editora': 'Editora', 'assunto': 'Assunto',
           'bibliotecario': 'Bibliotecário', 'quantidade': 'Qtd', 'disponibilidade': 'Disp',
           'emprestado_a': 'Emprestado', 'entrada': 'Entrada', 'saida': 'Saída'}
    larg = {'id': 0.8, 'titulo': 4.0, 'autor': 3.0, 'estante': 1.3, 'prateleira': 1.4,
            'editora': 2.2, 'assunto': 2.2, 'bibliotecario': 2.2, 'quantidade': 0.9,
            'disponibilidade': 1.0, 'emprestado_a': 3.0, 'entrada': 1.6, 'saida': 1.6}

    # Usa só as colunas que realmente existem nos dados (funciona com ou sem Estante/Prateleira)
    cols = [c for c in ordem if any((c in l) for l in lv)] or ordem

    hd = [rot.get(c, c) for c in cols]
    dt = [[('' if l.get(c) is None else str(l.get(c, ''))) for c in cols] for l in lv]
    widths = [larg.get(c, 1.5) * cm for c in cols]

    tb = Table([hd] + dt, colWidths=widths, repeatRows=1)
    tb.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    return tb


def _pdf_ler_resultados(self):
    """Lê EXATAMENTE o que está no painel 'Resultados Encontrados'."""
    tree = getattr(self, 'resultados_tree', None)
    if tree is None:
        return []
    try:
        col_names = tree.cget('columns')
        if isinstance(col_names, str):
            col_names = col_names.split()
        col_names = list(col_names)
    except Exception:
        col_names = []

    out = []
    try:
        for iid in tree.get_children():
            vals = tree.item(iid)['values']
            d = {}
            for i, name in enumerate(col_names):
                d[name] = vals[i] if i < len(vals) else ''
            out.append(d)
    except Exception:
        return []
    return out


def _pdf_abrir(caminho):
    try:
        if os.name == 'nt':
            os.startfile(caminho)
        elif sys.platform == 'darwin':
            _pdf_subprocess.call(['open', caminho])
        else:
            _pdf_subprocess.call(['xdg-open', caminho])
    except Exception:
        messagebox.showinfo("ℹ️", "PDF salvo em:\n" + caminho)


def _pdf_exportar_v2(self):
    """Exporta TODO o catálogo para PDF."""
    if not self.livros:
        messagebox.showwarning("⚠️", "Catálogo vazio.")
        return

    if not HAS_PDF:
        _pdf_html_fallback(self.livros, "FEEU - Catálogo")
        return

    try:
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm

        n = filedialog.asksaveasfilename(defaultextension=".pdf", filetypes=[("PDF", "*.pdf")])
        if not n:
            return

        doc = SimpleDocTemplate(n, pagesize=landscape(A4),
                                rightMargin=1.2 * cm, leftMargin=1.2 * cm,
                                topMargin=1.2 * cm, bottomMargin=1.2 * cm)
        el = [Paragraph("FEEU - Catálogo", getSampleStyleSheet()['Heading1']),
              Spacer(1, 0.4 * cm),
              _pdf_montar_tabela(self.livros)]
        doc.build(el)

        messagebox.showinfo("✅", "PDF gerado em:\n" + n)
    except Exception:
        messagebox.showerror("❌", _pdf_traceback.format_exc())


def _pdf_imprimir_v2(self, tipo='completo'):
    """Imprime o catálogo completo OU o resultado da busca (lido do painel)."""
    if tipo == 'busca':
        if not getattr(self, 'busca_ativa', False):
            messagebox.showwarning(
                "⚠️ Nenhuma busca ativa",
                "Faça uma busca primeiro (🔍 Buscar) e depois clique em '🖨️ Busca'.\n"
                "Para imprimir tudo, use '🖨️ Catálogo'.")
            return
        lv = _pdf_ler_resultados(self)
        titulo = "FEEU - Resultados da Busca"
    else:
        lv = self.livros
        titulo = "FEEU - Catálogo"

    if not lv:
        messagebox.showwarning("⚠️", "Nada para imprimir.")
        return

    if not HAS_PDF:
        _pdf_html_fallback(lv, titulo)
        return

    try:
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm

        tmp = _pdf_tempfile.NamedTemporaryFile(delete=False, suffix='.pdf', prefix='Cat_').name

        doc = SimpleDocTemplate(tmp, pagesize=landscape(A4),
                                rightMargin=1.2 * cm, leftMargin=1.2 * cm,
                                topMargin=1.2 * cm, bottomMargin=1.2 * cm)
        el = [Paragraph(titulo, getSampleStyleSheet()['Heading1']),
              Spacer(1, 0.4 * cm),
              _pdf_montar_tabela(lv)]
        doc.build(el)

        _pdf_abrir(tmp)
    except Exception:
        messagebox.showerror("❌", _pdf_traceback.format_exc())


# Reatribui os métodos (a última definição vence, sobrescrevendo patches antigos)
LivroCatalogApp.exportar_pdf = _pdf_exportar_v2
LivroCatalogApp.imprimir_catalogo = _pdf_imprimir_v2
# =====================================================================

# =====================================================================
# PATCH v3 — CABEÇALHO AZUL NO EXCEL/CALC  +  IMPRESSÃO DA BUSCA
# - Excel (.xlsx) e Calc (.ods): pinta a linha de título de azul (branco/negrito)
# - Imprimir Catálogo: imprime EXATAMENTE a grade principal
# - Imprimir Busca:    imprime EXATAMENTE o painel "Resultados Encontrados"
# Cola ANTES de:  if __name__ == "__main__":
# =====================================================================
import tempfile as _v3_temp
import subprocess as _v3_sub
import traceback as _v3_tb

# Larguras (cm) por campo, para o PDF caber bem em paisagem
_V3_LARG = {
    'id': 0.9, 'titulo': 4.2, 'autor': 3.2, 'estante': 1.4, 'prateleira': 1.5,
    'editora': 2.4, 'assunto': 2.4, 'bibliotecario': 2.4, 'quantidade': 1.0,
    'disponibilidade': 1.2, 'emprestado_a': 3.2, 'entrada': 1.7, 'saida': 1.7,
}


def _v3_ler_grid(tree):
    """Lê cabeçalho + linhas de uma grade, seja Treeview ou tabela custom.
    Retorna (chaves, cabecalhos, linhas). linhas = lista de listas de str."""
    chaves, cabecalhos, linhas = [], [], []
    try:
        # --- tabela custom (minha classe TabelaCatalogo) ---
        if hasattr(tree, 'data') and hasattr(tree, 'visible_cols'):
            vcols = tree.visible_cols or tree.columns
            chaves = [c['key'] for c in vcols]
            cabecalhos = [c['text'] for c in vcols]
            for row in tree.data:
                linhas.append(['' if row.get(k) is None else str(row.get(k)) for k in chaves])
            return chaves, cabecalhos, linhas

        # --- Treeview padrão do Tkinter ---
        if hasattr(tree, 'get_children'):
            cols = tree.cget('columns')
            if isinstance(cols, str):
                cols = cols.split()
            cols = list(cols)
            chaves = cols
            cabecalhos = [str(tree.heading(c)['text']) for c in cols]
            for iid in tree.get_children():
                vals = tree.item(iid)['values']
                linhas.append([str(vals[i]) if i < len(vals) else '' for i in range(len(cols))])
            return chaves, cabecalhos, linhas
    except Exception:
        pass
    return chaves, cabecalhos, linhas


def _v3_montar_pdf(chaves, cabecalhos, linhas):
    """Monta a tabela do PDF com IMPORT LOCAL do reportlab (evita NameError)."""
    from reportlab.platypus import Table, TableStyle
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4, landscape

    # larguras por chave, com ajuste automático para caber na página
    widths = [_V3_LARG.get(k, 1.8) for k in chaves]
    pw, ph = landscape(A4)
    util = (pw - 1.6 * cm) / cm          # largura útil em cm (margens 0,8cm)
    total = sum(widths) or 1.0
    if total > util:
        f = util / total
        widths = [w * f for w in widths]

    tb = Table([cabecalhos] + linhas, colWidths=[w * cm for w in widths], repeatRows=1)
    tb.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563eb')),   # ✅ cabeçalho azul
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f1f5f9')]),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
    ]))
    return tb


def _v3_abrir(caminho):
    try:
        if os.name == 'nt':
            os.startfile(caminho)
        elif sys.platform == 'darwin':
            _v3_sub.call(['open', caminho])
        else:
            _v3_sub.call(['xdg-open', caminho])
    except Exception:
        messagebox.showinfo("ℹ️", "PDF salvo em:\n" + caminho)


# ---------------------------------------------------------------------
# EXPORTAR EXCEL / CALC  ->  CABEÇALHO AZUL
# ---------------------------------------------------------------------
def _gerar_arquivo_planilha_v3(self, fmt):
    if fmt == 'xlsx' and not HAS_XLSX:
        messagebox.showerror("📦", "Instale openpyxl"); return
    if fmt == 'ods' and not HAS_ODS:
        messagebox.showerror("📦", "Instale odfpy"); return
    if not self.livros:
        messagebox.showwarning("⚠️", "Catálogo vazio."); return

    h = ['ID', 'Título', 'Autor', 'Estante', 'Prateleira', 'Editora', 'Assunto',
         'Bibliotecário', 'Qtd', 'Disp', 'Emprestado', 'Entrada', 'Saída']
    ch = ['id', 'titulo', 'autor', 'estante', 'prateleira', 'editora', 'assunto',
          'bibliotecario', 'quantidade', 'disponibilidade', 'emprestado_a', 'entrada', 'saida']

    def linha_de(l):
        out = []
        for k in ch:
            v = l.get(k, '')
            out.append('' if v is None else v)
        return out

    if fmt == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook(); ws = wb.active; ws.title = "Catalogo"
            ws.append(h)

            # ✅ estilo do cabeçalho: azul, branco, negrito, centralizado
            fill = PatternFill('solid', fgColor='2563EB')
            font = Font(bold=True, color='FFFFFF')
            alig = Alignment(horizontal='center', vertical='center')
            thin = Side(style='thin', color='CBD5E1')
            bord = Border(left=thin, right=thin, top=thin, bottom=thin)
            for ci in range(1, len(h) + 1):
                c = ws.cell(row=1, column=ci)
                c.fill = fill; c.font = font; c.alignment = alig; c.border = bord
                ws.column_dimensions[get_column_letter(ci)].width = max(len(str(h[ci - 1])) + 3, 10)
            ws.freeze_panes = 'A2'   # congela o cabeçalho

            for l in self.livros:
                ws.append(linha_de(l))

            n = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel", "*.xlsx")])
            if n:
                wb.save(n); messagebox.showinfo("✅", f"Exportado em {n}")
        except Exception:
            messagebox.showerror("❌", _v3_tb.format_exc())
        return

    # --- ODS (Calc) ---
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table as _T, TableRow as _TR, TableCell as _TC
        from odf.text import P as _P

        doc = OpenDocumentSpreadsheet()

        # ✅ estilo de célula azul para o cabeçalho
        hdr_style = None
        try:
            from odf.style import Style, TableCellProperties, TextProperties
            hs = Style(name='CabAzul', family='table-cell')
            hs.addElement(TableCellProperties(backgroundcolor='#2563eb'))
            hs.addElement(TextProperties(color='#ffffff', fontweight='bold'))
            doc.automaticstyles.addElement(hs)
            hdr_style = 'CabAzul'
        except Exception:
            hdr_style = None

        def cel(texto, estilo=None):
            c = _TC(stylename=estilo) if estilo else _TC()
            c.addElement(_P(text=str(texto)))
            return c

        t = _T(name="Catalogo")
        r = _TR()
        for x in h:
            r.addElement(cel(x, hdr_style))
        t.addElement(r)

        for l in self.livros:
            r = _TR()
            for d in linha_de(l):
                r.addElement(cel(d))
            t.addElement(r)

        doc.spreadsheet.addElement(t)
        n = filedialog.asksaveasfilename(defaultextension=".ods", filetypes=[("ODS", "*.ods")])
        if n:
            doc.save(n); messagebox.showinfo("✅", f"Exportado em {n}")
    except Exception:
        messagebox.showerror("❌", _v3_tb.format_exc())


# ---------------------------------------------------------------------
# IMPRIMIR CATÁLOGO / BUSCA  ->  lê direto da tela (com cabeçalho azul)
# ---------------------------------------------------------------------
def _imprimir_catalogo_v3(self, tipo='completo'):
    if not HAS_PDF:
        messagebox.showerror("📦", "Instale reportlab"); return

    if tipo == 'busca':
        if not getattr(self, 'busca_ativa', False):
            messagebox.showwarning(
                "⚠️ Nenhuma busca ativa",
                "Faça uma busca primeiro (🔍 Buscar) e, com os resultados no painel,\n"
                "clique em '🖨️ Busca'.\nPara imprimir tudo, use '🖨️ Catálogo'.")
            return
        chaves, cabecalhos, linhas = _v3_ler_grid(self.resultados_tree)
        titulo = "FEEU - Resultados da Busca"
        if not linhas:
            messagebox.showwarning(
                "⚠️ Nenhum resultado no painel",
                "O painel de resultados está vazio.\nFaça a busca novamente e tente de novo.")
            return
    else:
        chaves, cabecalhos, linhas = _v3_ler_grid(self.tree)
        titulo = "FEEU - Catálogo"
        if not linhas:
            messagebox.showwarning("⚠️", "Catálogo vazio."); return

    try:
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm

        tmp = _v3_temp.NamedTemporaryFile(delete=False, suffix='.pdf', prefix='Cat_').name
        doc = SimpleDocTemplate(tmp, pagesize=landscape(A4),
                                rightMargin=0.8 * cm, leftMargin=0.8 * cm,
                                topMargin=1.0 * cm, bottomMargin=1.0 * cm)
        el = [Paragraph(titulo, getSampleStyleSheet()['Heading1']),
              Spacer(1, 0.4 * cm),
              _v3_montar_pdf(chaves, cabecalhos, linhas)]
        doc.build(el)
        _v3_abrir(tmp)
    except Exception:
        messagebox.showerror("❌", _v3_tb.format_exc())


# Reatribui (a última definição vence, sobrescrevendo as versões antigas)
LivroCatalogApp._gerar_arquivo_planilha = _gerar_arquivo_planilha_v3
LivroCatalogApp.imprimir_catalogo = _imprimir_catalogo_v3
# =====================================================================

# =====================================================================
# PATCH v5 — (A) BOTÃO ✖ FECHAR BUSCA   (B) TEMA ESCURO PERSISTENTE
#  - Garante o botão "✖ Fechar" no painel de resultados (sem duplicar)
#  - "✖ Fechar" limpa termo + filtro + recolhe o painel
#  - Grava o tema em config.json e recarrega no boot
#  - Pinta a TELA DE LOGIN com o tema salvo (claro ou escuro)
#  - Login compatível com SHA-256 E PBKDF2 (não quebra seu banco)
# Cola ANTES de:  if __name__ == "__main__":
# =====================================================================
import json as _v5_json

# --- dicionário mínimo de cores (fallback se o seu módulo não tiver TEMAS) ---
_V5_FALLBACK = {
    'claro': {'bg_janela': '#f1f5f9', 'bg_card': '#ffffff', 'texto': '#0f172a',
              'texto_suave': '#64748b', 'titulo': '#1e40af', 'borda': '#cbd5e1',
              'campo_bg': '#ffffff', 'campo_fg': '#0f172a'},
    'escuro': {'bg_janela': '#0f172a', 'bg_card': '#1e293b', 'texto': '#e2e8f0',
               'texto_suave': '#94a3b8', 'titulo': '#93c5fd', 'borda': '#475569',
               'campo_bg': '#1e293b', 'campo_fg': '#e2e8f0'},
}

def _v5_temas_base():
    return globals().get('TEMAS') or _V5_FALLBACK

def _v5_tema_dict(nome):
    """Retorna o dict de cores do tema, garantindo as chaves mínimas."""
    base = _v5_temas_base()
    out = dict(_V5_FALLBACK.get(nome, _V5_FALLBACK['claro']))
    out.update(base.get(nome, base.get('claro', {})))
    return out

def _v5_cfg_path():
    try:
        if getattr(sys, 'frozen', False):
            d = os.path.dirname(sys.executable)
        else:
            d = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        d = os.getcwd()
    return os.path.join(d, 'config.json')

def _v5_cfg_load():
    try:
        p = _v5_cfg_path()
        if os.path.exists(p):
            with open(p, 'r', encoding='utf-8') as f:
                d = _v5_json.load(f)
            return d if isinstance(d, dict) else {}
    except Exception:
        pass
    return {}

def _v5_cfg_save(tema_nome):
    try:
        d = _v5_cfg_load()
        d['tema'] = tema_nome
        with open(_v5_cfg_path(), 'w', encoding='utf-8') as f:
            _v5_json.dump(d, f, indent=2, ensure_ascii=False)
    except Exception:
        pass

def _v5_skin_styles(t):
    """Aplica o mínimo de estilo ttk para a tela de login respeitar o tema."""
    try:
        s = ttk.Style()
        try: s.theme_use('clam')
        except Exception: pass
        s.configure('TFrame', background=t['bg_janela'])
        s.configure('TLabel', background=t['bg_janela'], foreground=t['texto'])
        s.configure('TLabelframe', background=t['bg_janela'], bordercolor=t['borda'])
        s.configure('TLabelframe.Label', background=t['bg_janela'], foreground=t['titulo'])
        s.configure('TEntry', fieldbackground=t['campo_bg'], foreground=t['campo_fg'],
                    insertcolor=t['campo_fg'], bordercolor=t['borda'])
        s.configure('TCombobox', fieldbackground=t['campo_bg'], foreground=t['campo_fg'],
                    arrowcolor=t['texto'], bordercolor=t['borda'])
        s.configure('TButton', background=t.get('bg_card', t['bg_janela']),
                    foreground=t['texto'], bordercolor=t['borda'])
    except Exception:
        pass

# ---------------------------------------------------------------------
# (BOOT) carrega o tema ANTES de tudo e prepara a tela de login
# ---------------------------------------------------------------------
_v5_orig_init = LivroCatalogApp.__init__

def _v5_new_init(self, root):
    global TEMA_ATUAL
    nome = _v5_cfg_load().get('tema', 'claro')
    TEMA_ATUAL = _v5_tema_dict(nome)
    _v5_skin_styles(TEMA_ATUAL)                 # estilos da tela de login
    _v5_orig_init(self, root)                   # roda o __init__ original (login + tela)
    # pós-boot: sincroniza o tema salvo com a tela principal
    try:
        if not self.root.winfo_exists():
            return
        if not getattr(self, 'config', None):
            self.config = {}
        self.config['tema'] = nome
        self.tema_cores = TEMA_ATUAL
        if hasattr(self, 'aplicar_tema'):
            self.aplicar_tema()
    except Exception:
        pass

LivroCatalogApp.__init__ = _v5_new_init

# ---------------------------------------------------------------------
# (ALTERNAR) grava o tema em disco e reaplica
# ---------------------------------------------------------------------
def _v5_new_alternar(self):
    global TEMA_ATUAL
    cur = 'claro'
    try:
        cur = (self.config or {}).get('tema', 'claro')
    except Exception:
        pass
    novo = 'escuro' if cur == 'claro' else 'claro'
    try:
        self.config['tema'] = novo
    except Exception:
        self.config = {'tema': novo}
    TEMA_ATUAL = _v5_tema_dict(novo)
    try:
        self.tema_cores = TEMA_ATUAL
    except Exception:
        pass
    _v5_cfg_save(novo)                          # ✅ persiste em config.json
    if hasattr(self, '_salvar_config'):         # se o seu código também salva, mantém coerente
        try:
            self._salvar_config()
        except Exception:
            pass
    if hasattr(self, 'aplicar_tema'):
        try:
            self.aplicar_tema()
        except Exception:
            pass
    # atualiza o texto do botão de tema, se existir
    try:
        if hasattr(self, 'btn_tema'):
            self.btn_tema.config(text='☀️ Tema Claro' if novo == 'escuro' else '🌙 Tema Escuro')
    except Exception:
        pass

LivroCatalogApp.alternar_tema = _v5_new_alternar

# ---------------------------------------------------------------------
# (LOGIN) tela pintada pelo tema salvo + senha SHA-256 ou PBKDF2
# ---------------------------------------------------------------------
def _v5_verifica_senha(senha, armazenada):
    if not armazenada:
        return False
    arm = str(armazenada).strip()
    # PBKDF2
    if arm.startswith('pbkdf2_sha256$'):
        try:
            import hmac as _hmac
            _, it, salt, hh = arm.split('$')
            dk = hashlib.pbkdf2_hmac('sha256', senha.encode('utf-8'),
                                     bytes.fromhex(salt), int(it))
            return _hmac.compare_digest(dk.hex(), hh)
        except Exception:
            return False
    # SHA-256 simples (padrão da 4.4.1)
    if len(arm) == 64:
        try:
            import hmac as _hmac
            return _hmac.compare_digest(hashlib.sha256(senha.encode('utf-8')).hexdigest(), arm)
        except Exception:
            return hashlib.sha256(senha.encode('utf-8')).hexdigest() == arm
    # texto puro (último recurso)
    return arm == senha

def _v5_new_autenticar(self):
    global TEMA_ATUAL
    t = TEMA_ATUAL if isinstance(TEMA_ATUAL, dict) else _v5_tema_dict('claro')
    bg = t.get('bg_janela', '#f8fafc')
    card = t.get('bg_card', '#ffffff')
    fg = t.get('texto', '#0f172a')
    fgs = t.get('texto_suave', '#64748b')
    fgt = t.get('titulo', '#1e40af')

    # último usuário
    last_user = ''
    try:
        lf = os.path.join(self.diretorio_app, 'last_user.conf')
        if os.path.exists(lf):
            with open(lf, 'r', encoding='utf-8') as f:
                last_user = f.read().strip()
    except Exception:
        pass

    login_win = tk.Toplevel(self.root)
    login_win.title("🔐 Login do Sistema")
    login_win.geometry("520x580")
    login_win.resizable(False, False)
    login_win.grab_set()
    login_win.attributes("-topmost", True)
    try:
        login_win.configure(bg=bg)
    except Exception:
        pass
    try:
        self.centralizar_janela(login_win)
    except Exception:
        pass

    main_frame = tk.Frame(login_win, bg=bg); main_frame.pack(fill="both", expand=True)
    welcome = tk.Frame(main_frame, bg=bg); welcome.pack(fill="x", pady=(15, 10))

    sym = tk.Canvas(welcome, width=60, height=60, bg=card, highlightthickness=1, relief="solid")
    sym.pack(side="left", padx=(30, 10))
    cx = cy = 30
    sym.create_oval(4, 4, 56, 56, outline=self.cores['verde'], width=3)
    sym.create_oval(12, 12, 48, 48, outline=self.cores['dourado'], width=2)
    rt = 18
    sym.create_polygon(cx, cy - rt,
                       cx - rt * math.cos(math.radians(30)), cy + rt * math.sin(math.radians(30)),
                       cx + rt * math.cos(math.radians(30)), cy + rt * math.sin(math.radians(30)),
                       outline=self.cores['azul'], width=2, fill='')
    sym.create_oval(cx - 2, cy - 2, cx + 2, cy + 2, fill=self.cores['ponto'])

    # bandeirinha (mantém o visual da 4.4.1)
    flag = tk.Canvas(welcome, width=70, height=50, bg=card, highlightthickness=1, relief="solid")
    flag.pack(side="right", padx=(10, 30))
    wf, hf = 70, 50
    flag.create_rectangle(0, 0, wf, hf, fill='#009C3B', outline='')
    flag.create_polygon(wf // 2, 5, wf - 5, hf // 2, wf // 2, hf - 5, 5, hf // 2, fill='#FFDF00', outline='')
    flag.create_oval(wf // 2 - 14, hf // 2 - 14, wf // 2 + 14, hf // 2 + 14, fill='#002776', outline='')

    tk.Label(welcome, text="📚 Catálogo de Livros FEEU", font=('DejaVu Sans', 16, 'bold'),
             foreground=fgt, background=bg, wraplength=480).pack(side='bottom', pady=(5, 10))
    tk.Label(welcome, text="Bem-vindo ao Sistema", font=('DejaVu Sans', 11),
             foreground=fgs, background=bg).pack(side='bottom')

    form = tk.Frame(main_frame, bg=card, relief="solid", bd=1); form.pack(fill="x", padx=20, pady=15)
    inp = ttk.Frame(form, padding=15); inp.pack(fill='x')

    ttk.Label(inp, text="Usuário: ").grid(row=0, column=0, sticky='w', pady=5)
    conn = sqlite3.connect(self.db_file); cur = conn.cursor()
    try:
        cur.execute("SELECT nome FROM usuarios ORDER BY nome")
        users = [r[0] for r in cur.fetchall()]
    except Exception:
        users = []
    conn.close()

    user_var = tk.StringVar()
    uc = ttk.Combobox(inp, textvariable=user_var, values=users, state='readonly')
    uc.grid(row=0, column=1, padx=5, pady=5, sticky='we')
    uc.set(last_user if last_user in users else (users[0] if users else ""))

    ttk.Label(inp, text="Senha: ").grid(row=1, column=0, sticky='w', pady=5)
    pass_var = tk.StringVar()
    pf = ttk.Frame(inp); pf.grid(row=1, column=1, padx=5, pady=5, sticky='we')
    pe = ttk.Entry(pf, textvariable=pass_var, show='*'); pe.pack(side='left', fill='x', expand=True)
    vis = {'v': False}

    def toggle():
        vis['v'] = not vis['v']
        pe.config(show='' if vis['v'] else '*')
        olho.config(text='👁️' if vis['v'] else '🙈')

    olho = ttk.Button(pf, text='🙈', command=toggle, width=3); olho.pack(side='right')

    result = {"ok": False, "id": None, "nome": None, "perms": {}}

    def do_login():
        u = user_var.get().strip(); p = pass_var.get().strip()
        if not u:
            messagebox.showwarning("⚠️", "Selecione um usuário.", parent=login_win); return
        if not p:
            messagebox.showwarning("⚠️", "Digite a senha.", parent=login_win); return
        conn = sqlite3.connect(self.db_file); cur = conn.cursor()
        row = None
        try:
            cur.execute("SELECT * FROM usuarios WHERE nome=?", (u,))
            desc = [d[0] for d in cur.description]
            raw = cur.fetchone()
            if raw:
                row = dict(zip(desc, raw))
        except Exception:
            row = None
        conn_ok = conn
        if row and _v5_verifica_senha(p, row.get('senha', '')):
            # grava último usuário
            try:
                with open(os.path.join(self.diretorio_app, 'last_user.conf'), 'w', encoding='utf-8') as f:
                    f.write(u)
            except Exception:
                pass
            # monta permissões por NOME de coluna (à prova de ordem/colunas extras)
            def bp(k):
                return bool(row.get(k, 0))
            perms = {
                'inserir': bp('perm_inserir'), 'editar': bp('perm_editar'),
                'apagar': bp('perm_apagar'), 'exportar': bp('perm_exportar'),
                'importar': bp('perm_importar'), 'relatorio': bp('perm_relatorio'),
                'admin': bp('perm_admin'), 'emprestados': bp('perm_emprestados'),
            }
            result.update({"ok": True, "id": row.get('id'), "nome": row.get('nome'), "perms": perms})
            try: conn_ok.close()
            except Exception: pass
            login_win.destroy()
        else:
            try: conn_ok.close()
            except Exception: pass
            messagebox.showerror("❌", "Usuário ou senha incorretos.", parent=login_win)
            pe.delete(0, tk.END); pe.focus_set()

    bf = ttk.Frame(login_win, padding=10); bf.pack(fill='x')
    ttk.Button(bf, text="Entrar", command=do_login, style='Import.TButton').pack(side='left', expand=True, padx=5)
    ttk.Button(bf, text="Cancelar", command=login_win.destroy).pack(side='left', expand=True, padx=5)
    login_win.bind("<Return>", lambda e: do_login())

    pe.focus_set()
    try: login_win.after(150, lambda: pe.focus_force())   # cursor piscando na senha
    except Exception: pass
    login_win.lift(); login_win.focus_force()
    self.root.wait_window(login_win)

    if result["ok"]:
        self.usuario_id = result["id"]; self.usuario_nome = result["nome"]; self.permissoes = result["perms"]
        return True
    return False

LivroCatalogApp.autenticar = _v5_new_autenticar

# ---------------------------------------------------------------------
# (BUSCA) botão ✖ garantido + fechar limpa tudo
# ---------------------------------------------------------------------
def _v5_new_fechar(self):
    try:
        self.main_paned.forget(self.resultados_frame)
    except Exception:
        pass
    self.busca_ativa = False
    try:
        self.ultima_busca = []
    except Exception:
        pass
    try:
        self.busca_var.set('')
    except Exception:
        pass
    try:
        self.filtro_var.set('Todos')
    except Exception:
        pass
    try:
        self.carregar_dados()
    except Exception:
        pass
    try:
        self.limpar_form()
    except Exception:
        pass
    try:
        self.status_var.set(f"Exibindo {len(self.livros)} livros")
    except Exception:
        pass
    try:
        if hasattr(self, '_atualizar_estado_btn_busca'):
            self._atualizar_estado_btn_busca()
    except Exception:
        pass

LivroCatalogApp.fechar_painel_resultados = _v5_new_fechar

def _v5_tem_botao_fechar(frame):
    try:
        for w in frame.winfo_children():
            txt = ''
            try: txt = str(w.cget('text'))
            except Exception: pass
            if ('Fechar' in txt) or ('✖' in txt) or ('x' == txt.strip().lower()):
                return True
    except Exception:
        pass
    return False

def _v5_garantir_x(self):
    try:
        rf = getattr(self, 'resultados_frame', None)
        if not rf:
            return
        if getattr(rf, '_v5_x', False):
            return
        if _v5_tem_botao_fechar(rf):
            rf._v5_x = True
            return
        b = ttk.Button(rf, text="✖ Fechar", command=self.fechar_painel_resultados)
        b.pack(side='right', padx=10, pady=5)
        rf._v5_x = True
    except Exception:
        pass

_v5_orig_cw = getattr(LivroCatalogApp, 'create_widgets', None)
if _v5_orig_cw is not None:
    def _v5_new_cw(self):
        _v5_orig_cw(self)
        _v5_garantir_x(self)
    LivroCatalogApp.create_widgets = _v5_new_cw
# =====================================================================

# =====================================================================
# PATCH v6 — TEMA PERSISTENTE  +  PAINÉIS RESPONSIVOS  +  BOTÃO ✖ VISÍVEL
# Cola este bloco ANTES da linha:  if __name__ == "__main__":
# Não apague nada acima. Este bloco vence os patches antigos de tema.
# =====================================================================
import json as json   # <- garante o nome global 'json' (corrige _carregar_config/_salvar_config)

# Referência ao __init__ "puro" (sem o patch de tema que usava caminho errado).
_BASE_INIT = globals().get('_v5_orig_init')
if _BASE_INIT is None:
    _BASE_INIT = LivroCatalogApp.__init__


def _v6_cfg_nome():
    """Lê o tema do config CANÔNICO (mesma pasta do .py), sem depender do cwd."""
    try:
        p = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')
        if os.path.exists(p):
            with open(p, 'r', encoding='utf-8') as f:
                return (json.load(f) or {}).get('tema', 'claro')
    except Exception:
        pass
    return 'claro'


def _v6_skin_login(t):
    """Pinta os campos da tela de login no tema salvo (antes do boot terminar)."""
    try:
        s = ttk.Style()
        try:
            s.theme_use('clam')
        except Exception:
            pass
        s.configure('TFrame', background=t['bg_janela'])
        s.configure('TLabel', background=t['bg_janela'], foreground=t['texto'])
        s.configure('TLabelframe', background=t['bg_janela'], bordercolor=t['borda'])
        s.configure('TLabelframe.Label', background=t['bg_janela'], foreground=t['titulo'])
        s.configure('TEntry', fieldbackground=t['campo_bg'], foreground=t['campo_fg'],
                    insertcolor=t['campo_fg'], bordercolor=t['borda'])
        s.configure('TCombobox', fieldbackground=t['campo_bg'], foreground=t['campo_fg'],
                    arrowcolor=t['texto'], bordercolor=t['borda'])
        s.configure('TButton', background=t.get('bg_card', t['bg_janela']),
                    foreground=t['texto'], bordercolor=t['borda'])
    except Exception:
        pass


def _v6_new_init(self, root):
    # 1) tema global ANTES do boot (p/ calendário e afins nascerem no tema certo)
    try:
        globals()['TEMA_ATUAL'] = TEMAS.get(_v6_cfg_nome(), TEMAS['claro'])
        _v6_skin_login(globals()['TEMA_ATUAL'])     # tela de login no tema salvo
    except Exception:
        pass

    # 2) boot puro: agora _carregar_config funciona (json existe) -> login já no tema certo
    _BASE_INIT(self, root)

    # 3) coerência final do tema a partir do CANÔNICO
    try:
        nome = self.config.get('tema', 'claro')
        self.tema_cores = TEMAS.get(nome, TEMAS['claro'])
        if hasattr(self, 'aplicar_tema'):
            self.aplicar_tema()
    except Exception:
        pass

    # 4) layout do painel de resultados (✖ sempre visível) + painéis responsivos
    try:
        self._v6_reorganizar_resultados()
    except Exception:
        pass
    try:
        self._v6_instalar_resize_panes()
    except Exception:
        pass


def _v6_reorganizar_resultados(self):
    """Fixa o botão ✖ no canto superior direito, sem competir por espaço."""
    rf = getattr(self, 'resultados_frame', None)
    if not rf:
        return
    btn = None
    rc = None
    for w in rf.winfo_children():
        try:
            if isinstance(w, ttk.Button) and 'Fechar' in str(w.cget('text')):
                btn = w
            elif isinstance(w, ttk.Frame):
                rc = w
        except Exception:
            pass
    if rc:
        try:
            rc.pack_configure(pady=(28, 0))   # reserva faixa no topo p/ o botão
        except Exception:
            pass
    if btn:
        try:
            btn.pack_forget()
            btn.place(relx=1.0, x=-6, y=4, anchor='ne')   # flutua no canto sup. direito
            btn.lift()
        except Exception:
            pass


def _v6_ajustar_panes(self):
    """Mantém proporções dos painéis ao redimensionar a janela."""
    try:
        self.main_paned.update_idletasks()
        H = self.main_paned.winfo_height()
        if H < 120:
            return
        panes = [str(p) for p in self.main_paned.panes()]
        rf = str(self.resultados_frame)
        if rf in panes:                      # 3 painéis: catálogo / resultados / cadastro
            try:
                self.main_paned.sashpos(0, int(H * 0.42))
            except Exception:
                pass
            try:
                self.main_paned.sashpos(1, int(H * 0.72))
            except Exception:
                pass
        else:                                # 2 painéis: catálogo / cadastro
            try:
                self.main_paned.sashpos(0, int(H * 0.62))
            except Exception:
                pass
    except Exception:
        pass


def _v6_instalar_resize_panes(self):
    def _on_cfg(event):
        try:
            if event.widget is self.root:
                self._v6_ajustar_panes()
        except Exception:
            pass
    self.root.bind('<Configure>', _on_cfg, add='+')
    self.root.after(250, self._v6_ajustar_panes)


def _v6_new_alternar(self):
    """Alterna e GRAVA o tema no config CANÔNICO (mesma pasta do .py)."""
    atual = self.config.get('tema', 'claro')
    novo = 'escuro' if atual == 'claro' else 'claro'
    self.config['tema'] = novo
    self.tema_cores = TEMAS.get(novo, TEMAS['claro'])
    try:
        globals()['TEMA_ATUAL'] = self.tema_cores
    except Exception:
        pass
    try:
        self._salvar_config()              # -> diretorio_app/config.json (agora funciona)
    except Exception:
        pass
    try:
        self.aplicar_tema()
    except Exception:
        pass
    try:
        if hasattr(self, 'btn_tema'):
            self.btn_tema.config(text='☀️ Tema Claro' if novo == 'escuro' else '🌙 Tema Escuro')
    except Exception:
        pass


# --- wrappers finos: reajustam os painéis quando a busca abre/fecha ---
_busca_orig = LivroCatalogApp.buscar
def _v6_busca_wrap(self, *a, **k):
    r = _busca_orig(self, *a, **k)
    try:
        self.root.after(60, self._v6_ajustar_panes)
    except Exception:
        pass
    return r

_fecha_orig = LivroCatalogApp.fechar_painel_resultados
def _v6_fecha_wrap(self, *a, **k):
    r = _fecha_orig(self, *a, **k)
    try:
        self.root.after(60, self._v6_ajustar_panes)
    except Exception:
        pass
    return r

_limpa_orig = LivroCatalogApp.limpar_busca
def _v6_limpa_wrap(self, *a, **k):
    r = _limpa_orig(self, *a, **k)
    try:
        self.root.after(60, self._v6_ajustar_panes)
    except Exception:
        pass
    return r


# --- reatribui (a última definição vence) ---
LivroCatalogApp.__init__ = _v6_new_init
LivroCatalogApp.alternar_tema = _v6_new_alternar
LivroCatalogApp._v6_reorganizar_resultados = _v6_reorganizar_resultados
LivroCatalogApp._v6_ajustar_panes = _v6_ajustar_panes
LivroCatalogApp._v6_instalar_resize_panes = _v6_instalar_resize_panes
LivroCatalogApp.buscar = _v6_busca_wrap
LivroCatalogApp.fechar_painel_resultados = _v6_fecha_wrap
LivroCatalogApp.limpar_busca = _v6_limpa_wrap
# =====================================================================
# FIM PATCH v6  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v7 — LOGIN LEGÍVEL NO ESCURO + LAYOUT PERSISTENTE/RESPONSIVO + i18n
#  (1) Tela de login com contraste correto no tema escuro
#  (2) Posições dos divisores + tamanho da janela SALVOS e restaurados
#  (3) Alturas mínimas por painel (não esmaga em 1024x768)
#  (4) Idiomas pt-BR / en / es / eo com dropdown salvo (tela principal)
# Tudo em try/except: se algo falhar, desativa só aquela parte.
# Cola ANTES de:  if __name__ == "__main__":
# =====================================================================
import json as _v7_json
import re as _v7_re

# ---- dicionários de idioma (chaves = texto pt-BR EXATO da tela) ----
_V7_TR = {
 'pt-BR': {},  # pt-BR é a própria chave (lookup retorna o texto original)
 'en': {
   ' 🔍 Busca e Filtros ': ' 🔍 Search and Filters ',
   ' 📤 Exportar, Importar e Imprimir ': ' 📤 Export, Import and Print ',
   ' 📖 Catálogo Completo ': ' 📖 Full Catalog ',
   ' 🔍 Resultados Encontrados ': ' 🔍 Results Found ',
   ' Cadastro / Edição / Localização / Empréstimos ': ' Entry / Edit / Location / Loans ',
   'Filtrar por: ': 'Filter by: ', 'Termo: ': 'Term: ',
   'Exportar: ': 'Export: ', 'Importar: ': 'Import: ', 'Imprimir: ': 'Print: ',
   'Título:': 'Title:', 'Autor:': 'Author:', 'Estante:': 'Shelf:', 'Prateleira:': 'Rack:',
   'Editora:': 'Publisher:', 'Assunto:': 'Subject:', 'Bibliotecário:': 'Librarian:',
   'Quantidade:': 'Qty:', 'Emprestado a:': 'Loaned to:', 'Entrada:': 'In:', 'Saída:': 'Out:',
   '🔍 Buscar': '🔍 Search', '🔄 Limpar': '🔄 Clear',
   '📊 Excel': '📊 Excel', '📈 Calc': '📈 Calc', '📄 PDF': '📄 PDF',
   '📥 Planilha': '📥 Spreadsheet', '🖨️ Catálogo': '🖨️ Catalog', '🖨️ Busca': '🖨️ Search',
   '❓ Ajuda': '❓ Help', 'ℹ️ Sobre': 'ℹ️ About',
   '➕ Adicionar': '➕ Add', '💾 Salvar': '💾 Save', '🗑️ Excluir': '🗑️ Delete',
   '📤 Emprestar': '📤 Loan', '📥 Devolver': '📥 Return', '📚 Emprestados': '📚 Loaned',
   '🔑 Usuários': '🔑 Users', '🔒 Senha Admin': '🔒 Admin Pwd',
   '🚪 Sair': '🚪 Exit', '🔙 Voltar ao Login': '🔙 Back to Login', '✖ Fechar': '✖ Close',
   '🌙 Tema Escuro': '🌙 Dark Theme', '☀️ Tema Claro': '☀️ Light Theme',
   'ID': 'ID', 'Título': 'Title', 'Autor': 'Author', 'Estante': 'Shelf', 'Prateleira': 'Rack',
   'Editora': 'Publisher', 'Assunto': 'Subject', 'Bibliotecário': 'Librarian',
   'Qtd.': 'Qty', 'Disp.': 'Avail.', 'Emprestado a': 'Loaned to', 'Entrada': 'In', 'Saída': 'Out',
   'Desenvolvido pelos Irmãos Rodolpho e Alexandre': 'Developed by Brothers Rodolpho and Alexandre',
   'Catálogo de Livros da Fraternidade Eclética Espiritualista Universal':
       'Book Catalog of Fraternidade Eclética Espiritualista Universal',
   "*PDF abre no visualizador para Ctrl+P   •   '🖨️ Busca' só fica ativo quando houver resultados de busca":
       "*PDF opens in viewer for Ctrl+P   •   '🖨️ Search' is enabled only when there are search results",
 },
 'es': {
   ' 🔍 Busca e Filtros ': ' 🔍 Búsqueda y Filtros ',
   ' 📤 Exportar, Importar e Imprimir ': ' 📤 Exportar, Importar e Imprimir ',
   ' 📖 Catálogo Completo ': ' 📖 Catálogo Completo ',
   ' 🔍 Resultados Encontrados ': ' 🔍 Resultados Encontrados ',
   ' Cadastro / Edição / Localização / Empréstimos ': ' Alta / Edición / Ubicación / Préstamos ',
   'Filtrar por: ': 'Filtrar por: ', 'Termo: ': 'Término: ',
   'Exportar: ': 'Exportar: ', 'Importar: ': 'Importar: ', 'Imprimir: ': 'Imprimir: ',
   'Título:': 'Título:', 'Autor:': 'Autor:', 'Estante:': 'Estante:', 'Prateleira:': 'Estante:',
   'Editora:': 'Editorial:', 'Assunto:': 'Asunto:', 'Bibliotecário:': 'Bibliotecario:',
   'Quantidade:': 'Cant.:', 'Emprestado a:': 'Prestado a:', 'Entrada:': 'Entrada:', 'Saída:': 'Salida:',
   '🔍 Buscar': '🔍 Buscar', '🔄 Limpar': '🔄 Limpiar',
   '📊 Excel': '📊 Excel', '📈 Calc': '📈 Calc', '📄 PDF': '📄 PDF',
   '📥 Planilha': '📥 Planilla', '🖨️ Catálogo': '🖨️ Catálogo', '🖨️ Busca': '🖨️ Búsqueda',
   '❓ Ajuda': '❓ Ayuda', 'ℹ️ Sobre': 'ℹ️ Acerca de',
   '➕ Adicionar': '➕ Añadir', '💾 Salvar': '💾 Guardar', '🗑️ Excluir': '🗑️ Eliminar',
   '📤 Emprestar': '📤 Prestar', '📥 Devolver': '📥 Devolver', '📚 Emprestados': '📚 Prestados',
   '🔑 Usuários': '🔑 Usuarios', '🔒 Senha Admin': '🔒 Clave Admin',
   '🚪 Sair': '🚪 Salir', '🔙 Voltar ao Login': '🔙 Volver al Login', '✖ Fechar': '✖ Cerrar',
   '🌙 Tema Escuro': '🌙 Tema Oscuro', '☀️ Tema Claro': '☀️ Tema Claro',
   'ID': 'ID', 'Título': 'Título', 'Autor': 'Autor', 'Estante': 'Estante', 'Prateleira': 'Estante',
   'Editora': 'Editorial', 'Assunto': 'Asunto', 'Bibliotecário': 'Bibliotecario',
   'Qtd.': 'Cant.', 'Disp.': 'Disp.', 'Emprestado a': 'Prestado a', 'Entrada': 'Entrada', 'Salida': 'Salida',
   'Desenvolvido pelos Irmãos Rodolpho e Alexandre': 'Desarrollado por los Hermanos Rodolpho y Alexandre',
   'Catálogo de Livros da Fraternidade Eclética Espiritualista Universal':
       'Catálogo de Libros de la Fraternidade Eclética Espiritualista Universal',
   "*PDF abre no visualizador para Ctrl+P   •   '🖨️ Busca' só fica ativo quando houver resultados de busca":
       "*PDF abre en el visor para Ctrl+P   •   '🖨️ Búsqueda' se activa solo con resultados de búsqueda",
 },
 'eo': {
   ' 🔍 Busca e Filtros ': ' 🔍 Serĉo kaj Filtroj ',
   ' 📤 Exportar, Importar e Imprimir ': ' 📤 Eksporti, Importi kaj Presi ',
   ' 📖 Catálogo Completo ': ' 📖 Plena Katalogo ',
   ' 🔍 Resultados Encontrados ': ' 🔍 Trovitaj Rezultoj ',
   ' Cadastro / Edição / Localização / Empréstimos ': ' Enigo / Redakto / Loko / Pruntoj ',
   'Filtrar por: ': 'Filtri laŭ: ', 'Termo: ': 'Termo: ',
   'Exportar: ': 'Eksporti: ', 'Importar: ': 'Importi: ', 'Imprimir: ': 'Presi: ',
   'Título:': 'Titolo:', 'Autor:': 'Aŭtoro:', 'Estante:': 'Ŝranko:', 'Prateleira:': 'Breto:',
   'Editora:': 'Eldonejo:', 'Assunto:': 'Temo:', 'Bibliotecário:': 'Bibliotekisto:',
   'Quantidade:': 'Kvanto:', 'Emprestado a:': 'Pruntita al:', 'Entrada:': 'Eniro:', 'Saída:': 'Eliro:',
   '🔍 Buscar': '🔍 Serĉi', '🔄 Limpar': '🔄 Viŝi',
   '📊 Excel': '📊 Excel', '📈 Calc': '📈 Calc', '📄 PDF': '📄 PDF',
   '📥 Planilha': '📥 Tabelo', '🖨️ Catálogo': '🖨️ Katalogo', '🖨️ Busca': '🖨️ Serĉo',
   '❓ Ajuda': '❓ Helpo', 'ℹ️ Sobre': 'ℹ️ Pri',
   '➕ Adicionar': '➕ Aldoni', '💾 Salvar': '💾 Konservi', '🗑️ Excluir': '🗑️ Forigi',
   '📤 Emprestar': '📤 Prunti', '📥 Devolver': '📥 Redoni', '📚 Emprestados': '📚 Pruntitaj',
   '🔑 Usuários': '🔑 Uzantoj', '🔒 Senha Admin': '🔒 Pasvorto',
   '🚪 Sair': '🚪 Eliri', '🔙 Voltar ao Login': '🔙 Reen al Saluto', '✖ Fechar': '✖ Fermi',
   '🌙 Tema Escuro': '🌙 Malhela Temo', '☀️ Tema Claro': '☀️ Hela Temo',
   'ID': 'ID', 'Título': 'Titolo', 'Autor': 'Aŭtoro', 'Estante': 'Ŝranko', 'Prateleira': 'Breto',
   'Editora': 'Eldonejo', 'Assunto': 'Temo', 'Bibliotecário': 'Bibliotekisto',
   'Qtd.': 'Kv.', 'Disp.': 'Disp.', 'Emprestado a': 'Pruntita al', 'Entrada': 'Eniro', 'Saída': 'Eliro',
   'Desenvolvido pelos Irmãos Rodolpho e Alexandre': 'Farita de la Fratoj Rodolpho kaj Alexandre',
   'Catálogo de Livros da Fraternidade Eclética Espiritualista Universal':
       'Libro-Katalogo de Fraternidade Eclética Espiritualista Universal',
   "*PDF abre no visualizador para Ctrl+P   •   '🖨️ Busca' só fica ativo quando houver resultados de busca":
       "*PDF malfermiĝas en vidigilo por Ctrl+P   •   '🖨️ Serĉo' aktiviĝas nur kun serĉ-rezultoj",
 },
}

_V7_IDIOMAS = [('pt-BR', '🇧🇷 Português'), ('en', '🇺 English'),
               ('es', '🇪🇸 Español'), ('eo', '🌐 Esperanto')]

_V7_DIAS = {
 'pt-BR': {'Monday': 'Segunda-Feira', 'Tuesday': 'Terça-Feira', 'Wednesday': 'Quarta-Feira',
           'Thursday': 'Quinta-Feira', 'Friday': 'Sexta-Feira', 'Saturday': 'Sábado', 'Sunday': 'Domingo'},
 'en': {'Monday': 'Monday', 'Tuesday': 'Tuesday', 'Wednesday': 'Wednesday', 'Thursday': 'Thursday',
        'Friday': 'Friday', 'Saturday': 'Saturday', 'Sunday': 'Sunday'},
 'es': {'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles', 'Thursday': 'Jueves',
        'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'},
 'eo': {'Monday': 'Lundo', 'Tuesday': 'Mardo', 'Wednesday': 'Merkredo', 'Thursday': 'Ĵado',
        'Friday': 'Vendredo', 'Saturday': 'Sabato', 'Sunday': 'Dimanĉo'},
}

# ---- config unificado (mesmo config.json do programa) ----
def _v7_cfg_path():
    try:
        d = os.path.dirname(os.path.abspath(__file__))
    except Exception:
        d = os.getcwd()
    return os.path.join(d, 'config.json')

def _v7_read_cfg():
    try:
        p = _v7_cfg_path()
        if os.path.exists(p):
            with open(p, 'r', encoding='utf-8') as f:
                d = _v7_json.load(f)
            return d if isinstance(d, dict) else {}
    except Exception:
        pass
    return {}

def _v7_write_cfg(d):
    try:
        with open(_v7_cfg_path(), 'w', encoding='utf-8') as f:
            _v7_json.dump(d, f, indent=2, ensure_ascii=False)
    except Exception:
        pass

def _v7_t(self, pt):
    """Traduz um texto pt pelo idioma atual (se faltar, devolve o próprio pt)."""
    return _V7_TR.get(getattr(self, '_v7_idioma', 'pt-BR'), {}).get(pt, pt)

# ---- (1) estilo completo aplicado ANTES do login (corrige contraste) ----
def _v7_aplicar_style(t):
    try:
        s = ttk.Style()
        try:
            s.theme_use('clam')
        except Exception:
            pass
        s.configure('TFrame', background=t['bg_janela'])
        s.configure('TLabel', background=t['bg_janela'], foreground=t['texto'])
        s.configure('TLabelframe', background=t['bg_janela'], bordercolor=t['borda'])
        s.configure('TLabelframe.Label', background=t['bg_janela'], foreground=t['titulo'])
        s.configure('TEntry', fieldbackground=t['campo_bg'], foreground=t['campo_fg'],
                    insertcolor=t['campo_fg'], bordercolor=t['borda'])
        # ✅ o 'map' readonly era o que escondia o texto do combobox no escuro
        s.configure('TCombobox', fieldbackground=t['campo_bg'], foreground=t['campo_fg'],
                    arrowcolor=t['texto'], bordercolor=t['borda'])
        s.map('TCombobox', fieldbackground=[('readonly', t['campo_bg'])],
              foreground=[('readonly', t['campo_fg'])])
        s.configure('TButton', background=t['bg_botao'], foreground=t['texto'], bordercolor=t['borda'])
        s.map('TButton', background=[('active', t['bg_botao_hover'])], foreground=[('active', t['texto'])])
    except Exception:
        pass

_V7_BASE_INIT = globals().get('_BASE_INIT') or globals().get('_v5_orig_init')

def _v7_new_init(self, root):
    cfg = _v7_read_cfg()
    tema_nome = cfg.get('tema', 'claro')
    t = TEMAS.get(tema_nome, TEMAS['claro'])
    try:
        globals()['TEMA_ATUAL'] = t
    except Exception:
        pass
    _v7_aplicar_style(t)                       # ✅ tela de login já nasce no tema certo
    if _V7_BASE_INIT is not None:
        _V7_BASE_INIT(self, root)
    else:
        LivroCatalogApp.__init__(self, root)
    # pós-boot: idioma salvo + layout salvo
    try:
        self._v7_idioma = cfg.get('idioma', 'pt-BR')
        self._v7_aplicar_idioma(self._v7_idioma)
    except Exception:
        self._v7_idioma = 'pt-BR'
    try:
        self.root.after(350, self._v7_restaurar_layout)
    except Exception:
        pass

# ---- (2)/(3) layout persistente e responsivo ----
def _v7_ajustar_panes(self):
    """Garante alturas mínimas por painel (não esmaga o formulário)."""
    try:
        self.main_paned.update_idletasks()
        H = self.main_paned.winfo_height()
        if H < 150:
            return
        panes = [str(p) for p in self.main_paned.panes()]
        rf = str(self.resultados_frame)
        min_cat, min_res, min_form = 120, 110, 200
        if rf in panes and len(panes) >= 3:
            try:
                c0 = self.main_paned.sashpos(0)
                c1 = self.main_paned.sashpos(1)
            except Exception:
                return
            if c0 < min_cat:
                c0 = min_cat
            if c1 - c0 < min_res:
                c1 = c0 + min_res
            if H - c1 < min_form:
                c1 = H - min_form
            if c1 - c0 < min_res:
                c0 = max(min_cat, c1 - min_res)
            try:
                self.main_paned.sashpos(0, c0)
                self.main_paned.sashpos(1, c1)
            except Exception:
                pass
        elif len(panes) >= 2:
            try:
                c0 = self.main_paned.sashpos(0)
            except Exception:
                return
            if c0 < min_cat:
                c0 = min_cat
            if H - c0 < min_form:
                c0 = H - min_form
            if c0 < min_cat:
                c0 = min_cat
            try:
                self.main_paned.sashpos(0, c0)
            except Exception:
                pass
    except Exception:
        pass

def _v7_proporcao_inicial(self):
    try:
        self.main_paned.update_idletasks()
        H = self.main_paned.winfo_height()
        if H < 150:
            return
        panes = [str(p) for p in self.main_paned.panes()]
        if str(self.resultados_frame) in panes and len(panes) >= 3:
            self.main_paned.sashpos(0, int(H * 0.42))
            self.main_paned.sashpos(1, int(H * 0.70))
        elif len(panes) >= 2:
            self.main_paned.sashpos(0, int(H * 0.62))
        self._v7_ajustar_panes()
    except Exception:
        pass

def _v7_clamp_geometry(self, g):
    try:
        m = _v7_re.match(r'(\d+)x(\d+)\+(-?\d+)\+(-?\d+)', g)
        if not m:
            return None
        W, H, X, Y = map(int, m.groups())
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        W = min(W, sw); H = min(H, sh)
        X = max(0, min(X, sw - W)); Y = max(0, min(Y, sh - H))
        return f"{W}x{H}+{X}+{Y}"
    except Exception:
        return None

def _v7_restaurar_layout(self):
    try:
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        self.root.minsize(min(1000, sw - 16), min(500, sh - 16))   # ✅ cabe em 1024x768
        cfg = _v7_read_cfg()
        if cfg.get('zoomed'):
            try:
                self.root.state('zoomed')
            except Exception:
                pass
        elif cfg.get('geometry'):
            g = self._v7_clamp_geometry(cfg['geometry'])
            if g:
                try:
                    self.root.geometry(g)
                except Exception:
                    pass
        self.root.update_idletasks()
        fracs = cfg.get('sashes')
        if fracs:
            H = self.main_paned.winfo_height()
            for i, f in enumerate(fracs):
                try:
                    self.main_paned.sashpos(i, int(float(f) * H))
                except Exception:
                    pass
            self._v7_ajustar_panes()
        else:
            self._v7_proporcao_inicial()
    except Exception:
        pass

def _v7_salvar_layout(self):
    try:
        cfg = _v7_read_cfg()
        try:
            cfg['zoomed'] = (self.root.state() == 'zoomed')
        except Exception:
            cfg['zoomed'] = False
        if not cfg['zoomed']:
            cfg['geometry'] = self.root.geometry()
        panes = self.main_paned.panes()
        H = self.main_paned.winfo_height() or 1
        fracs = []
        for i in range(len(panes) - 1):
            try:
                fracs.append(round(self.main_paned.sashpos(i) / H, 4))
            except Exception:
                pass
        cfg['sashes'] = fracs
        _v7_write_cfg(cfg)
    except Exception:
        pass

# ---- (4) i18n: registra widgets uma vez (em pt) e reaplica idioma ----
def _v7_walk(w):
    yield w
    try:
        for c in w.winfo_children():
            for x in _v7_walk(c):
                yield x
    except Exception:
        pass

def _v7_registrar_idioma(self):
    if getattr(self, '_v7_i18n_reg', False):
        return
    self._v7_map = {}
    try:
        for w in _v7_walk(self.root):
            try:
                t = w.cget('text')
            except Exception:
                continue
            if isinstance(t, str) and t.strip():
                self._v7_map[w] = t
    except Exception:
        pass
    # guarda o pt original dos cabeçalhos de coluna (lista compartilhada)
    try:
        for c in self.COLUNAS:
            c.setdefault('_pt', c.get('text', ''))
    except Exception:
        pass
    try:
        self._v7_win_title_pt = self.root.title()
    except Exception:
        pass
    # injeta o dropdown de idioma no canto superior direito
    try:
        parent = self.btn_tema.master
        rot_atual = dict(_V7_IDIOMAS).get(_v7_read_cfg().get('idioma', 'pt-BR'), '🇧🇷 Português')
        self._v7_idioma_var = tk.StringVar(value=rot_atual)
        cb = ttk.Combobox(parent, textvariable=self._v7_idioma_var,
                          values=[r for _, r in _V7_IDIOMAS], state='readonly', width=14)
        cb.pack(fill='x', pady=(0, 5))
        cb.bind('<<ComboboxSelected>>', lambda e: self._v7_set_idioma_rot(self._v7_idioma_var.get()))
    except Exception:
        pass
    self._v7_i18n_reg = True

def _v7_set_idioma_rot(self, rot):
    try:
        id_ = 'pt-BR'
        for k, r in _V7_IDIOMAS:
            if r == rot:
                id_ = k
                break
        cfg = _v7_read_cfg()
        cfg['idioma'] = id_
        _v7_write_cfg(cfg)
        self._v7_aplicar_idioma(id_)
    except Exception:
        pass

def _v7_aplicar_idioma(self, id_):
    self._v7_idioma = id_
    tr = _V7_TR.get(id_, {})
    # widgets estáticos
    for w, pt in list(getattr(self, '_v7_map', {}).items()):
        try:
            w.configure(text=tr.get(pt, pt))
        except Exception:
            pass
    # cabeçalhos de coluna
    try:
        for c in self.COLUNAS:
            c['text'] = tr.get(c.get('_pt', c.get('text', '')), c.get('_pt', c.get('text', '')))
        for tb in (getattr(self, 'tree', None), getattr(self, 'resultados_tree', None)):
            if tb is not None:
                try:
                    tb._draw_header()
                    tb._draw_body()
                except Exception:
                    pass
    except Exception:
        pass
    # título da janela
    try:
        self.root.title(tr.get(getattr(self, '_v7_win_title_pt', self.root.title()),
                               getattr(self, '_v7_win_title_pt', self.root.title())))
    except Exception:
        pass
    # botão de tema (depende de tema + idioma)
    self._v7_fix_btn_tema()
    # sincroniza o dropdown
    try:
        if hasattr(self, '_v7_idioma_var'):
            self._v7_idioma_var.set(dict(_V7_IDIOMAS).get(id_, '🇧 Português'))
    except Exception:
        pass

def _v7_fix_btn_tema(self):
    try:
        tema = self.config.get('tema', 'claro')
        pt = '🌙 Tema Escuro' if tema == 'claro' else '☀️ Tema Claro'
        self.btn_tema.config(text=self._v7_t(pt))
    except Exception:
        pass

# ---- relógio com dias por idioma ----
def _v7_relogio(self):
    try:
        agora = datetime.now()
        dias = _V7_DIAS.get(getattr(self, '_v7_idioma', 'pt-BR'), _V7_DIAS['pt-BR'])
        self.lbl_hora.config(text=agora.strftime('%H:%M'))
        self.lbl_data.config(text=agora.strftime('%d/%m/%Y'))
        self.lbl_dia.config(text=dias.get(agora.strftime('%A'), ''))
    except Exception:
        pass
    finally:
        try:
            self.clock_job = self.root.after(1000, self.atualizar_relogio)
        except Exception:
            pass

# ---- wrappers de create_widgets / aplicar_tema / confirmar_saida ----
_V7_CW_ORIG = LivroCatalogApp.create_widgets
def _v7_cw_wrap(self):
    _V7_CW_ORIG(self)
    try:
        self._v7_registrar_idioma()
        self._v7_aplicar_idioma(getattr(self, '_v7_idioma', _v7_read_cfg().get('idioma', 'pt-BR')))
    except Exception:
        pass

_V7_AT_ORIG = LivroCatalogApp.aplicar_tema
def _v7_at_wrap(self):
    _V7_AT_ORIG(self)
    try:
        self._v7_fix_btn_tema()
    except Exception:
        pass

_V7_CS_ORIG = LivroCatalogApp.confirmar_saida
def _v7_cs_wrap(self):
    try:
        self._v7_salvar_layout()
    except Exception:
        pass
    _V7_CS_ORIG(self)

# ---- reatribui (a última definição vence) ----
LivroCatalogApp.__init__ = _v7_new_init
LivroCatalogApp.create_widgets = _v7_cw_wrap
LivroCatalogApp.aplicar_tema = _v7_at_wrap
LivroCatalogApp.confirmar_saida = _v7_cs_wrap
LivroCatalogApp.atualizar_relogio = _v7_relogio
LivroCatalogApp._v7_ajustar_panes = _v7_ajustar_panes
LivroCatalogApp._v7_proporcao_inicial = _v7_proporcao_inicial
LivroCatalogApp._v7_restaurar_layout = _v7_restaurar_layout
LivroCatalogApp._v7_salvar_layout = _v7_salvar_layout
LivroCatalogApp._v7_clamp_geometry = _v7_clamp_geometry
LivroCatalogApp._v7_registrar_idioma = _v7_registrar_idioma
LivroCatalogApp._v7_aplicar_idioma = _v7_aplicar_idioma
LivroCatalogApp._v7_set_idioma_rot = _v7_set_idioma_rot
LivroCatalogApp._v7_fix_btn_tema = _v7_fix_btn_tema
LivroCatalogApp._v7_t = _v7_t
# faz o redimensionador do patch v6 usar a lógica de MÍNIMOS (respeita seu arrasto)
try:
    LivroCatalogApp._v6_ajustar_panes = _v7_ajustar_panes
except Exception:
    pass
# =====================================================================
# FIM PATCH v7  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v8 — LAYOUT: CADASTRO FIXO NA BASE + CATÁLOGO MAIOR +
#              DIVISÃO 50/50 (CATÁLOGO / RESULTADOS)
#  - Cadastro fica compacto, "colado" acima do rodapé (altura do conteúdo)
#  - Catálogo Completo ocupa o máximo de espaço
#  - Com Resultados aberto, o espaço de cima é dividido 50/50
#  - Mantém o ajuste ao redimensionar / buscar / limpar / fechar
#  - Tudo em try/except: se algo falhar, não altera nada e não quebra
# Cola ANTES de:  if __name__ == "__main__":
# =====================================================================

def _v8_adjust(self):
    """Reposiciona os divisores: cadastro fixo embaixo; resto dividido acima."""
    try:
        paned = self.main_paned
        panes = paned.panes()
        n = len(panes)
        if n < 2:
            return
        paned.update_idletasks()
        H = paned.winfo_height()
        if H < 120:
            return

        # O Cadastro é SEMPRE o último painel (Resultados entra no índice 1).
        ff = self.root.nametowidget(panes[-1])

        # Altura "natural" do cadastro (campos + botões), com limites seguros.
        try:
            req = ff.winfo_reqheight()
        except Exception:
            req = 200
        cad_h = max(150, min(int(req) + 12, 280))
        if cad_h > H - 90:
            cad_h = max(120, H - 90)

        top = H - cad_h                 # espaço disponível para Catálogo (+Resultados)
        last = n - 2                    # divisor imediatamente acima do Cadastro

        # 1) "Prega" o Cadastro na base (divisor acima dele = top)
        try:
            paned.sashpos(last, top)
        except Exception:
            pass

        # 2) Se houver Resultados (3 painéis), divide o espaço de cima 50/50
        if n >= 3:
            try:
                paned.sashpos(last - 1, int(top * 0.5))
            except Exception:
                pass
    except Exception:
        pass


def _v8_on_configure(self, event):
    """Ao redimensionar a janela, reaplica o layout (com debounce)."""
    try:
        if event.widget is self.root:
            deb = getattr(self, '_v8_deb', None)
            if deb:
                self.root.after_cancel(deb)
            self._v8_deb = self.root.after(120, self._v8_adjust)
    except Exception:
        pass


# ---- captura as versões atuais (encadeia com patches anteriores) ----
_V8_PREV_INIT    = LivroCatalogApp.__init__
_V8_PREV_BUSCAR  = LivroCatalogApp.buscar
_V8_PREV_FECHAR  = LivroCatalogApp.fechar_painel_resultados
_V8_PREV_LIMPAR  = LivroCatalogApp.limpar_busca


def _v8_wrap_init(self, root):
    _V8_PREV_INIT(self, root)
    # Neutraliza ajustadores de layout antigos (evita briga de sashes).
    for _m in ('_v6_ajustar_panes', '_v7_ajustar_panes',
               '_v7_proporcao_inicial', '_v6_instalar_resize_panes'):
        try:
            setattr(LivroCatalogApp, _m, lambda s: None)
        except Exception:
            pass
    # Instala o nosso controle de layout.
    try:
        self.root.bind('<Configure>', self._v8_on_configure, add='+')
    except Exception:
        pass
    try:
        self.root.after(450, self._v8_adjust)
        self.root.after(1000, self._v8_adjust)
    except Exception:
        pass


def _v8_wrap_buscar(self, *a, **k):
    r = _V8_PREV_BUSCAR(self, *a, **k)
    try:
        self.root.after(80, self._v8_adjust)
    except Exception:
        pass
    return r


def _v8_wrap_fechar(self, *a, **k):
    r = _V8_PREV_FECHAR(self, *a, **k)
    try:
        self.root.after(80, self._v8_adjust)
    except Exception:
        pass
    return r


def _v8_wrap_limpar(self, *a, **k):
    r = _V8_PREV_LIMPAR(self, *a, **k)
    try:
        self.root.after(80, self._v8_adjust)
    except Exception:
        pass
    return r


# ---- reatribui (a última definição vence) ----
LivroCatalogApp.__init__                  = _v8_wrap_init
LivroCatalogApp.buscar                    = _v8_wrap_buscar
LivroCatalogApp.fechar_painel_resultados  = _v8_wrap_fechar
LivroCatalogApp.limpar_busca              = _v8_wrap_limpar
LivroCatalogApp._v8_adjust                = _v8_adjust
LivroCatalogApp._v8_on_configure          = _v8_on_configure
# =====================================================================
# FIM PATCH v8  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v9 — (1) DROPDOWN DE FONTE 8–96 + SCROLL HORIZONTAL NA GRADE
#            (2) CAMPO "SAÍDA" INICIA VAZIO
#            (3) CORRIGE PERSISTÊNCIA DO TEMA (import json) + CONTRASTE DO LOGIN
#  - Grade: fonte escolhida + linhas auto-ajustáveis + barra de rolagem horizontal
#  - Botões de ação têm teto de fonte (não estouram a barra)
#  - "Saída" não é pré-preenchida no cadastro nem no Limpar
#  - import json conserta _carregar_config/_salvar_config (tema passa a salvar)
#  - estilos ttk aplicados ANTES do login (campos legíveis no tema escuro)
#  - TUDO em try/except: se algo falhar, desativa só aquela parte
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================
import json  # <- conserta _carregar_config/_salvar_config do núcleo (estava faltando)

# ---- captura as versões atuais para encadear (compatível com patches v5/v7/v8) ----
_V9_PREV_INIT = LivroCatalogApp.__init__
_V9_PREV_AT   = LivroCatalogApp.aplicar_tema

# ---------------------------------------------------------------------
# (A) TELA DE LOGIN NO TEMA SALVO (contraste correto no escuro)
# ---------------------------------------------------------------------
def _v9_skin_login(t):
    """Configura o ttk.Style com o tema ANTES da tela de login existir."""
    try:
        s = ttk.Style()
        try:
            s.theme_use('clam')
        except Exception:
            pass
        s.configure('TFrame', background=t['bg_janela'])
        s.configure('TLabel', background=t['bg_janela'], foreground=t['texto'])
        s.configure('TLabelframe', background=t['bg_janela'], bordercolor=t['borda'])
        s.configure('TLabelframe.Label', background=t['bg_janela'], foreground=t['titulo'])
        s.configure('TEntry', fieldbackground=t['campo_bg'], foreground=t['campo_fg'],
                    insertcolor=t['campo_fg'], bordercolor=t['borda'])
        s.configure('TCombobox', fieldbackground=t['campo_bg'], foreground=t['campo_fg'],
                    arrowcolor=t['texto'], bordercolor=t['borda'])
        s.map('TCombobox', fieldbackground=[('readonly', t['campo_bg'])],
              foreground=[('readonly', t['campo_fg'])])
        s.configure('TButton', background=t.get('bg_card', t['bg_janela']),
                    foreground=t['texto'], bordercolor=t['borda'])
        s.map('TButton',
              background=[('active', t.get('bg_botao_hover', t['borda']))],
              foreground=[('active', t['texto'])])
    except Exception:
        pass

# ---------------------------------------------------------------------
# (B) TabelaCatalogo: cabeçalho/corpo parametrizados por fonte
#     (cópia fiel do original, trocando o "9" e o "28" por valores dinâmicos)
# ---------------------------------------------------------------------
def _v9_draw_header(self):
    self.header.delete('all')
    x = 0
    t = self.tema
    fs = getattr(self, '_font_size', 9)
    hh = getattr(self, '_header_h', 28)
    for idx, c in enumerate(self.visible_cols):
        w = self.col_widths.get(c['key'], c['base'])
        bg = t['header_par'] if idx % 2 == 0 else t['header_impar']
        self.header.create_rectangle(x, 0, x + w, hh, fill=bg, outline=t['borda'])
        arrow = ''
        if self.sort_col == c['key']:
            arrow = ' ▼' if self.sort_reverse else ' ▲'
        self.header.create_text(
            x + w // 2, hh // 2, text=c['text'] + arrow, anchor='center',
            font=('DejaVu Sans', fs, 'bold'), fill=t['header_texto'])
        x += w
    self.header.configure(scrollregion=(0, 0, x, hh))

def _v9_draw_body(self):
    if not self.winfo_exists():
        return
    self.body.delete('all')
    if not self.visible_cols:
        return
    t = self.tema
    fs = getattr(self, '_font_size', 9)
    rh = self.row_height
    char_px = max(5, int(fs * 0.62))          # largura média de 1 char p/ truncar certo
    total_w = sum(self.col_widths.values())
    total_h = len(self.data) * rh
    cw = max(self.body.winfo_width(), 1)
    ch = max(self.body.winfo_height(), 1)
    self.body.configure(scrollregion=(0, 0, max(total_w, cw), max(total_h, ch)))
    if not self.data:
        self.body.create_text(cw // 2, ch // 2, text="Nenhum registro.",
                              fill=t['texto_suave'], font=('DejaVu Sans', fs))
        return
    y0 = self.body.canvasy(0)
    first = max(0, int(y0 // rh) - 1)
    visible_count = int(ch // rh) + 3
    last = min(len(self.data), first + visible_count)
    for i in range(first, last):
        row = self.data[i]
        y = i * rh
        selected = row.get('id') == self.selected_id
        indisponivel = str(row.get('disponibilidade', '')).strip() == 'Não'
        x = 0
        for idx, c in enumerate(self.visible_cols):
            w = self.col_widths.get(c['key'], c['base'])
            if selected:
                bg = t['selecao']
            elif indisponivel:
                bg = t['indisponivel_par'] if i % 2 == 0 else t['indisponivel_impar']
            else:
                if i % 2 == 0:
                    bg = t['celula_par'] if idx % 2 == 0 else t['celula_impar']
                else:
                    bg = t['celula_par_alt'] if idx % 2 == 0 else t['celula_impar_alt']
            self.body.create_rectangle(x, y, x + w, y + rh, fill=bg, outline=t['borda_cell'])
            val = row.get(c['key'], '')
            txt = '' if val is None else str(val)
            max_chars = max(3, int(w / char_px))
            if len(txt) > max_chars:
                txt = txt[:max_chars - 1] + '…'
            anchor = c.get('anchor', 'w')
            tx = x + w // 2 if anchor == 'center' else x + 4
            if selected:
                cor = t['selecao_texto']
            elif indisponivel:
                cor = t.get('indisponivel_texto', t['texto'])
            else:
                cor = t['texto']
            self.body.create_text(tx, y + rh // 2, text=txt, anchor=anchor,
                                  fill=cor, font=('DejaVu Sans', fs))
            x += w

# ---- rolagem horizontal (sincroniza cabeçalho + corpo) ----
def _v9_scroll_x(self, *a):
    self.body.xview(*a)
    self.header.xview(*a)

def _v9_on_xscroll(self, first, last):
    try:
        self.hbar.set(first, last)
    except Exception:
        pass
    try:
        self.header.xview_moveto(first)
    except Exception:
        pass

def _v9_shift_wheel(self, event):
    try:
        self.body.xview_scroll(int(-1 * (event.delta / 120)), 'units')
    except Exception:
        pass

def _v9_ensure_hbar(self):
    """Adiciona a barra de rolagem horizontal UMA vez, reempacotando na ordem correta."""
    if getattr(self, '_v9_hbar_done', False):
        return
    try:
        self.hbar = ttk.Scrollbar(self, orient='horizontal', command=self._v9_scroll_x)
        self.body.configure(xscrollcommand=self._v9_on_xscroll)
        self.body.bind('<Shift-MouseWheel>', self._v9_shift_wheel, add='+')
        self.body.bind('<Shift-Button-4>', lambda e: self._v9_scroll_x('scroll', -1, 'units'), add='+')
        self.body.bind('<Shift-Button-5>', lambda e: self._v9_scroll_x('scroll', 1, 'units'), add='+')
        # reempacota: topo, base(hbar), direita(vbar), meio(corpo) -> nada some
        for w in (self.header, self.body, self.vbar):
            w.pack_forget()
        self.header.pack(side='top', fill='x')
        self.hbar.pack(side='bottom', fill='x')
        self.vbar.pack(side='right', fill='y')
        self.body.pack(side='left', fill='both', expand=True)
    except Exception:
        pass
    self._v9_hbar_done = True

def _v9_set_font_size(self, size):
    try:
        size = int(size)
    except Exception:
        size = 9
    self._font_size = size
    self.row_height = max(22, int(size * 2.4) + 6)     # linhas crescem com a fonte
    self._header_h = max(24, int(size * 2.0) + 8)
    try:
        self.header.configure(height=self._header_h)
    except Exception:
        pass
    self._v9_ensure_hbar()
    self._draw_header()
    self._draw_body()

# reatribui na classe da tabela (vale p/ instâncias novas e existentes)
TabelaCatalogo._draw_header   = _v9_draw_header
TabelaCatalogo._draw_body     = _v9_draw_body
TabelaCatalogo._v9_scroll_x   = _v9_scroll_x
TabelaCatalogo._v9_on_xscroll = _v9_on_xscroll
TabelaCatalogo._v9_shift_wheel = _v9_shift_wheel
TabelaCatalogo._v9_ensure_hbar = _v9_ensure_hbar
TabelaCatalogo.set_font_size  = _v9_set_font_size

# ---------------------------------------------------------------------
# (C) APLICA A FONTE EM TUDO (com teto nos botões)
# ---------------------------------------------------------------------
def _v9_aplicar_fonte(self, n):
    try:
        n = int(n)
    except Exception:
        n = 10
    n = max(8, min(96, n))
    self._v9_cur_size = n
    try:
        s = ttk.Style()
        s.configure('TLabel', font=('DejaVu Sans', n))
        s.configure('Header.TLabel', font=('DejaVu Sans', n + 2, 'bold'))
        s.configure('Status.TLabel', font=('DejaVu Sans', max(8, n - 1), 'italic'))
        s.configure('Footer.TLabel', font=('DejaVu Sans', max(8, n - 1), 'italic'))
        s.configure('TLabelframe.Label', font=('DejaVu Sans', n, 'bold'))
        s.configure('Treeview', font=('DejaVu Sans', n), rowheight=max(22, int(n * 2.4) + 6))
        s.configure('Treeview.Heading', font=('DejaVu Sans', n, 'bold'))
        s.configure('TEntry', font=('DejaVu Sans', n))
        s.configure('TCombobox', font=('DejaVu Sans', n))
        s.configure('TCheckbutton', font=('DejaVu Sans', n))
        s.configure('TRadiobutton', font=('DejaVu Sans', n))
        bn = min(n, 13)   # teto: barra de botões não pode estourar
        s.configure('TButton', font=('DejaVu Sans', bn, 'bold'))
        for st in ('Excel.TButton', 'Calc.TButton', 'PDF.TButton', 'Print.TButton',
                   'Import.TButton', 'Adicionar.TButton', 'Salvar.TButton',
                   'Excluir.TButton', 'Limpar.TButton', 'Sair.TButton',
                   'Voltar.TButton', 'Emprestar.TButton', 'Devolver.TButton',
                   'Sobre.TButton', 'Ajuda.TButton'):
            s.configure(st, font=('DejaVu Sans', bn, 'bold'))
    except Exception:
        pass
    # grade (células + cabeçalho + scroll horizontal)
    for tb in (getattr(self, 'tree', None), getattr(self, 'resultados_tree', None)):
        if tb is not None and hasattr(tb, 'set_font_size'):
            try:
                tb.set_font_size(n)
            except Exception:
                pass
    try:
        if hasattr(self, '_v9_fontvar'):
            self._v9_fontvar.set(str(n))
    except Exception:
        pass

def _v9_set_from_var(self):
    try:
        v = int(self._v9_fontvar.get())
    except Exception:
        return
    v = max(8, min(96, v))
    self._v9_fontvar.set(str(v))
    self.config['tamanho_fonte'] = v
    try:
        self._salvar_config()
    except Exception:
        pass
    self._v9_aplicar_fonte(v)

def _v9_criar_seletor_fonte(self):
    """Cria o dropdown de fonte no canto superior direito (editável, 8–96)."""
    try:
        parent = self.btn_tema.master
        cur = self.config.get('tamanho_fonte', 10)
        try:
            cur = int(cur)
        except Exception:
            cur = 10
        cur = max(8, min(96, cur))
        self._v9_fontvar = tk.StringVar(value=str(cur))
        vals = [8, 9, 10, 11, 12, 14, 16, 18, 20, 22, 24, 26, 28, 32, 36, 40, 44, 48, 54, 60, 72, 80, 96]
        cb = ttk.Combobox(parent, textvariable=self._v9_fontvar, values=vals,
                          width=4, state='normal')
        try:
            cb.pack(fill='x', pady=(0, 5), before=self.btn_tema)   # fica no topo da coluna
        except Exception:
            cb.pack(fill='x', pady=(0, 5))
        try:
            ToolTip(cb, "Tamanho da fonte (8–96).\nA grade ganha rolagem horizontal e linhas auto-ajustáveis.")
        except Exception:
            pass

        def _commit(e=None):
            try:
                self._v9_set_from_var()
            except Exception:
                pass
        cb.bind('<<ComboboxSelected>>', _commit)
        cb.bind('<Return>', _commit)
        cb.bind('<FocusOut>', _commit)
        self._v9_font_cb = cb
    except Exception:
        pass

# ---------------------------------------------------------------------
# (D) "SAÍDA" INICIA VAZIA (cadastro e Limpar)
# ---------------------------------------------------------------------
def _v9_limpar_form(self):
    for n in self.entries:
        self.entries[n].delete(0, tk.END)
    self.entries['entrada'].insert(0, datetime.now().strftime('%d/%m/%Y'))
    # ✅ SAÍDA NÃO é pré-preenchida (só faz sentido quando o livro sai)
    self.entries['quantidade'].insert(0, '1')
    self.livro_selecionado_id = None

# ---------------------------------------------------------------------
# (E) NEUTRALIZA o auto-scaling de fonte por resize (quem manda é o dropdown)
# ---------------------------------------------------------------------
def _v9_noop_scaling(self, *a, **k):
    pass

# ---------------------------------------------------------------------
# (F) WRAPPERS: aplicar_tema (reaplica fonte) e __init__ (boot completo)
# ---------------------------------------------------------------------
def _v9_new_at(self):
    _V9_PREV_AT(self)
    if getattr(self, '_v9_cur_size', None):
        try:
            self._v9_aplicar_fonte(self._v9_cur_size)
        except Exception:
            pass

def _v9_new_init(self, root):
    # 1) tema no boot ANTES do login (campos legíveis no escuro)
    try:
        if getattr(sys, 'frozen', False):
            d = os.path.dirname(sys.executable)
        else:
            d = os.path.dirname(os.path.abspath(__file__))
        p = os.path.join(d, 'config.json')
        cfg = {}
        if os.path.exists(p):
            with open(p, 'r', encoding='utf-8') as f:
                cfg = json.load(f) or {}
        t = TEMAS.get(cfg.get('tema', 'claro'), TEMAS['claro'])
        globals()['TEMA_ATUAL'] = t
        _v9_skin_login(t)
    except Exception:
        pass
    # 2) init original (login + tela) — já encadeia v5/v7/v8 se existirem
    _V9_PREV_INIT(self, root)
    # 3) pós-boot: saída vazia + seletor de fonte + fonte salva
    try:
        if self.root.winfo_exists():
            if hasattr(self, 'entries') and 'saida' in self.entries:
                self.entries['saida'].delete(0, tk.END)        # ✅ saída vazia
            self._v9_criar_seletor_fonte()
            try:
                self._v9_aplicar_fonte(int(self.config.get('tamanho_fonte', 10)))
            except Exception:
                self._v9_aplicar_fonte(10)
    except Exception:
        pass

# ---- reatribui na aplicação (a última definição vence) ----
LivroCatalogApp.aplicar_tema   = _v9_new_at
LivroCatalogApp.limpar_form    = _v9_limpar_form
LivroCatalogApp._apply_scaling = _v9_noop_scaling
LivroCatalogApp.__init__       = _v9_new_init
LivroCatalogApp._v9_aplicar_fonte      = _v9_aplicar_fonte
LivroCatalogApp._v9_criar_seletor_fonte = _v9_criar_seletor_fonte
LivroCatalogApp._v9_set_from_var       = _v9_set_from_var
# =====================================================================
# FIM PATCH v9  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================


# =====================================================================
# PATCH v10 — (A) COLUNAS SEM SOBREPOR (mede o texto + largura mínima)
#               (B) "EMPRESTADO A" = DROPDOWN CONTEXTUAL + STATUS DE QTD
#  - O texto de cada célula é medido e truncado com "…" => nunca invade a vizinha
#  - Entrada/Saída/Qtd/Disp têm largura mínima garantida (datas aparecem inteiras)
#  - Ao selecionar um livro: campo = último leitor; seta = todos os leitores;
#    e um aviso "📖 N emprestado(s) • M no estoque" liga leitores à quantidade
#  - TUDO em try/except: se algo falhar, desativa só aquele detalhe (não quebra)
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================
import tkinter.font as _v10_tkfont

# ---------------------------------------------------------------------
# (A) TabelaCatalogo: medir texto + larguras mínimas
# ---------------------------------------------------------------------
def _v10_ensure_mfont(self):
    """Cria/recria a fonte usada para MEDIR a largura dos textos."""
    try:
        sz = getattr(self, '_font_size', 9)
        if getattr(self, '_mfont', None) is None or getattr(self, '_mfont_size', None) != sz:
            self._mfont = _v10_tkfont.Font(family='DejaVu Sans', size=sz)
            self._mfont_size = sz
    except Exception:
        self._mfont = None

def _v10_fit(self, text, max_px):
    """Trunca 'text' com '…' até caber em max_px (medido). À prova de fonte."""
    if not text:
        return text
    try:
        self._v10_ensure_mfont()
        if self._mfont is None:
            raise RuntimeError('sem fonte')
        if self._mfont.measure(text) <= max_px:
            return text
        n = len(text)
        est = max(1, int(n * max_px / max(1, self._mfont.measure(text))) - 1)
        while est > 1 and self._mfont.measure(text[:est] + '…') > max_px:
            est -= 1
        while est < n and self._mfont.measure(text[:est + 1] + '…') <= max_px:
            est += 1
        return text[:max(1, est)] + '…'
    except Exception:
        # fallback (estimativa) se medir falhar
        mc = max(3, int(max_px / 8))
        return text if len(text) <= mc else text[:mc - 1] + '…'

def _v10_min_widths(self):
    """Largura mínima por coluna, medida pelo conteúdo típico."""
    self._v10_ensure_mfont()
    def m(s):
        try:
            return self._mfont.measure(s) if self._mfont else len(s) * 8
        except Exception:
            return len(s) * 8
    return {
        'id': m('9999') + 12,
        'entrada': m('00/00/0000') + 12,
        'saida': m('00/00/0000') + 12,
        'quantidade': m('9999') + 12,
        'disponibilidade': m('Não') + 16,
        'estante': m('WWWW') + 12,
        'prateleira': m('WWWWW') + 12,
        'titulo': 70, 'autor': 60, 'editora': 60, 'assunto': 60,
        'bibliotecario': 60, 'emprestado_a': 80,
    }

def _v10_update_visible_columns(self):
    """Recalcula colunas visíveis + aplica larguras mínimas + distribui sobra."""
    try:
        avail = self.body.winfo_width()
        if avail < 50:
            self.visible_cols = self.columns[:]
            self.col_widths = {c['key']: c['base'] for c in self.columns}
            return
        essential = [c for c in self.columns if c.get('priority', 1) == 0]
        optional = [c for c in self.columns if c.get('priority', 1) != 0]
        optional.sort(key=lambda c: (c.get('priority', 1), self.columns.index(c)))
        chosen = essential[:]
        total = sum(c['base'] for c in chosen)
        for c in optional:
            if total + c['base'] <= avail:
                chosen.append(c)
                total += c['base']
        self.visible_cols = [c for c in self.columns if c in chosen] or self.columns[:]

        base = {c['key']: c['base'] for c in self.visible_cols}
        mw = self._v10_min_widths()
        with_min = {k: max(base[k], mw.get(k, 0)) for k in base}
        # só usa os mínimos se couberem; senão fica no base (o _fit trunca sem trepar)
        self.col_widths = with_min if sum(with_min.values()) <= avail else base

        extra = avail - sum(self.col_widths.values())
        if extra > 0:
            stretch = [c for c in self.visible_cols if c.get('stretch')]
            if not stretch:
                stretch = [c for c in self.visible_cols if c['key'] in ('titulo', 'autor')]
            if not stretch:
                stretch = self.visible_cols[:]
            weights = [self.col_widths[c['key']] for c in stretch]
            soma = sum(weights) or 1
            used = 0
            for i, c in enumerate(stretch):
                add = int(extra * weights[i] / soma)
                self.col_widths[c['key']] += add
                used += add
            resto = extra - used
            if resto > 0:
                self.col_widths[stretch[0]['key']] += resto
    except Exception:
        # em caso de erro, mantém o comportamento simples
        self.col_widths = {c['key']: c['base'] for c in getattr(self, 'visible_cols', self.columns)}

def _v10_draw_header(self):
    try:
        self.header.delete('all')
        t = self.tema
        fs = getattr(self, '_font_size', 9)
        hh = getattr(self, '_header_h', 28)
        x = 0
        for idx, c in enumerate(self.visible_cols):
            w = self.col_widths.get(c['key'], c['base'])
            bg = t['header_par'] if idx % 2 == 0 else t['header_impar']
            self.header.create_rectangle(x, 0, x + w, hh, fill=bg, outline=t['borda'])
            arrow = ''
            if self.sort_col == c['key']:
                arrow = ' ▼' if self.sort_reverse else ' ▲'
            txt = self._v10_fit(c['text'] + arrow, max(20, w - 6))
            self.header.create_text(x + w // 2, hh // 2, text=txt, anchor='center',
                                    font=('DejaVu Sans', fs, 'bold'), fill=t['header_texto'])
            x += w
        self.header.configure(scrollregion=(0, 0, x, hh))
    except Exception:
        pass

def _v10_draw_body(self):
    try:
        if not self.winfo_exists():
            return
        self.body.delete('all')
        if not self.visible_cols:
            return
        t = self.tema
        fs = getattr(self, '_font_size', 9)
        rh = self.row_height
        total_w = sum(self.col_widths.values())
        total_h = len(self.data) * rh
        cw = max(self.body.winfo_width(), 1)
        ch = max(self.body.winfo_height(), 1)
        self.body.configure(scrollregion=(0, 0, max(total_w, cw), max(total_h, ch)))
        if not self.data:
            self.body.create_text(cw // 2, ch // 2, text="Nenhum registro.",
                                  fill=t['texto_suave'], font=('DejaVu Sans', fs))
            return
        y0 = self.body.canvasy(0)
        first = max(0, int(y0 // rh) - 1)
        visible_count = int(ch // rh) + 3
        last = min(len(self.data), first + visible_count)
        for i in range(first, last):
            row = self.data[i]
            y = i * rh
            selected = row.get('id') == self.selected_id
            indisponivel = str(row.get('disponibilidade', '')).strip() == 'Não'
            x = 0
            for idx, c in enumerate(self.visible_cols):
                w = self.col_widths.get(c['key'], c['base'])
                if selected:
                    bg = t['selecao']
                elif indisponivel:
                    bg = t['indisponivel_par'] if i % 2 == 0 else t['indisponivel_impar']
                else:
                    if i % 2 == 0:
                        bg = t['celula_par'] if idx % 2 == 0 else t['celula_impar']
                    else:
                        bg = t['celula_par_alt'] if idx % 2 == 0 else t['celula_impar_alt']
                self.body.create_rectangle(x, y, x + w, y + rh, fill=bg, outline=t['borda_cell'])
                val = row.get(c['key'], '')
                txt = '' if val is None else str(val)
                txt = self._v10_fit(txt, max(12, w - 8))   # ✅ mede e trunca => nunca trepa
                anchor = c.get('anchor', 'w')
                tx = x + w // 2 if anchor == 'center' else x + 4
                if selected:
                    cor = t['selecao_texto']
                elif indisponivel:
                    cor = t.get('indisponivel_texto', t['texto'])
                else:
                    cor = t['texto']
                self.body.create_text(tx, y + rh // 2, text=txt, anchor=anchor,
                                      fill=cor, font=('DejaVu Sans', fs))
                x += w
    except Exception:
        pass

# reatribui na classe da tabela
TabelaCatalogo._v10_ensure_mfont        = _v10_ensure_mfont
TabelaCatalogo._v10_fit                 = _v10_fit
TabelaCatalogo._v10_min_widths          = _v10_min_widths
TabelaCatalogo._update_visible_columns  = _v10_update_visible_columns
TabelaCatalogo._draw_header             = _v10_draw_header
TabelaCatalogo._draw_body               = _v10_draw_body

# mantém o seletor de fonte (se existir) sincronizado com a medição
_v10_prev_sf = getattr(TabelaCatalogo, 'set_font_size', None)
def _v10_set_font_size(self, size):
    if _v10_prev_sf:
        try:
            _v10_prev_sf(self, size)
        except Exception:
            self._font_size = size
            self.row_height = max(22, int(size * 2.4) + 6)
    else:
        self._font_size = size
        self.row_height = max(22, int(size * 2.4) + 6)
    self._mfont = None                       # força recriar na próxima medição
    self._v10_ensure_mfont()
    try: self._update_visible_columns()
    except Exception: pass
    try: self._draw_header()
    except Exception: pass
    try: self._draw_body()
    except Exception: pass
TabelaCatalogo.set_font_size = _v10_set_font_size

# ---------------------------------------------------------------------
# (B) "Emprestado a" = dropdown contextual + status de quantidade
# ---------------------------------------------------------------------
def _v10_atualiza_status_emp(self, n, qtd):
    """Cria (uma vez) e atualiza o aviso '📖 N emprestado(s) • M no estoque'."""
    try:
        if not hasattr(self, '_v10_emp_lbl'):
            ff = self.entries['titulo'].master.master   # ef -> ff (frame do formulário)
            lbl = tk.Label(ff, text='', font=('DejaVu Sans', 9, 'bold'),
                           fg='#7a8699', anchor='w')
            lbl.grid(row=2, column=6, columnspan=2, sticky='w', padx=8, pady=5)
            self._v10_emp_lbl = lbl
        if n:
            self._v10_emp_lbl.config(text=f"📖 {n} emprestado(s)  •  {qtd} no estoque")
        else:
            self._v10_emp_lbl.config(text=f"📖 nenhum empréstimo  •  {qtd} no estoque")
    except Exception:
        pass

_v10_prev_cfid = LivroCatalogApp._carregar_form_por_id
def _v10_cfid(self, livro_id):
    self._emp_lista_livro = []                 # zera antes (segurança)
    _v10_prev_cfid(self, livro_id)
    try:
        l = next((x for x in self.livros if x['id'] == livro_id), None)
        if not l and getattr(self, 'ultima_busca', []):
            l = next((x for x in self.ultima_busca if x['id'] == livro_id), None)
        if not l:
            return
        raw = str(l.get('emprestado_a', '') or '')
        lista = [n.strip() for n in raw.split(';') if n.strip()]
        self._emp_lista_livro = lista
        # seta = leitores atuais (último primeiro) + histórico, sem repetir
        try:
            hist = list(self.entries['emprestado_a'].cget('values') or [])
        except Exception:
            hist = []
        vals = []
        for n in lista[::-1]:
            if n and n not in vals:
                vals.append(n)
        for n in hist:
            if n and n not in vals:
                vals.append(n)
        self.entries['emprestado_a']['values'] = vals
        # campo = último leitor
        self.entries['emprestado_a'].delete(0, tk.END)
        if lista:
            self.entries['emprestado_a'].insert(0, lista[-1])
        self._v10_atualiza_status_emp(len(lista), l.get('quantidade', 0))
    except Exception:
        pass
LivroCatalogApp._carregar_form_por_id = _v10_cfid

_v10_prev_limpar = LivroCatalogApp.limpar_form
def _v10_limpar(self, *a, **k):
    r = _v10_prev_limpar(self, *a, **k)
    try:
        self._emp_lista_livro = []
        if hasattr(self, '_v10_emp_lbl'):
            self._v10_emp_lbl.config(text='')
        self.atualizar_comboboxes()           # values volta ao histórico geral
    except Exception:
        pass
    return r
LivroCatalogApp.limpar_form = _v10_limpar

_v10_prev_salvar = LivroCatalogApp.salvar
def _v10_salvar(self, *a, **k):
    # reconstrói a lista completa a partir do que o usuário deixou no campo,
    # SEM perder os demais leitores (só o último fica visível no campo)
    try:
        if getattr(self, 'livro_selecionado_id', None) is not None \
                and hasattr(self, '_emp_lista_livro'):
            lista = list(getattr(self, '_emp_lista_livro', []) or [])
            val = self.entries['emprestado_a'].get().strip()
            if val == '':
                nova = ''                       # apagou => devolveu todos
            elif val in lista:
                nova = '; '.join(lista)         # só navegou => mantém o conjunto
            else:
                nova = '; '.join(lista + [val]) # nome novo => adiciona
            self.entries['emprestado_a'].delete(0, tk.END)
            self.entries['emprestado_a'].insert(0, nova)
    except Exception:
        pass
    return _v10_prev_salvar(self, *a, **k)
LivroCatalogApp.salvar = _v10_salvar
# =====================================================================
# FIM PATCH v10  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v11 — (A) COLUNAS REDIMENSIONÁVEIS + SALVAS  (B) BOTÃO ✖ NO TOPO
#  (A) Arraste a BORDA DIREITA de um cabeçalho para redimensionar a coluna.
#      Ao soltar, as larguras são gravadas em config.json (chave col_widths)
#      e restauradas na próxima abertura.
#  (B) O botão "✖ Fechar" do painel de resultados passa para o canto
#      superior direito (linha do título), nunca ficando escondido.
#  - Encadeia com o que já existe (v3/v5/v6/v7/v8/v9/v10): só acrescenta.
#  - TUDO em try/except: se algo falhar, desativa só aquele detalhe.
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================
import json as _v11_json

# ---- leitura / gravação das larguras salvas (independente da app) ----
def _v11_load_widths(path):
    if not path:
        return None
    try:
        if not os.path.exists(path):
            return None
        with open(path, 'r', encoding='utf-8') as f:
            d = _v11_json.load(f) or {}
        cw = d.get('col_widths')
        if not isinstance(cw, dict) or not cw:
            return None
        out = {}
        for k, v in cw.items():
            try:
                iv = int(float(v))
                if iv >= 20:
                    out[str(k)] = iv
            except Exception:
                pass
        return out or None
    except Exception:
        return None

def _v11_save_widths(path, widths):
    if not path:
        return
    try:
        d = {}
        if os.path.exists(path):
            with open(path, 'r', encoding='utf-8') as f:
                d = _v11_json.load(f) or {}
        d['col_widths'] = {str(k): int(v) for k, v in widths.items()}
        with open(path, 'w', encoding='utf-8') as f:
            _v11_json.dump(d, f, indent=2, ensure_ascii=False)
    except Exception:
        pass

# ---- captura as versões ATUAIS (encadeia com v10/original sem perder nada) ----
_V11_PREV_UC    = TabelaCatalogo._update_visible_columns
_V11_PREV_DRAW  = TabelaCatalogo._draw_header
_V11_PREV_CLICK = TabelaCatalogo._on_header_click

# ---------------------------------------------------------------------
# (A) REDIMENSIONAR COLUNAS
# ---------------------------------------------------------------------
def _v11_hit_grip(self, x):
    """Retorna a chave da coluna cuja alça (borda direita) está sob x."""
    for (x0, x1, key) in getattr(self, '_header_bounds', []):
        if abs(x - x1) <= 5:
            return key
    return None

def _v11_header_motion(self, event):
    """Muda o cursor para ↔ quando o mouse passa sobre uma alça."""
    try:
        if getattr(self, '_v11_resizing', None):
            self.header.configure(cursor='sb_h_double')
            return
        if self._v11_hit_grip(event.x) is not None:
            self.header.configure(cursor='sb_h_double')
        else:
            self.header.configure(cursor='')
    except Exception:
        pass

def _v11_header_press(self, event):
    """Clique na alça = inicia resize; clique no meio = ordena (original)."""
    try:
        key = self._v11_hit_grip(event.x)
    except Exception:
        key = None
    if key is not None:
        self._v11_resizing = key
        self._v11_start_x = event.x
        self._v11_start_w = self.col_widths.get(key, 80)
        return  # NÃO ordena
    _V11_PREV_CLICK(self, event)  # comportamento original (ordenar)

def _v11_header_drag(self, event):
    """Arrasto: ajusta a largura da coluna e redesenha."""
    key = getattr(self, '_v11_resizing', None)
    if not key:
        return
    try:
        dx = event.x - self._v11_start_x
        nw = max(28, int(self._v11_start_w + dx))
        self.col_widths[key] = nw
        self._draw_header()  # recompute bounds (mantém a alça sob o mouse)
        self._draw_body()
    except Exception:
        pass

def _v11_header_release(self, event):
    """Solta: congela as larguras de TODAS as colunas e salva no config."""
    key = getattr(self, '_v11_resizing', None)
    if not key:
        return
    try:
        self._v11_resizing = None
        self.header.configure(cursor='')
        widths = {}
        for c in self.columns:
            k = c['key']
            widths[k] = int(self.col_widths.get(k, c.get('base', 80)))
        self._v11_override = widths
        _v11_save_widths(getattr(self, '_v11_cfg_path', None), widths)
    except Exception:
        pass

def _v11_draw_header(self):
    """Desenha (preservando o v10: fonte/altura) + guarda as alças + instala binds 1x."""
    _V11_PREV_DRAW(self)
    if not getattr(self, '_v11_binds', False):
        try:
            self.header.bind('<Motion>', self._v11_header_motion, add='+')
            self.header.bind('<B1-Motion>', self._v11_header_drag, add='+')
            self.header.bind('<ButtonRelease-1>', self._v11_header_release, add='+')
        except Exception:
            pass
        self._v11_binds = True
    # computa as posições x de cada coluna (sincronizadas com o desenho)
    bounds = []
    x = 0
    for c in self.visible_cols:
        w = self.col_widths.get(c['key'], c.get('base', 80))
        bounds.append((x, x + w, c['key']))
        x += w
    self._header_bounds = bounds

def _v11_update_visible_columns(self):
    """Roda o cálculo original/v10 e, se houver larguras salvas, aplica-as."""
    _V11_PREV_UC(self)
    ov = getattr(self, '_v11_override', None)
    if ov:
        for k in list(self.col_widths.keys()):
            if k in ov:
                try:
                    self.col_widths[k] = max(28, int(ov[k]))
                except Exception:
                    pass

# reatribui na classe da tabela
TabelaCatalogo._update_visible_columns = _v11_update_visible_columns
TabelaCatalogo._draw_header            = _v11_draw_header
TabelaCatalogo._on_header_click        = _v11_header_press
TabelaCatalogo._v11_hit_grip           = _v11_hit_grip
TabelaCatalogo._v11_header_motion      = _v11_header_motion
TabelaCatalogo._v11_header_press       = _v11_header_press
TabelaCatalogo._v11_header_drag        = _v11_header_drag
TabelaCatalogo._v11_header_release     = _v11_header_release

# ---------------------------------------------------------------------
# (B) BOTÃO ✖ NO TOPO DIREITO  +  boot das larguras salvas
# ---------------------------------------------------------------------
def _v11_fix_close_btn(self):
    """Move o '✖ Fechar' do painel de resultados para a linha do título."""
    try:
        rf = getattr(self, 'resultados_frame', None)
        if rf is None:
            return
        btn = None
        rc = None
        for w in rf.winfo_children():
            if isinstance(w, (tk.Button, ttk.Button)):
                try:
                    t = str(w.cget('text'))
                except Exception:
                    t = ''
                if ('Fechar' in t) or ('✖' in t):
                    btn = w
            elif isinstance(w, ttk.Frame):
                rc = w
        if btn is not None:
            for forget in ('pack_forget', 'grid_forget', 'place_forget'):
                try:
                    getattr(btn, forget)()
                except Exception:
                    pass
            try:
                btn.place(relx=1.0, x=-8, y=3, anchor='ne')  # canto sup. direito
                btn.lift()
            except Exception:
                pass
        if rc is not None:
            try:
                rc.pack_configure(pady=(28, 0))  # reserva a faixa do topo p/ o botão
            except Exception:
                pass
    except Exception:
        pass

def _v11_reaplicar_layout(self):
    """Reposiciona o botão e reaplica as larguras (após o layout estabilizar)."""
    self._v11_fix_close_btn()
    for tb in (getattr(self, 'tree', None), getattr(self, 'resultados_tree', None)):
        if tb is None:
            continue
        try:
            tb._update_visible_columns()
            tb._draw_header()
            tb._draw_body()
        except Exception:
            pass

_V11_PREV_APP_INIT = LivroCatalogApp.__init__

def _v11_app_init(self, root):
    _V11_PREV_APP_INIT(self, root)
    # injeta o caminho do config + as larguras salvas em cada tabela
    try:
        path = getattr(self, 'config_file', None)
        ov = _v11_load_widths(path)
        for tb in (getattr(self, 'tree', None), getattr(self, 'resultados_tree', None)):
            if tb is None:
                continue
            tb._v11_cfg_path = path
            tb._v11_override = ov
    except Exception:
        pass
    # após o layout assentar, reposiciona o botão e aplica as larguras
    try:
        self.root.after(120, lambda: self._v11_reaplicar_layout())
    except Exception:
        pass

LivroCatalogApp.__init__              = _v11_app_init
LivroCatalogApp._v11_fix_close_btn    = _v11_fix_close_btn
LivroCatalogApp._v11_reaplicar_layout = _v11_reaplicar_layout
# =====================================================================
# FIM PATCH v11  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v10 (reentrega) — "EMPRESTADO A" = DROPDOWN NO FORMULÁRIO
#  Ao selecionar um livro: campo = último leitor; setinha = leitores
#  atuais (do + recente ao + antigo) + histórico; e um aviso liga os
#  leitores à quantidade: "📖 N emprestado(s) • M no estoque".
#  O banco continua guardando a lista COMPLETA (não perde leitores).
#  Idempotente: se já aplicado, não reaplica. Cola ANTES de __main__.
# =====================================================================
if not getattr(LivroCatalogApp, '_v10_done', False):

    _v10_orig_cfid   = LivroCatalogApp._carregar_form_por_id
    _v10_orig_limpar = LivroCatalogApp.limpar_form
    _v10_orig_salvar = LivroCatalogApp.salvar

    def _v10_status(self, n, qtd):
        """Cria (1x) e atualiza o aviso de empréstimos ao lado do campo."""
        try:
            if not hasattr(self, '_v10_emp_lbl'):
                ff = self.entries['titulo'].master.master   # ef -> ff
                lbl = tk.Label(ff, text='', font=('DejaVu Sans', 9, 'bold'),
                               fg='#7a8699', anchor='w')
                lbl.grid(row=2, column=6, columnspan=2, sticky='w', padx=8, pady=5)
                self._v10_emp_lbl = lbl
            if n:
                self._v10_emp_lbl.config(text=f"📖 {n} emprestado(s)  •  {qtd} no estoque")
            else:
                self._v10_emp_lbl.config(text=f"📖 nenhum empréstimo  •  {qtd} no estoque")
        except Exception:
            pass

    def _v10_cfid(self, livro_id):
        self._emp_lista_livro = []
        _v10_orig_cfid(self, livro_id)
        try:
            l = next((x for x in self.livros if x['id'] == livro_id), None)
            if not l and getattr(self, 'ultima_busca', []):
                l = next((x for x in self.ultima_busca if x['id'] == livro_id), None)
            if not l:
                return
            lista = [n.strip() for n in str(l.get('emprestado_a', '') or '').split(';') if n.strip()]
            self._emp_lista_livro = lista
            # setinha = leitores atuais (último primeiro) + histórico, sem repetir
            try:
                hist = list(self.entries['emprestado_a'].cget('values') or [])
            except Exception:
                hist = []
            vals = []
            for n in lista[::-1]:
                if n and n not in vals:
                    vals.append(n)
            for n in hist:
                if n and n not in vals:
                    vals.append(n)
            self.entries['emprestado_a']['values'] = vals
            # campo = último leitor
            self.entries['emprestado_a'].delete(0, tk.END)
            if lista:
                self.entries['emprestado_a'].insert(0, lista[-1])
            self._v10_status(len(lista), l.get('quantidade', 0))
        except Exception:
            pass

    def _v10_limpar(self, *a, **k):
        r = _v10_orig_limpar(self, *a, **k)
        try:
            self._emp_lista_livro = []
            if hasattr(self, '_v10_emp_lbl'):
                self._v10_emp_lbl.config(text='')
            self.atualizar_comboboxes()
        except Exception:
            pass
        return r

    def _v10_salvar(self, *a, **k):
        # reconstrói a lista completa a partir do campo, SEM perder leitores
        try:
            if getattr(self, 'livro_selecionado_id', None) is not None \
                    and hasattr(self, '_emp_lista_livro'):
                lista = list(getattr(self, '_emp_lista_livro', []) or [])
                val = self.entries['emprestado_a'].get().strip()
                if val == '':
                    nova = ''                          # apagou => devolveu todos
                elif val in lista:
                    nova = '; '.join(lista)            # só navegou => mantém conjunto
                else:
                    nova = '; '.join(lista + [val])    # nome novo => adiciona
                self.entries['emprestado_a'].delete(0, tk.END)
                self.entries['emprestado_a'].insert(0, nova)
        except Exception:
            pass
        return _v10_orig_salvar(self, *a, **k)

    LivroCatalogApp._carregar_form_por_id = _v10_cfid
    LivroCatalogApp.limpar_form           = _v10_limpar
    LivroCatalogApp.salvar                = _v10_salvar
    LivroCatalogApp._v10_done             = True
# =====================================================================

# =====================================================================
# PATCH v13 — LAYOUT COMPACTO PARA MONITORES ANTIGOS (800x600 / 1024x768)
#  Objetivo: dar o MÁXIMO de espaço ao "Catálogo Completo".
#  (1) Botões Tema/Ajuda/Sobre/Fonte/Idioma vão para UMA LINHA (horizontal),
#      o que "baixa" o cabeçalho de ~150px para ~70px.
#  (2) Quadros "Busca e Filtros" e "Exportar..." ficam com padding mínimo;
#      em telas estreitas a linha de dica some sozinha.
#  (3) Cabeçalho responsivo: o título re-quebra por largura e o relógio
#      se esconde abaixo de 900px, para o topo nunca estourar a altura.
#  - Reorganiza widgets JÁ EXISTENTES (não reescreve o create_widgets),
#    logo NÃO afeta banco / salvar / importar / tema / idioma / fonte.
#  - Cada passo em try/except: se algo falhar, pula só aquele detalhe.
#  - Auto-aplica no boot e a cada redimensionamento (debounce).
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================

# --- helpers (sem self) -------------------------------------------------
def _v13_walk(w):
    """Percorre a árvore de widgets a partir de w."""
    yield w
    try:
        for c in w.winfo_children():
            for x in _v13_walk(c):
                yield x
    except Exception:
        pass

def _v13_is_idioma(w):
    """Diz se um Combobox é o seletor de IDIOMA (senão, é o de fonte)."""
    try:
        vals = [str(x) for x in w.cget('values')]
    except Exception:
        return False
    marcas = ('Portugu', 'English', 'Espa', 'Esperanto', 'Fran', 'Deutsch')
    return any(m in v for v in vals for m in marcas)

# --- (1) botões na horizontal ------------------------------------------
def _v13_reorganizar_botoes(self):
    """Tira os 5 controles da coluna vertical e os põe numa linha só."""
    if getattr(self, '_v13_btns_done', False):
        return
    self._v13_btns_done = True
    try:
        hbtns = self.btn_tema.master              # frame que hoje é a coluna
        filhos = list(hbtns.winfo_children())

        # identifica cada controle (à prova da ordem em que foram criados)
        found = {'tema': self.btn_tema, 'ajuda': None, 'sobre': None,
                 'fonte': None, 'idioma': None}
        for w in filhos:
            if w is self.btn_tema:
                continue
            if isinstance(w, ttk.Combobox):
                if _v13_is_idioma(w):
                    found['idioma'] = w
                else:
                    found['fonte'] = w
            elif isinstance(w, tk.Button):
                txt = str(w.cget('text'))
                if ('Ajuda' in txt) or ('❓' in txt):
                    found['ajuda'] = w
                elif ('Sobre' in txt) or ('ℹ' in txt):
                    found['sobre'] = w

        # desempacota tudo e reempacota na horizontal (ordem pedida)
        for w in filhos:
            try:
                w.pack_forget()
            except Exception:
                pass

        colocados = set()
        for k in ('tema', 'ajuda', 'sobre', 'fonte', 'idioma'):
            w = found.get(k)
            if w:
                try:
                    w.pack(side='left', padx=2, pady=2)
                    colocados.add(id(w))
                except Exception:
                    pass
        # segurança: qualquer widget "sobrando" também vai p/ linha (não some)
        for w in filhos:
            if id(w) not in colocados and w is not self.btn_tema:
                try:
                    w.pack(side='left', padx=2, pady=2)
                except Exception:
                    pass

        # o frame vira uma linha à direita, bem fina
        try:
            hbtns.pack_configure(side='right', padx=6, pady=2)
        except Exception:
            pass
    except Exception:
        pass

# --- localiza quadros / título (sem depender de nomes de variável) -----
def _v13_find_lf(self, substr):
    """Acha um LabelFrame pelo texto do título (ex.: 'Busca')."""
    try:
        main = self.symbol_canvas.master.master   # header.master == main
        for w in _v13_walk(main):
            if isinstance(w, ttk.LabelFrame) and substr in str(w.cget('text')):
                return w
    except Exception:
        pass
    return None

def _v13_localizar_titulo(self):
    """Guarda referência do Label do título (para re-quebrar depois)."""
    try:
        header = self.symbol_canvas.master
        hbtns = self.btn_tema.master
        tf = None
        for w in header.winfo_children():
            if isinstance(w, ttk.Frame) and w is not hbtns:
                tf = w
                break
        if tf:
            for w in tf.winfo_children():
                if isinstance(w, tk.Label) and 'Fraternidade' in str(w.cget('text')):
                    self._v13_titulo = w
                    return
    except Exception:
        pass
    self._v13_titulo = None

# --- (2) compacta os quadros -------------------------------------------
def _v13_aplicar_compactacao(self):
    """Reduz paddings do main / header / busca / exportar (1x só)."""
    if getattr(self, '_v13_pad_done', False):
        return
    self._v13_pad_done = True
    try:
        self.symbol_canvas.master.master.configure(padding=4)   # main: 10 -> 4
    except Exception:
        pass
    try:
        self.symbol_canvas.master.pack_configure(pady=(0, 4))   # header
    except Exception:
        pass
    lf_b = self._v13_find_lf('Busca')
    if lf_b:
        try:
            lf_b.configure(padding=2)
        except Exception:
            pass
    lf_e = self._v13_find_lf('Exportar')
    if lf_e:
        try:
            lf_e.configure(padding=2)
        except Exception:
            pass
        # guarda a linha de dica (*PDF abre...) para esconder em telas pequenas
        for w in _v13_walk(lf_e):
            if isinstance(w, ttk.Label) and str(w.cget('text')).startswith('*PDF'):
                self._v13_dica = w
                break

# --- (3) cabeçalho responsivo por largura ------------------------------
def _v13_layout(self):
    """Re-quebra o título e esconde relógio/dica conforme a largura."""
    try:
        larg = self.root.winfo_width()
    except Exception:
        return
    if larg < 50:
        return

    # título: degraus de quebra (sem geometria fina -> robusto)
    t = getattr(self, '_v13_titulo', None)
    if t:
        try:
            if larg < 950:
                t.configure(wraplength=200)
            elif larg < 1300:
                t.configure(wraplength=380)
            else:
                t.configure(wraplength=900)
        except Exception:
            pass

    # relógio: some abaixo de 900px (não é crítico p/ gerir os livros)
    rf = getattr(self, 'relogio_frame', None)
    if rf:
        try:
            if larg < 900:
                rf.pack_forget()
            else:
                rf.pack(pady=(2, 2))
        except Exception:
            pass

    # dica do PDF: some abaixo de 1000px (grid_remove lembra a posição)
    dica = getattr(self, '_v13_dica', None)
    if dica:
        try:
            if larg < 1000:
                dica.grid_remove()
            else:
                dica.grid()
        except Exception:
            pass

def _v13_on_configure(self, event):
    """Reaplica o layout ao redimensionar (com debounce)."""
    try:
        if event.widget is self.root:
            d = getattr(self, '_v13_deb', None)
            if d:
                self.root.after_cancel(d)
            self._v13_deb = self.root.after(150, self._v13_layout)
    except Exception:
        pass

# --- boot: encadeia por cima do __init__ atual (seja qual for) ---------
def _v13_new_init(self, root):
    _V13_PREV_INIT(self, root)
    if not self.root.winfo_exists():
        return
    try:
        self._v13_reorganizar_botoes()
    except Exception:
        pass
    try:
        self._v13_localizar_titulo()
    except Exception:
        pass
    try:
        self._v13_aplicar_compactacao()
    except Exception:
        pass
    try:
        # add='+' -> não sobrescreve os handlers de Configure dos patches anteriores
        self.root.bind('<Configure>', self._v13_on_configure, add='+')
        self.root.after(300, self._v13_layout)
        self.root.after(800, self._v13_layout)
    except Exception:
        pass

_V13_PREV_INIT = LivroCatalogApp.__init__
LivroCatalogApp.__init__                 = _v13_new_init
LivroCatalogApp._v13_reorganizar_botoes  = _v13_reorganizar_botoes
LivroCatalogApp._v13_find_lf             = _v13_find_lf
LivroCatalogApp._v13_localizar_titulo    = _v13_localizar_titulo
LivroCatalogApp._v13_aplicar_compactacao = _v13_aplicar_compactacao
LivroCatalogApp._v13_layout              = _v13_layout
LivroCatalogApp._v13_on_configure        = _v13_on_configure
# =====================================================================
# FIM PATCH v13  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v14 — (1) EXE SALVA DB/CONFIG/TEMA/IDIOMA  (2) CANCELAR NÃO CRASHA
#  - Reescreve __file__ (global do módulo) p/ a pasta do .exe quando
#    congelado, com fallback p/ Documentos se a pasta não for gravável.
#    => banco, config.json, tema, fonte e idioma passam a persistir.
#  - Torna o boot à prova de "janela já destruída" (Cancelar no login).
#  - Idempotente e todo em try/except: se falhar, não quebra o resto.
# Cola ANTES de:  if __name__ == "__main__":
# =====================================================================
def _v14_gravavel(p):
    try:
        os.makedirs(p, exist_ok=True)
        t = os.path.join(p, '.feeu_write_test')
        with open(t, 'w', encoding='utf-8') as f:
            f.write('1')
        os.remove(t)
        return True
    except Exception:
        return False

if getattr(sys, 'frozen', False):
    try:
        _v14_base = os.path.dirname(os.path.abspath(sys.executable))   # pasta do .exe
        if not _v14_gravavel(_v14_base):
            _v14_base = (os.path.join(os.path.expanduser('~'), 'Documents', 'FEEU_Catalogo')
                         if os.name == 'nt' else os.path.join(os.path.expanduser('~'), '.feeucatalogo'))
            _v14_gravavel(_v14_base)
        # reescreve a GLOBAL __file__ => o __init__ usa a pasta certa
        __file__ = os.path.join(_v14_base, 'Catalogo_de_Livros.py')
    except Exception:
        pass

# ---- boot à prova de "root já destruído" (corrige o Cancelar) ----
_V14_PREV_INIT = LivroCatalogApp.__init__

def _v14_safe_init(self, root):
    try:
        _V14_PREV_INIT(self, root)
    except Exception:
        # se o init abortou (ex.: login cancelado já destruiu o root), não faz mais nada
        return
    # só mexe no layout se a janela ainda estiver viva
    try:
        if not root.winfo_exists():
            return
    except Exception:
        return
    for _m in ('_v13_reorganizar_botoes', '_v13_localizar_titulo',
               '_v13_aplicar_compactacao'):
        try:
            getattr(self, _m)()
        except Exception:
            pass
    try:
        root.after(300, self._v13_layout)
        root.after(800, self._v13_layout)
    except Exception:
        pass

LivroCatalogApp.__init__ = _v14_safe_init
# =====================================================================

# =====================================================================
# PATCH v16 — (A) CARET DESTACADO POR TEMA  (B) "GLOW" NOS CAMPOS VAZIOS
#               (C) REMOVE OS BOTÕES ➕ DE ESTANTE / PRATELEIRA
#  - Caret mais grosso (insertwidth=2) e com cor de contraste por tema.
#  - Campo VAZIO ganha fundo âmbar ("preencha-me"); some ao preencher.
#    (O caret piscante é único por natureza; o glow cobre "todos os campos".)
#  - Some o ➕ de Estante/Prateleira (mantém o combobox p/ padronizar).
#  - Idempotente e todo em try/except: se falhar, não quebra o resto.
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================

# --- cores do destaque, por tema -------------------------------------
_V16_CARET   = {'claro': '#b45309', 'escuro': '#fbbf24'}   # cor do cursor
_V16_VAZIO   = {'claro': '#fde68a', 'escuro': '#78350f'}   # fundo do campo vazio

def _v16_nome_tema(self):
    try:
        return (self.config or {}).get('tema', 'claro')
    except Exception:
        return 'claro'

def _v16_style_caret(self):
    """Caret mais grosso + cor de contraste (aplica no style, vale p/ todos)."""
    try:
        caret = _V16_CARET.get(self._v16_nome_tema(), _V16_CARET['claro'])
        s = ttk.Style()
        s.configure('TEntry',    insertwidth=2, insertbackground=caret)
        s.configure('TCombobox', insertwidth=2, insertbackground=caret)
    except Exception:
        pass

def _v16_atualizar_campo(self, w):
    """Pinta o fundo de âmbar se o campo estiver vazio; senão, cor normal."""
    try:
        nome = self._v16_nome_tema()
        vazio_bg = _V16_VAZIO.get(nome, _V16_VAZIO['claro'])
        cheio_bg = (self.tema_cores or {}).get('campo_bg', '#ffffff')
        txt = w.get().strip()
        w.configure(fieldbackground=(vazio_bg if not txt else cheio_bg))
    except Exception:
        pass

def _v16_atualizar_todos(self):
    try:
        for w in (self.entries or {}).values():
            self._v16_atualizar_campo(w)
    except Exception:
        pass

def _v16_instalar_binds(self):
    """Reage a digitação / seleção / saída de foco p/ recalcular o glow."""
    try:
        for w in (self.entries or {}).values():
            w.bind('<KeyRelease>',        lambda e, w=w: self._v16_atualizar_campo(w), add='+')
            w.bind('<FocusOut>',          lambda e, w=w: self._v16_atualizar_campo(w), add='+')
            if isinstance(w, ttk.Combobox):
                w.bind('<<ComboboxSelected>>', lambda e, w=w: self._v16_atualizar_campo(w), add='+')
    except Exception:
        pass

def _v16_remover_mais(self):
    """Some o botão ➕ dos campos Estante e Prateleira (mantém o combobox)."""
    for nome in ('estante', 'prateleira'):
        try:
            ef = self.entries[nome].master          # frame que envolve o campo
            for w in list(ef.winfo_children()):
                if isinstance(w, tk.Button) and '➕' in str(w.cget('text')):
                    w.destroy()
        except Exception:
            pass

# --- encadeia nos métodos existentes (a última definição vence) ------
_V16_CW   = LivroCatalogApp.create_widgets
_V16_AT   = LivroCatalogApp.aplicar_tema
_V16_LF   = LivroCatalogApp.limpar_form
_V16_CFI  = LivroCatalogApp._carregar_form_por_id

def _v16_cw_wrap(self):
    _V16_CW(self)                                   # monta a tela normalmente
    self._v16_remover_mais()                        # tira os ➕
    self._v16_instalar_binds()                      # liga o glow
    self._v16_style_caret()                         # caret destacado
    self._v16_atualizar_todos()                     # glow inicial nos vazios

def _v16_at_wrap(self):
    _V16_AT(self)                                   # aplica o tema normalmente
    self._v16_style_caret()                         # caret segue o tema
    self._v16_atualizar_todos()                     # glow segue o tema

def _v16_lf_wrap(self):
    _V16_LF(self)                                   # limpa normalmente
    self._v16_atualizar_todos()                     # recalcula o glow

def _v16_cfi_wrap(self, livro_id):
    _V16_CFI(self, livro_id)                        # carrega o registro
    self._v16_atualizar_todos()                     # glow nos que vierem vazios

LivroCatalogApp.create_widgets        = _v16_cw_wrap
LivroCatalogApp.aplicar_tema          = _v16_at_wrap
LivroCatalogApp.limpar_form           = _v16_lf_wrap
LivroCatalogApp._carregar_form_por_id = _v16_cfi_wrap
LivroCatalogApp._v16_style_caret      = _v16_style_caret
LivroCatalogApp._v16_atualizar_campo  = _v16_atualizar_campo
LivroCatalogApp._v16_atualizar_todos  = _v16_atualizar_todos
LivroCatalogApp._v16_instalar_binds   = _v16_instalar_binds
LivroCatalogApp._v16_remover_mais     = _v16_remover_mais
LivroCatalogApp._v16_nome_tema        = _v16_nome_tema
# =====================================================================
# FIM PATCH v16  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================
# =====================================================================
# PATCH v17 — (1) QUANTIDADE NÃO INICIA COM 1  (2) GLOW/CARET GARANTIDOS
#               (3) BOTÃO "🗑️ ZERAR CATÁLOGO" (com backup + confirmação)
#  - Quantidade inicia VAZIA (e fica âmbar pelo glow, chamando atenção).
#  - Re-garante o glow (fundo âmbar nos vazios) + caret grosso/colorido,
#    mesmo que o patch do glow não esteja no seu arquivo.
#  - Novo botão "🗑️ Zerar Catálogo": só admin; faz BACKUP automático do
#    banco ANTES; pede para digitar EXCLUIR; apaga livros + histórico de
#    sugestões (mantém usuários, tema, idioma e fonte).
#  - Agnóstico à estrutura: usa só atributos que certamente existem.
#  - Idempotente: colar/executar mais de uma vez não duplica nada.
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================
import shutil as _v17_shutil
import traceback as _v17_tb

_V17_CARET = {'claro': '#b45309', 'escuro': '#fbbf24'}   # cor do cursor
_V17_VAZIO = {'claro': '#fde68a', 'escuro': '#78350f'}   # fundo do campo vazio

def _v17_nome_tema(self):
    try:
        return (getattr(self, 'config', None) or {}).get('tema', 'claro')
    except Exception:
        return 'claro'

def _v17_style_caret(self):
    """Caret mais grosso + cor de contraste (vale p/ todos os campos)."""
    try:
        caret = _V17_CARET.get(self._v17_nome_tema(), _V17_CARET['claro'])
        s = ttk.Style()
        s.configure('TEntry', insertwidth=2, insertbackground=caret)
        s.configure('TCombobox', insertwidth=2, insertbackground=caret)
    except Exception:
        pass

def _v17_atualizar_campo(self, w):
    """Pinta o fundo de âmbar se o campo estiver vazio; senão, cor normal."""
    try:
        nome = self._v17_nome_tema()
        vazio_bg = _V17_VAZIO.get(nome, _V17_VAZIO['claro'])
        cheio_bg = (getattr(self, 'tema_cores', None) or {}).get('campo_bg', '#ffffff')
        w.configure(fieldbackground=(vazio_bg if not w.get().strip() else cheio_bg))
    except Exception:
        pass

def _v17_atualizar_todos(self):
    try:
        for w in (getattr(self, 'entries', None) or {}).values():
            self._v17_atualizar_campo(w)
    except Exception:
        pass

def _v17_instalar_binds(self):
    """Reage a digitação/seleção p/ recalcular o glow (1x só)."""
    if getattr(self, '_v17_glow_done', False):
        return
    try:
        for w in (getattr(self, 'entries', None) or {}).values():
            w.bind('<KeyRelease>', lambda e, w=w: self._v17_atualizar_campo(w), add='+')
            w.bind('<FocusOut>', lambda e, w=w: self._v17_atualizar_campo(w), add='+')
            if isinstance(w, ttk.Combobox):
                w.bind('<<ComboboxSelected>>', lambda e, w=w: self._v17_atualizar_campo(w), add='+')
        self._v17_glow_done = True
    except Exception:
        pass

# ---- limpar_form: NÃO pré-preenche Quantidade nem Saída ----
def _v17_limpar_form(self):
    try:
        for n in (getattr(self, 'entries', None) or {}):
            self.entries[n].delete(0, tk.END)
        if 'entrada' in self.entries:
            self.entries['entrada'].insert(0, datetime.now().strftime('%d/%m/%Y'))
        # Quantidade e Saída ficam VAZIAS (o glow as marca em âmbar)
    except Exception:
        pass
    self.livro_selecionado_id = None
    self._v17_atualizar_todos()

# ---- botão Zerar Catálogo + diálogo com backup + confirmação ----
def _v17_zerar_catalogo(self):
    perms = getattr(self, 'permissoes', None) or {}
    if not perms.get('admin'):
        messagebox.showwarning("🔒 Permissão insuficiente",
                               "Apenas o administrador pode zerar o catálogo.")
        return
    n_livros = len(getattr(self, 'livros', []) or [])

    d = tk.Toplevel(self.root)
    d.title("🗑️ Zerar Catálogo")
    d.transient(self.root); d.grab_set(); d.resizable(False, False)
    try:
        self.centralizar_janela(d)
    except Exception:
        pass

    f = ttk.Frame(d, padding=15); f.pack(fill='both', expand=True)
    ttk.Label(f, text="⚠️ ATENÇÃO: esta ação é irreversível.",
              font=('DejaVu Sans', 11, 'bold'), foreground='#dc2626').pack(anchor='w')
    ttk.Label(f, text=(f"Serão apagados TODOS os {n_livros} livro(s) do catálogo e o "
                       f"histórico de sugestões (autores, editoras, etc.).\n"
                       f"Usuários, tema, idioma e fonte NÃO são afetados.\n"
                       f"Um BACKUP do banco será salvo ANTES de apagar."),
              wraplength=460, justify='left').pack(anchor='w', pady=(8, 12))
    ttk.Label(f, text="Para confirmar, digite exatamente: EXCLUIR").pack(anchor='w')
    ent = ttk.Entry(f, width=20); ent.pack(anchor='w', pady=(4, 12))

    bf = ttk.Frame(f); bf.pack(fill='x')
    btn_ok = ttk.Button(bf, text="🗑️ Apagar tudo", state='disabled')
    btn_ok.pack(side='left', padx=5)
    ttk.Button(bf, text="Cancelar", command=d.destroy).pack(side='left', padx=5)

    def _on_txt(e=None):
        try:
            btn_ok.config(state='normal' if ent.get().strip() == 'EXCLUIR' else 'disabled')
        except Exception:
            pass
    ent.bind('<KeyRelease>', _on_txt)

    def _do():
        # 1) backup OBRIGATÓRIO antes de apagar
        try:
            base = getattr(self, 'diretorio_app', os.getcwd())
            bdir = os.path.join(base, 'backups')
            os.makedirs(bdir, exist_ok=True)
            ts = datetime.now().strftime('%Y%m%d_%H%M%S')
            bkp = os.path.join(bdir, f"catalogo_feeu_ANTES_ZERAR_{ts}.db")
            _v17_shutil.copy2(self.db_file, bkp)
        except Exception:
            messagebox.showerror("❌", "Não foi possível criar o backup de segurança.\n"
                                 "O catálogo NÃO foi apagado.\n\n" + _v17_tb.format_exc(),
                                 parent=d)
            return
        # 2) apaga livros + histórico de sugestões
        try:
            conn = sqlite3.connect(self.db_file); cur = conn.cursor()
            cur.execute("DELETE FROM livros")
            try:
                cur.execute("DELETE FROM historico_campos")
            except Exception:
                pass
            conn.commit(); conn.close()
        except Exception:
            messagebox.showerror("❌", "Erro ao apagar os registros.\n\n" + _v17_tb.format_exc(),
                                 parent=d)
            return
        try: self.carregar_dados()
        except Exception: pass
        try: self.atualizar_comboboxes()
        except Exception: pass
        try: self.limpar_form()
        except Exception: pass
        d.destroy()
        messagebox.showinfo("✅", f"Catálogo zerado.\n\nBackup de segurança salvo em:\n{bkp}",
                            parent=self.root)

    btn_ok.config(command=_do)
    ent.focus_set()
    d.bind('<Return>', lambda e: (_do() if ent.get().strip() == 'EXCLUIR' else None))

def _v17_criar_botao_zerar(self):
    """Cria o botão na barra de ações (1x só), com cor de perigo."""
    if getattr(self, '_v17_zerar_btn', None):
        return
    try:
        bf = self.btn_excluir.master   # frame da barra de botões
        btn = tk.Button(bf, text="🗑️ Zerar Catálogo", command=self._v17_zerar_catalogo,
                        bg='#7f1d1d', fg='white', activebackground='#991b1b',
                        activeforeground='white', font=('DejaVu Sans', 9, 'bold'),
                        relief='raised', padx=8, pady=4, cursor='hand2')
        btn.pack(side='left', padx=5)
        self._v17_zerar_btn = btn
    except Exception:
        pass

# ---- wrapper do __init__: esvazia Quantidade + instala glow + botão ----
_V17_PREV_INIT = LivroCatalogApp.__init__

def _v17_new_init(self, root):
    _V17_PREV_INIT(self, root)
    try:
        if not self.root.winfo_exists():   # login cancelado: não faz mais nada
            return
    except Exception:
        return
    try:                                   # (1) Quantidade NÃO inicia com 1
        if 'quantidade' in self.entries:
            self.entries['quantidade'].delete(0, tk.END)
    except Exception:
        pass
    try:                                   # (2) glow + caret
        self._v17_style_caret()
        self._v17_instalar_binds()
        self._v17_atualizar_todos()
    except Exception:
        pass
    try:                                   # (3) botão zerar
        self._v17_criar_botao_zerar()
    except Exception:
        pass

# ---- reaplica glow/caret ao trocar de tema (encadeia aplicar_tema) ----
_V17_PREV_AT = getattr(LivroCatalogApp, 'aplicar_tema', None)

def _v17_new_at(self):
    if _V17_PREV_AT:
        _V17_PREV_AT(self)
    try:
        self._v17_style_caret()
        self._v17_atualizar_todos()
    except Exception:
        pass

# ---- reatribui (a última definição vence) ----
LivroCatalogApp.__init__        = _v17_new_init
LivroCatalogApp.limpar_form     = _v17_limpar_form
LivroCatalogApp.aplicar_tema    = _v17_new_at
LivroCatalogApp._v17_style_caret      = _v17_style_caret
LivroCatalogApp._v17_atualizar_campo  = _v17_atualizar_campo
LivroCatalogApp._v17_atualizar_todos  = _v17_atualizar_todos
LivroCatalogApp._v17_instalar_binds   = _v17_instalar_binds
LivroCatalogApp._v17_zerar_catalogo   = _v17_zerar_catalogo
LivroCatalogApp._v17_criar_botao_zerar = _v17_criar_botao_zerar
LivroCatalogApp._v17_nome_tema        = _v17_nome_tema
# =====================================================================
# FIM PATCH v17  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v18 — (1) CARET VISÍVEL NOS DOIS TEMAS
#               (2) ENTER = TAB NOS CAMPOS  (3) LOOP DE FOCO (volta ao Título)
#  - Caret: insertcolor por tema (claro=laranja-queimado, escuro=amarelo-vivo)
#    + insertwidth=2; reaplicado ao TROCAR de tema e no boot.
#  - Enter e Tab pulam de campo em campo no formulário (mesma função).
#  - Sequência de foco definida à mão, com Shift+Tab no reverso e LOOP:
#    ... Entrada -> 📅Entrada -> Saída -> 📅Saída -> (volta) Título.
#  - Botões de ação (Adicionar/Excluir/...) ficam FORA do ciclo de Enter/Tab
#    (segurança: evita disparo acidental).
#  - Remove o ➕ de Estante/Prateleira se ainda existir (idempotente).
#  - Foco no Título ao limpar/abrir.
#  - Idempotente e todo em try/except: se falhar, não quebra o resto.
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================

# --- (1) caret com alto contraste por tema ---------------------------
_V18_CARET = {'claro': '#9a3412', 'escuro': '#fde047'}   # laranja-queimado / amarelo-vivo

def _v18_caret(self):
    """Define a cor/espessura do cursor conforme o tema ATUAL (fonte da verdade = config)."""
    try:
        nome = (getattr(self, 'config', None) or {}).get('tema', 'claro')
        cor = _V18_CARET.get(nome, _V18_CARET['claro'])
        s = ttk.Style()
        # insertcolor = propriedade CORRETA do ttk; insertwidth = caret mais grosso
        s.configure('TEntry',    insertcolor=cor, insertwidth=2)
        s.configure('TCombobox', insertcolor=cor, insertwidth=2)
    except Exception:
        pass

# --- (2)/(3) sequência de foco + navegação ---------------------------
_V18_ORDEM = ['titulo', 'autor', 'estante', 'prateleira', 'editora',
              'assunto', 'bibliotecario', 'quantidade', 'emprestado_a',
              'entrada', 'saida']

def _v18_remover_mais(self):
    """Some o ➕ de Estante/Prateleira, se ainda existir (não afeta os 📅)."""
    for k in ('estante', 'prateleira'):
        try:
            ef = self.entries[k].master
            for w in list(ef.winfo_children()):
                if isinstance(w, tk.Button) and '➕' in str(w.cget('text')):
                    w.destroy()
        except Exception:
            pass

def _v18_montar_nav(self):
    """Monta a sequência (campos + 📅 nas posições certas) e binda Tab/Enter/Shift+Tab."""
    try:
        self._v18_remover_mais()
        seq = []
        for k in _V18_ORDEM:
            w = self.entries.get(k)
            if w is None:
                continue
            seq.append(w)
            if k in ('entrada', 'saida'):           # inclui o 📅 logo após o seu campo
                try:
                    for sib in w.master.winfo_children():
                        if isinstance(sib, tk.Button) and '📅' in str(sib.cget('text')):
                            seq.append(sib)
                            break
                except Exception:
                    pass
        self._v18_seq = seq

        for w in seq:
            w.bind('<Tab>',          lambda e: self._v18_go(+1, e))   # vence o ciclo nativo
            w.bind('<Return>',       lambda e: self._v18_go(+1, e))   # Enter = Tab
            w.bind('<Shift-Tab>',    lambda e: self._v18_go(-1, e))   # reverso
            w.bind('<ISO_Left_Tab>', lambda e: self._v18_go(-1, e))   # reverso (Linux)
    except Exception:
        pass

def _v18_go(self, delta, event):
    """Avança (+1) ou recua (-1) na sequência, com loop fechado."""
    seq = getattr(self, '_v18_seq', None) or []
    if not seq:
        return None
    try:
        i = seq.index(event.widget)
    except ValueError:
        return None
    j = (i + delta) % len(seq)                    # % => loop (último->Título / Título->último)
    try:
        seq[j].focus_set()
    except Exception:
        pass
    return "break"                                # não deixa o Tab/Enter vazar pro nativo

def _v18_focar_titulo(self):
    try:
        self.root.after(50, lambda: self.entries['titulo'].focus_set())
    except Exception:
        pass

# --- wrappers (encadeiam sem reescrever o seu código) ----------------
_V18_PREV_CW = LivroCatalogApp.create_widgets
_V18_PREV_AT = LivroCatalogApp.aplicar_tema
_V18_PREV_LF = LivroCatalogApp.limpar_form

def _v18_cw_wrap(self):
    _V18_PREV_CW(self)                            # monta a tela normalmente
    try:
        self._v18_montar_nav()                    # sequência + binds
    except Exception:
        pass
    try:
        self._v18_caret()                         # caret no tema do boot
    except Exception:
        pass
    try:
        self.root.after(250, self._v18_focar_titulo)
    except Exception:
        pass

def _v18_at_wrap(self):
    _V18_PREV_AT(self)                            # aplica o tema normalmente
    try:
        self._v18_caret()                         # ✅ recalcula o caret p/ o tema novo
    except Exception:
        pass

def _v18_lf_wrap(self):
    _V18_PREV_LF(self)                            # limpa normalmente
    try:
        self._v18_focar_titulo()                  # caret volta p/ o Título
    except Exception:
        pass

# --- reatribui (a última definição vence) ----------------------------
LivroCatalogApp.create_widgets = _v18_cw_wrap
LivroCatalogApp.aplicar_tema   = _v18_at_wrap
LivroCatalogApp.limpar_form    = _v18_lf_wrap
LivroCatalogApp._v18_caret        = _v18_caret
LivroCatalogApp._v18_montar_nav   = _v18_montar_nav
LivroCatalogApp._v18_remover_mais = _v18_remover_mais
LivroCatalogApp._v18_go           = _v18_go
LivroCatalogApp._v18_focar_titulo = _v18_focar_titulo
# =====================================================================
# FIM PATCH v18  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v19 — IMPORTAÇÃO COM REGRAS: PERGUNTA EM CADA DUPLICATA
#  - Registros NOVOS (sem igual no programa) -> entram DIRETO, sem pergunta.
#  - Registros JÁ EXISTENTES (mesma identidade) -> abrem uma janela com:
#      ➕ Somar este            ⏭️ Ignorar este
#      ➕ Somar este + próximos  ⏭️ Ignorar este + próximos
#      ❌ Cancelar tudo  (desfaz o lote inteiro, inclusive os novos)
#  - Mostra "No programa: X | Na planilha: Y | Se somar: X+Y".
#  - Gravação atômica: commit só no fim; cancelar/erro -> rollback.
#  - Reusa os leitores e o _checar_duplicata que já existem (4.14.0).
#  - Idempotente e em try/except: se falhar, não quebra o resto.
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================

# ---------------------------------------------------------------------
# Janela de decisão para UMA duplicata (retorna a escolha do usuário)
# ---------------------------------------------------------------------
def _v19_decidir(self, info, dup_idx):
    """Abre diálogo modal e devolve: somar | ignorar | somar_todos |
    ignorar_todos | cancelar."""
    res = {'v': 'cancelar'}   # fechar no X = cancelar (sem ambiguidade)

    d = tk.Toplevel(self.root)
    d.title("⚠️ Registro já existente")
    d.transient(self.root)
    d.grab_set()
    d.resizable(False, False)
    d.protocol("WM_DELETE_WINDOW", lambda: d.destroy())  # res['v'] já é 'cancelar'

    f = ttk.Frame(d, padding=14)
    f.pack(fill='both', expand=True)

    ttk.Label(f, text=f"O livro abaixo JÁ EXISTE no catálogo (duplicata #{dup_idx}).",
              font=('DejaVu Sans', 10, 'bold'), foreground='#b45309').pack(anchor='w')
    ttk.Label(f, text="O que deseja fazer com ESTE item?",
              font=('DejaVu Sans', 9)).pack(anchor='w', pady=(2, 8))

    # --- dados do item ---
    det = ttk.Frame(f)
    det.pack(anchor='w', fill='x')
    ttk.Label(det, text="Título:", font=('DejaVu Sans', 9, 'bold')).grid(row=0, column=0, sticky='w')
    ttk.Label(det, text=str(info.get('titulo', '')), wraplength=380).grid(row=0, column=1, sticky='w', padx=4)
    ttk.Label(det, text="Autor:", font=('DejaVu Sans', 9, 'bold')).grid(row=1, column=0, sticky='w')
    ttk.Label(det, text=str(info.get('autor', '')), wraplength=380).grid(row=1, column=1, sticky='w', padx=4)
    est = str(info.get('estante', '') or '').strip()
    pra = str(info.get('prateleira', '') or '').strip()
    if est or pra:
        ttk.Label(det, text="Local:", font=('DejaVu Sans', 9, 'bold')).grid(row=2, column=0, sticky='w')
        ttk.Label(det, text=f"Estante '{est or '-'}'  /  Prateleira '{pra or '-'}'").grid(row=2, column=1, sticky='w', padx=4)

    # --- linha das quantidades (o coração da decisão) ---
    qa = int(info.get('qa', 0) or 0)
    qp = int(info.get('qp', 0) or 0)
    qf = ttk.Frame(f)
    qf.pack(anchor='w', fill='x', pady=(10, 4))
    ttk.Label(qf, text=f"No programa: {qa}", font=('DejaVu Sans', 10, 'bold')).pack(side='left', padx=(0, 10))
    ttk.Label(qf, text=f"Na planilha: {qp}", font=('DejaVu Sans', 10, 'bold'),
              foreground='#1e40af').pack(side='left', padx=(0, 10))
    ttk.Label(qf, text=f"Se somar: {qa + qp}", font=('DejaVu Sans', 10, 'bold'),
              foreground='#15803d').pack(side='left')
    if qp == 0:
        ttk.Label(f, text="(a planilha trouxe quantidade 0 — 'somar' não altera o estoque)",
                  font=('DejaVu Sans', 8), foreground='#64748b').pack(anchor='w')

    ttk.Separator(f, orient='horizontal').pack(fill='x', pady=10)

    # --- botões ---
    b1 = ttk.Frame(f); b1.pack(fill='x', pady=(0, 4))
    ttk.Button(b1, text="➕ Somar este",
               command=lambda: (res.__setitem__('v', 'somar'), d.destroy())).pack(side='left', padx=3, expand=True, fill='x')
    ttk.Button(b1, text="⏭️ Ignorar este",
               command=lambda: (res.__setitem__('v', 'ignorar'), d.destroy())).pack(side='left', padx=3, expand=True, fill='x')

    b2 = ttk.Frame(f); b2.pack(fill='x', pady=(0, 4))
    ttk.Button(b2, text="➕ Somar este + todos os próximos",
               command=lambda: (res.__setitem__('v', 'somar_todos'), d.destroy())).pack(side='left', padx=3, expand=True, fill='x')
    ttk.Button(b2, text="⏭️ Ignorar este + todos os próximos",
               command=lambda: (res.__setitem__('v', 'ignorar_todos'), d.destroy())).pack(side='left', padx=3, expand=True, fill='x')

    b3 = ttk.Frame(f); b3.pack(fill='x', pady=(4, 0))
    ttk.Button(b3, text="❌ Cancelar tudo (não grava nada do lote)",
               command=lambda: (res.__setitem__('v', 'cancelar'), d.destroy())).pack(side='left', padx=3, expand=True, fill='x')

    try:
        self.centralizar_janela(d)
    except Exception:
        pass
    d.lift(); d.focus_force()
    self.root.wait_window(d)
    return res['v']

# ---------------------------------------------------------------------
# Nova importação (substitui a ativa)
# ---------------------------------------------------------------------
def _v19_importar(self):
    if not self._checar_permissao('importar', 'Você não tem permissão para importar dados.'):
        return

    caminho = filedialog.askopenfilename(
        title="Selecionar",
        filetypes=[("Planilhas/CSV", "*.xlsx *.ods *.csv"),
                   ("Excel", "*.xlsx"), ("Calc", "*.ods"), ("CSV", "*.csv")])
    if not caminho:
        return
    ext = os.path.splitext(caminho)[1].lower()

    # fallback p/ CSV se a lib do formato faltar (comportamento que você já tem)
    if ext == '.xlsx' and not HAS_XLSX:
        if messagebox.askyesno("📦 Biblioteca ausente",
                               "openpyxl não está instalado.\n\nSelecionar um CSV no lugar?"):
            caminho = filedialog.askopenfilename(title="Selecionar CSV", filetypes=[("CSV", "*.csv")])
            if not caminho:
                return
            ext = '.csv'
        else:
            return
    if ext == '.ods' and not HAS_ODS:
        if messagebox.askyesno("📦 Biblioteca ausente",
                               "odfpy não está instalado.\n\nSelecionar um CSV no lugar?"):
            caminho = filedialog.askopenfilename(title="Selecionar CSV", filetypes=[("CSV", "*.csv")])
            if not caminho:
                return
            ext = '.csv'
        else:
            return

    # trava de arquivo aberto em outro programa
    try:
        with open(caminho, 'rb') as f:
            pass
    except PermissionError:
        messagebox.showerror("⛔ Erro de Acesso",
                             "A planilha está ABERTA em outro programa.\nFeche o Excel/LibreOffice e tente novamente.")
        return
    except OSError as e:
        messagebox.showerror("⛔ Erro de Acesso", f"Não foi possível abrir o arquivo:\n{e}")
        return

    # --- leitura (reusa os leitores que já existem) ---
    try:
        if ext == '.xlsx':
            dados = self._ler_xlsx(caminho)
        elif ext == '.ods':
            dados = self._ler_ods(caminho)
        elif ext == '.csv':
            dados = self._ler_csv(caminho)
        else:
            messagebox.showwarning("⚠️", "Formato não suportado. Use .xlsx, .ods ou .csv.")
            return
    except Exception:
        messagebox.showerror("❌", "Erro ao ler a planilha:\n" + traceback.format_exc())
        return

    if not dados:
        messagebox.showwarning("⚠️",
            "Nenhum dado reconhecido na planilha.\n\n"
            "Verifique se há cabeçalhos como: Título, Autor, Estante, Prateleira, "
            "Editora, Assunto, Quantidade, Entrada, Saída.")
        return

    # --- processamento com decisões (tudo numa transação) ---
    conn = sqlite3.connect(self.db_file)
    cur = conn.cursor()
    cols = self._colunas_livros_existentes(cur)

    ins = att = ign = 0
    dup_idx = 0
    decisao_global = None     # None | 'somar' | 'ignorar'
    cancelou = False

    def _aplicar_somar(dup, qp):
        nonlocal att
        lid, qat = dup
        nq = (qat if qat else 0) + qp
        sets = ["quantidade=?", "disponibilidade=?"]
        params = [nq, 'Sim' if nq > 0 else 'Não']
        # preenche estante/prateleira só se o registro atual estiver vazio nelas
        if 'estante' in cols and str(dup_info_est.get('est', '')).strip():
            sets.append("estante=CASE WHEN COALESCE(estante,'')='' THEN ? ELSE estante END")
            params.append(dup_info_est['est'])
        if 'prateleira' in cols and str(dup_info_est.get('prat', '')).strip():
            sets.append("prateleira=CASE WHEN COALESCE(prateleira,'')='' THEN ? ELSE prateleira END")
            params.append(dup_info_est['prat'])
        params.append(lid)
        cur.execute(f"UPDATE livros SET {', '.join(sets)} WHERE id=?", tuple(params))
        att += 1

    try:
        for l in dados:
            t = str(l.get('titulo', '') or '').strip()
            a = str(l.get('autor', '') or '').strip()
            if not t or not a:           # linha inválida -> ignora sem perguntar
                ign += 1
                continue

            e   = str(l.get('editora', '') or '').strip()
            ass = str(l.get('assunto', '') or '').strip()
            est = str(l.get('estante', '') or '').strip()
            prat = str(l.get('prateleira', '') or '').strip()
            bib = str(l.get('bibliotecario', '') or '').strip()
            emp = str(l.get('emprestado_a', '') or '').strip()
            q, ok = self._validar_quantidade(l.get('quantidade', '1'))
            if not ok:
                q = 1
            entrada_iso = self._data_br_para_iso(l.get('entrada', ''), manter_original=True)
            saida_iso   = self._data_br_para_iso(l.get('saida', ''), manter_original=True)
            disp_raw = str(l.get('disponibilidade', '') or '').strip().lower()
            if disp_raw in ('sim', 's', 'disponivel', 'yes', 'true', '1'):
                disp = 'Sim'
            elif disp_raw in ('nao', 'não', 'n', 'indisponivel', 'no', 'false', '0'):
                disp = 'Não'
            else:
                disp = 'Sim' if q > 0 else 'Não'

            dup = self._checar_duplicata(t, a, e, ass, est, prat, cursor=cur)

            if dup is None:
                # ---- NOVO: entra DIRETO, sem pergunta ----
                fields = ['titulo', 'autor']
                values = [t, a]
                for col, val in [('editora', e), ('assunto', ass), ('bibliotecario', bib),
                                 ('quantidade', q), ('disponibilidade', disp),
                                 ('emprestado_a', emp), ('entrada', entrada_iso),
                                 ('saida', saida_iso), ('estante', est), ('prateleira', prat)]:
                    if col in cols:
                        fields.append(col); values.append(val)
                ph = ','.join(['?'] * len(fields))
                cur.execute(f"INSERT INTO livros ({','.join(fields)}) VALUES ({ph})", tuple(values))
                ins += 1
                self._registrar_historico_livro(
                    {'autor': a, 'editora': e, 'assunto': ass, 'bibliotecario': bib,
                     'emprestado_a': emp, 'estante': est, 'prateleira': prat}, cursor=cur)
            else:
                # ---- JÁ EXISTE: pergunta (ou aplica decisão global) ----
                dup_idx += 1
                qa = dup[1] if dup[1] else 0
                info = {'titulo': t, 'autor': a, 'estante': est, 'prateleira': prat,
                        'qa': qa, 'qp': q}
                dup_info_est = {'est': est, 'prat': prat}   # p/ o UPDATE condicional

                if decisao_global == 'somar':
                    escolha = 'somar'
                elif decisao_global == 'ignorar':
                    escolha = 'ignorar'
                else:
                    escolha = self._v19_decidir(info, dup_idx)

                if escolha == 'cancelar':
                    cancelou = True
                    break
                elif escolha == 'somar_todos':
                    decisao_global = 'somar';   _aplicar_somar(dup, q)
                elif escolha == 'ignorar_todos':
                    decisao_global = 'ignorar'; ign += 1
                elif escolha == 'somar':
                    _aplicar_somar(dup, q)
                else:  # 'ignorar'
                    ign += 1

        # ---- commit ou rollback (atômico) ----
        if cancelou:
            conn.rollback()
            conn.close()
            messagebox.showinfo("🚫 Importação cancelada",
                f"Nada foi gravado (lote desfeito).\n\n"
                f"Até o cancelamento havia: {ins} novo(s), {att} somado(s), {ign} ignorado(s).")
            return

        conn.commit()
        conn.close()
        self._recarregar_manter_busca()
        self.atualizar_comboboxes()
        msg = f"✅ Importação concluída\n\nNovos inseridos: {ins}\nSomados (duplicatas): {att}\nIgnorados: {ign}"
        if dup_idx:
            msg += f"\n\n(Duplicatas encontradas: {dup_idx})"
        messagebox.showinfo("Importação", msg)

    except Exception:
        try:
            conn.rollback()
        except Exception:
            pass
        try:
            conn.close()
        except Exception:
            pass
        messagebox.showerror("❌", "Erro na importação (nada foi gravado):\n" + traceback.format_exc())

# reatribui (a última definição vence -> esta vira a importação ativa)
LivroCatalogApp._v19_decidir = _v19_decidir
LivroCatalogApp.importar_planilha = _v19_importar
# =====================================================================
# FIM PATCH v19  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v20 — "ABAS" POR MODO: libera espaço p/ o Catálogo em telas pequenas
#  - Seletor no topo com 2 botões-aba: 📖 Catálogo | 🔍 Ferramentas.
#  - Modo Catálogo (padrão): ESCONDE "Busca e Filtros" + "Exportar/...".
#    => a grade ganha ~120px (crucial em 1024x768 / 800x600).
#  - Modo Ferramentas: mostra os 2 quadros de volta (buscar/importar/exportar).
#  - Os Resultados ficam COM o Catálogo (melhor pro fluxo clique->formulário).
#  - Ctrl+F: vai p/ Ferramentas e foca o campo de busca.
#  - NÃO reescreve create_widgets (só pack_forget/pack em widgets já filhos
#    de main) => compatível com TODOS os patches anteriores.
#  - Idempotente e todo em try/except: se falhar, não muda nada.
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================

def _v20_find(self):
    """Localiza os 2 quadros e o campo de busca SEM depender de texto (i18n-safe)."""
    busca_lf = exp_lf = busca_ent = None
    try:
        exp_lf = self.btn_import.master                 # btn_import nasceu em 'exp'
    except Exception:
        pass
    try:
        fv = str(self.filtro_var)
        bv = str(self.busca_var)
        main = self.main_paned.master
        stack = [main]
        while stack:
            w = stack.pop()
            try:
                stack.extend(w.winfo_children())
            except Exception:
                continue
            if isinstance(w, ttk.Combobox) and str(w.cget('textvariable')) == fv:
                busca_lf = w.master                     # combobox de filtro nasceu em 'busca'
            if isinstance(w, (ttk.Entry, tk.Entry)) and str(w.cget('textvariable')) == bv:
                busca_ent = w
    except Exception:
        pass
    return busca_lf, exp_lf, busca_ent

def _v20_pintar_abas(self):
    """Pinta os 2 botões-aba conforme o tema (ativo = 'afundado'/destaque)."""
    try:
        t = getattr(self, 'tema_cores', None) or {}
        bg_ativo  = t.get('bg_card', '#ffffff')
        bg_inativo = t.get('bg_botao', '#e2e8f0')
        fg_ativo  = t.get('titulo', '#1e40af')
        fg_inativo = t.get('texto', '#0f172a')
        modo = getattr(self, '_v20_modo', 'catalogo')
        for chave, btn in (('catalogo', getattr(self, '_v20_btn_cat', None)),
                           ('ferramentas', getattr(self, '_v20_btn_fer', None))):
            if not btn:
                continue
            ativo = (modo == chave)
            btn.config(bg=(bg_ativo if ativo else bg_inativo),
                       fg=(fg_ativo if ativo else fg_inativo),
                       relief=('flat' if ativo else 'raised'),
                       bd=(0 if ativo else 1))
        # moldura do seletor acompanha o tema
        fr = getattr(self, '_v20_frame', None)
        if fr:
            fr.config(bg=t.get('bg_janela', '#f1f5f9'))
    except Exception:
        pass

def _v20_reflow(self):
    """Reajusta layout dos patches de layout após trocar de modo."""
    for m in ('_v8_adjust', '_v13_layout', '_v11_reaplicar_layout'):
        try:
            getattr(self, m)()
        except Exception:
            pass

def _v20_set_modo(self, modo):
    """Alterna a visibilidade dos 2 quadros, mantendo a ordem com before=."""
    try:
        self._v20_modo = modo
        b = getattr(self, '_v20_busca', None)
        e = getattr(self, '_v20_exp', None)
        mp = self.main_paned
        if modo == 'ferramentas':
            try: e.pack(fill='x', pady=5, before=mp)     # exp logo acima do catálogo
            except Exception: pass
            try: b.pack(fill='x', pady=5, before=e)      # busca acima de exp
            except Exception: pass
        else:
            try: b.pack_forget()
            except Exception: pass
            try: e.pack_forget()
            except Exception: pass
        self._v20_pintar_abas()
        try: self.root.after(60, self._v20_reflow)
        except Exception: pass
    except Exception:
        pass

def _v20_criar_abas(self):
    """Monta o seletor e reordena o pack (header -> seletor -> [busca,exp] -> catálogo)."""
    if getattr(self, '_v20_done', False):
        return
    b, e, bent = self._v20_find()
    if not b or not e:                                   # estrutura inesperada: não faz nada
        return
    self._v20_busca, self._v20_exp, self._v20_busca_ent = b, e, bent
    main = self.main_paned.master
    # desempacota os 3 pra reempilhar na ordem certa
    for w in (b, e, self.main_paned):
        try: w.pack_forget()
        except Exception: pass
    # seletor (2 botões-aba) logo abaixo do cabeçalho
    fr = tk.Frame(main, bg='#f1f5f9')
    fr.pack(fill='x', pady=(0, 5))
    self._v20_frame = fr
    self._v20_btn_cat = tk.Button(
        fr, text="  📖  Catálogo  ", font=('DejaVu Sans', 10, 'bold'), cursor='hand2',
        command=lambda: self._v20_set_modo('catalogo'))
    self._v20_btn_cat.pack(side='left', padx=(2, 0))
    self._v20_btn_fer = tk.Button(
        fr, text="  🔍  Ferramentas  ", font=('DejaVu Sans', 10, 'bold'), cursor='hand2',
        command=lambda: self._v20_set_modo('ferramentas'))
    self._v20_btn_fer.pack(side='left')
    tk.Label(fr, text="(Ctrl+F abre as Ferramentas)", font=('DejaVu Sans', 8),
             fg='#64748b').pack(side='left', padx=8)
    # modo padrão = Catálogo (quadros escondidos) -> reempilha só o catálogo
    self._v20_modo = 'catalogo'
    try: self.main_paned.pack(fill='both', expand=True, pady=5)
    except Exception: pass
    self._v20_pintar_abas()
    # atalho Ctrl+F -> Ferramentas + foco na busca
    try:
        def _cf(event=None):
            self._v20_set_modo('ferramentas')
            try:
                self.root.after(40, lambda: self._v20_busca_ent.focus_set())
            except Exception: pass
            return "break"
        self.root.bind_all('<Control-f>', _cf)
        self.root.bind_all('<Control-F>', _cf)
    except Exception: pass
    self._v20_done = True

# ---- wrappers (encadeiam sem reescrever o seu create_widgets/aplicar_tema) ----
_V20_PREV_CW = LivroCatalogApp.create_widgets
_V20_PREV_AT = LivroCatalogApp.aplicar_tema

def _v20_cw_wrap(self):
    _V20_PREV_CW(self)                                   # monta a tela (com todos os patches)
    try:
        self._v20_criar_abas()                           # depois, reorganiza em "abas"
    except Exception:
        pass

def _v20_at_wrap(self):
    _V20_PREV_AT(self)                                   # aplica o tema normalmente
    try:
        self._v20_pintar_abas()                          # repinta os botões-aba no tema novo
    except Exception:
        pass

LivroCatalogApp.create_widgets = _v20_cw_wrap
LivroCatalogApp.aplicar_tema   = _v20_at_wrap
LivroCatalogApp._v20_find        = _v20_find
LivroCatalogApp._v20_pintar_abas = _v20_pintar_abas
LivroCatalogApp._v20_reflow      = _v20_reflow
LivroCatalogApp._v20_set_modo    = _v20_set_modo
LivroCatalogApp._v20_criar_abas  = _v20_criar_abas
# =====================================================================
# FIM PATCH v20  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================

# =====================================================================
# PATCH v21 — CORRETOR ORTOGRÁFICO (sugestões + dicionário pessoal)
#  - Ao SAIR de um campo de texto com palavra não reconhecida -> fundo ROSA.
#  - Duplo-clique no campo (rosa ou não) OU botão "✓ Ortografia" abre o
#    corretor: palavra suspeita + sugestões + correção manual + botões
#    Substituir / Substituir todas / Ignorar / Adicionar ao dicionário.
#  - Dicionário pessoal persiste em dicionario_pessoal.txt (ao lado do db).
#  - Siglas (ALL-CAPS), números e palavras de 1 letra são ignorados.
#  - Idioma do corretor (pt/en/es) salvo em config['corretor_idioma'].
#  - SEM pyspellchecker: o botão avisa e o resto segue normal (não quebra).
#  - NÃO reescreve campos (só bind FocusOut/Double + 1 botão) -> compatível
#    com glow de vazio (v16/v17), caret (v18), Enter=Tab, autocomplete.
#  - Idempotente e todo em try/except.
# Cola ANTES de:  if __name__ == "__main__":   (NÃO apague nada acima)
# =====================================================================
try:
    from spellchecker import SpellChecker as _v21_SC
    HAS_SPELL = True
except Exception:
    HAS_SPELL = False

_V21_CAMPOS_TEXTO = ('titulo', 'autor', 'editora', 'assunto', 'bibliotecario', 'emprestado_a')
_V21_ROSA = {'claro': '#fecdd3', 'escuro': '#881337'}     # fundo de "atenção ortográfica"
_V21_LANGS = [('pt', 'Português'), ('en', 'English'), ('es', 'Español')]

# ---- dicionário pessoal (persistente) ------------------------------
def _v21_pessoal_path(self):
    return os.path.join(getattr(self, 'diretorio_app', os.getcwd()), 'dicionario_pessoal.txt')

def _v21_carregar_pessoal(self):
    if getattr(self, '_v21_pessoal', None) is not None:
        return self._v21_pessoal
    s = set()
    try:
        p = self._v21_pessoal_path()
        if os.path.exists(p):
            with open(p, 'r', encoding='utf-8') as f:
                for ln in f:
                    w = ln.strip().lower()
                    if w:
                        s.add(w)
    except Exception:
        pass
    self._v21_pessoal = s
    return s

def _v21_salvar_pessoal(self):
    try:
        with open(self._v21_pessoal_path(), 'w', encoding='utf-8') as f:
            for w in sorted(getattr(self, '_v21_pessoal', set())):
                f.write(w + '\n')
    except Exception:
        pass

# ---- spell checker com cache por idioma ----------------------------
def _v21_spell(self):
    lang = (self.config or {}).get('corretor_idioma', 'pt')
    if lang not in dict(_V21_LANGS):
        lang = 'pt'
    cache = getattr(self, '_v21_cache', None)
    if cache is None:
        cache = {}; self._v21_cache = cache
    sc = cache.get(lang)
    if sc is None:
        try:
            sc = _v21_SC(language=lang)
        except Exception:
            sc = _v21_SC()           # fallback pro dicionário default
        cache[lang] = sc
    return sc

def _v21_eh_suspeita(self, palavra):
    """True se a palavra NÃO é reconhecida (e não é sigla/número/curta/pessoal/ignorada)."""
    if not palavra or len(palavra) < 2:
        return False
    if palavra.isupper():                 # sigla/acrônimo (FEEU, RJ, V...)
        return False
    low = palavra.lower()
    if low in self._v21_carregar_pessoal():
        return False
    if low in getattr(self, '_v21_ignorar_sessao', set()):
        return False
    try:
        return palavra.lower() in self._v21_spell().unknown([palavra])
    except Exception:
        return False

def _v21_tokens(self, texto):
    """Palavras (com acento) na ordem de aparição, preservando repetições."""
    try:
        return re.findall(r"[^\W\d_]+", texto or '', re.UNICODE)
    except Exception:
        return []

def _v21_marcar_campo(self, w):
    """No FocusOut: se não-vazio e com suspeita -> rosa; senão deixa o glow decidir."""
    if not HAS_SPELL:
        return
    try:
        txt = w.get()
        if not txt.strip():
            return                       # campo vazio: não mexe (glow âmbar manda)
        tem_erro = any(self._v21_eh_suspeita(t) for t in self._v21_tokens(txt))
        nome = self._v21_nome_tema() if hasattr(self, '_v21_nome_tema') else \
            (self.config or {}).get('tema', 'claro')
        if tem_erro:
            w.configure(fieldbackground=_V21_ROSA.get(nome, _V21_ROSA['claro']))
        else:
            w.configure(fieldbackground=(self.tema_cores or {}).get('campo_bg', '#ffffff'))
    except Exception:
        pass

# ---- coleta de "jobs" (campo + palavra suspeita) -------------------
def _v21_coletar(self):
    jobs = []
    for nome in _V21_CAMPOS_TEXTO:
        w = (self.entries or {}).get(nome)
        if w is None:
            continue
        for tok in self._v21_tokens(w.get()):
            if self._v21_eh_suspeita(tok):
                jobs.append([w, nome, tok])
    return jobs

# ---- diálogo corretor ----------------------------------------------
def _v21_abrir_corretor(self, start_widget=None):
    if not HAS_SPELL:
        messagebox.showinfo("📦 Corretor indisponível",
            "O corretor ortográfico precisa da biblioteca 'pyspellchecker'.\n\n"
            "Instale com:\n    python -m pip install pyspellchecker\n\n"
            "(e, no .exe, recompile com --collect-all=spellchecker)")
        return
    self._v21_carregar_pessoal()
    if getattr(self, '_v21_ignorar_sessao', None) is None:
        self._v21_ignorar_sessao = set()

    st = {'jobs': self._v21_coletar(), 'idx': 0}
    if not st['jobs']:
        messagebox.showinfo("✅ Ortografia", "Nenhuma palavra não reconhecida nos campos de texto.")
        return
    if start_widget is not None:
        for i, j in enumerate(st['jobs']):
            if j[0] is start_widget:
                st['idx'] = i; break

    d = tk.Toplevel(self.root)
    d.title("📝 Corretor Ortográfico")
    d.transient(self.root); d.grab_set(); d.resizable(False, False)
    try: self.centralizar_janela(d)
    except Exception: pass

    f = ttk.Frame(d, padding=14); f.pack(fill='both', expand=True)
    lbl_campo = ttk.Label(f, text="", font=('DejaVu Sans', 9, 'bold'), foreground='#64748b')
    lbl_campo.pack(anchor='w')
    lbl_ctx = tk.Label(f, text="", font=('DejaVu Sans', 10), wraplength=440, justify='left',
                       bg=(self.tema_cores or {}).get('bg_card', '#fff'),
                       fg=(self.tema_cores or {}).get('texto', '#000'))
    lbl_ctx.pack(anchor='w', fill='x', pady=(2, 6))
    lbl_pal = tk.Label(f, text="", font=('DejaVu Sans', 11, 'bold'), fg='#dc2626')
    lbl_pal.pack(anchor='w')

    ttk.Label(f, text="Substituir por:").pack(anchor='w', pady=(8, 2))
    ent_corr = ttk.Entry(f, width=46); ent_corr.pack(anchor='w', fill='x')

    ttk.Label(f, text="Sugestões (clique para usar):").pack(anchor='w', pady=(8, 2))
    lf = ttk.Frame(f); lf.pack(fill='x')
    lb = tk.Listbox(lf, height=6, exportselection=False,
                    bg=(self.tema_cores or {}).get('campo_bg', '#fff'),
                    fg=(self.tema_cores or {}).get('texto', '#000'))
    lb.pack(side='left', fill='x', expand=True)
    sb = ttk.Scrollbar(lf, orient='vertical', command=lb.yview); sb.pack(side='right', fill='y')
    lb.configure(yscrollcommand=sb.set)
    lb.bind('<<ListboxSelect>>', lambda e: (ent_corr.delete(0, tk.END),
                                            ent_corr.insert(0, lb.get(lb.curselection()[0]))
                                            if lb.curselection() else None))

    lbl_status = ttk.Label(f, text="", font=('DejaVu Sans', 9), foreground='#64748b')
    lbl_status.pack(anchor='w', pady=(8, 4))

    def refresh():
        jobs = st['jobs']
        if not jobs:
            try: d.destroy()
            except Exception: pass
            messagebox.showinfo("✅ Ortografia", "Todas as palavras foram resolvidas.")
            return
        if st['idx'] >= len(jobs):
            st['idx'] = 0
        w, nome, pal = jobs[st['idx']]
        lbl_campo.config(text=f"Campo: {nome}    (ocorrência {st['idx'] + 1} de {len(jobs)})")
        lbl_ctx.config(text=w.get())
        lbl_pal.config(text=f"⚠ Palavra não reconhecida: {pal}")
        # sugestões
        lb.delete(0, tk.END)
        sug = []
        try:
            sug = self._v21_spell().candidates(pal) or []
            sug = sorted(sug, key=lambda s: (abs(len(s) - len(pal)), s))[:12]
        except Exception:
            sug = []
        for s in sug:
            lb.insert(tk.END, s)
        ent_corr.delete(0, tk.END)
        if sug:
            ent_corr.insert(0, sug[0])
        elif lb.size():
            lb.selection_set(0)
        lbl_status.config(text=f"Idioma do corretor: "
                               f"{dict(_V21_LANGS).get((self.config or {}).get('corretor_idioma','pt'),'pt')}")

    def rebuild_same_widget():
        cur_w = st['jobs'][st['idx']][0] if st['jobs'] else None
        st['jobs'] = self._v21_coletar()
        # tenta manter no mesmo campo
        if cur_w is not None:
            for i, j in enumerate(st['jobs']):
                if j[0] is cur_w:
                    st['idx'] = i; break
            else:
                st['idx'] = 0
        else:
            st['idx'] = 0
        # re-marcar cores dos campos afetados
        for nm in _V21_CAMPOS_TEXTO:
            ww = (self.entries or {}).get(nm)
            if ww is not None:
                self._v21_marcar_campo(ww)
        refresh()

    def substituir(todas):
        if not st['jobs']:
            return
        w, nome, pal = st['jobs'][st['idx']]
        corr = ent_corr.get()
        txt = w.get()
        try:
            novo = re.sub(r"\b" + re.escape(pal) + r"\b", lambda m: corr,
                          txt, count=(0 if todas else 1), flags=re.IGNORECASE)
        except Exception:
            novo = txt
        w.delete(0, tk.END); w.insert(0, novo)
        rebuild_same_widget()

    def ignorar():
        if not st['jobs']:
            return
        pal = st['jobs'][st['idx']][2]
        self._v21_ignorar_sessao.add(pal.lower())
        rebuild_same_widget()

    def adicionar_dic():
        if not st['jobs']:
            return
        pal = st['jobs'][st['idx']][2]
        self._v21_pessoal.add(pal.lower())
        self._v21_salvar_pessoal()
        rebuild_same_widget()

    def proxima():
        if not st['jobs']:
            return
        st['idx'] = (st['idx'] + 1) % len(st['jobs'])
        refresh()

    b1 = ttk.Frame(f); b1.pack(fill='x', pady=(4, 2))
    ttk.Button(b1, text="Substituir", command=lambda: substituir(False)).pack(side='left', padx=2, expand=True, fill='x')
    ttk.Button(b1, text="Substituir todas", command=lambda: substituir(True)).pack(side='left', padx=2, expand=True, fill='x')
    b2 = ttk.Frame(f); b2.pack(fill='x', pady=(0, 2))
    ttk.Button(b2, text="Ignorar palavra", command=ignorar).pack(side='left', padx=2, expand=True, fill='x')
    ttk.Button(b2, text="Adicionar ao dicionário", command=adicionar_dic).pack(side='left', padx=2, expand=True, fill='x')
    b3 = ttk.Frame(f); b3.pack(fill='x', pady=(2, 0))
    ttk.Button(b3, text="Próxima ▶", command=proxima).pack(side='left', padx=2, expand=True, fill='x')
    ttk.Button(b3, text="Fechar", command=d.destroy).pack(side='left', padx=2, expand=True, fill='x')

    # rodapé: troca de idioma do corretor
    bf = ttk.Frame(f); bf.pack(fill='x', pady=(10, 0))
    ttk.Label(bf, text="Idioma:").pack(side='left')
    lang_var = tk.StringVar(value=dict(_V21_LANGS).get((self.config or {}).get('corretor_idioma', 'pt'), 'Português'))
    cb = ttk.Combobox(bf, textvariable=lang_var, values=[r for _, r in _V21_LANGS],
                      state='readonly', width=12)
    cb.pack(side='left', padx=4)
    def _set_lang(e=None):
        cod = 'pt'
        for k, r in _V21_LANGS:
            if r == lang_var.get():
                cod = k; break
        try:
            self.config['corretor_idioma'] = cod; self._salvar_config()
        except Exception:
            pass
        self._v21_cache = {}             # força recarregar no novo idioma
        refresh()
    cb.bind('<<ComboboxSelected>>', _set_lang)

    refresh()
    try: ent_corr.focus_set()
    except Exception: pass

# ---- botão na barra + binds nos campos -----------------------------
def _v21_criar_botao(self):
    if getattr(self, '_v21_btn', None):
        return
    try:
        bf = self.btn_excluir.master
        btn = tk.Button(bf, text="✓ Ortografia", command=lambda: self._v21_abrir_corretor(None),
                        bg='#0e7490', fg='white', activebackground='#155e75',
                        activeforeground='white', font=('DejaVu Sans', 9, 'bold'),
                        relief='raised', padx=8, pady=4, cursor='hand2')
        btn.pack(side='left', padx=5)
        self._v21_btn = btn
    except Exception:
        pass

def _v21_instalar_binds(self):
    if getattr(self, '_v21_binds_done', False):
        return
    try:
        for nome in _V21_CAMPOS_TEXTO:
            w = (self.entries or {}).get(nome)
            if w is None:
                continue
            w.bind('<FocusOut>', lambda e, w=w: self._v21_marcar_campo(w), add='+')
            w.bind('<Double-Button-1>', lambda e, w=w: self._v21_abrir_corretor(w), add='+')
        self._v21_binds_done = True
    except Exception:
        pass

# ---- encadeia no create_widgets (roda depois de todos os patches) --
_V21_PREV_CW = LivroCatalogApp.create_widgets

def _v21_cw_wrap(self):
    _V21_PREV_CW(self)
    try: self._v21_criar_botao()
    except Exception: pass
    try: self._v21_instalar_binds()
    except Exception: pass

# reaplica a cor rosa correta ao trocar de tema
_V21_PREV_AT = LivroCatalogApp.aplicar_tema

def _v21_at_wrap(self):
    _V21_PREV_AT(self)
    try:
        for nome in _V21_CAMPOS_TEXTO:
            w = (self.entries or {}).get(nome)
            if w is not None:
                self._v21_marcar_campo(w)
    except Exception:
        pass

LivroCatalogApp.create_widgets      = _v21_cw_wrap
LivroCatalogApp.aplicar_tema        = _v21_at_wrap
LivroCatalogApp._v21_pessoal_path   = _v21_pessoal_path
LivroCatalogApp._v21_carregar_pessoal = _v21_carregar_pessoal
LivroCatalogApp._v21_salvar_pessoal = _v21_salvar_pessoal
LivroCatalogApp._v21_spell          = _v21_spell
LivroCatalogApp._v21_eh_suspeita    = _v21_eh_suspeita
LivroCatalogApp._v21_tokens         = _v21_tokens
LivroCatalogApp._v21_marcar_campo   = _v21_marcar_campo
LivroCatalogApp._v21_coletar        = _v21_coletar
LivroCatalogApp._v21_abrir_corretor = _v21_abrir_corretor
LivroCatalogApp._v21_criar_botao    = _v21_criar_botao
LivroCatalogApp._v21_instalar_binds = _v21_instalar_binds
# =====================================================================
# FIM PATCH v21  (a linha  if __name__ == "__main__":  vem logo abaixo)
# =====================================================================
# ==============================================================================
# EXECUÇÃO
# ==============================================================================
if __name__ == "__main__":
    root = None
    try:
        root = tk.Tk()
        app = LivroCatalogApp(root)
        try:
            if root.winfo_exists():
                root.mainloop()
        except tk.TclError:
            pass                       # janela fechada/cancelada: sai limpo
    except Exception as e:
        try:                           # grava o erro num arquivo (sem depender de console)
            with open(os.path.join(os.path.expanduser('~'), 'erro_catalogo.log'),
                      'a', encoding='utf-8') as f:
                f.write('\n===== ' + str(datetime.now()) + ' =====\n')
                f.write(traceback.format_exc())
        except Exception:
            pass
        if sys.stdout is not None:     # só imprime se houver console
            print("❌ Erro crítico:", e); traceback.print_exc()
        if sys.stdin is not None:      # só espera Enter se houver teclado
            input("\nPressione Enter para fechar...")